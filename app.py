import streamlit as st
import streamlit.components.v1 as components
import plotly.express as px
import plotly.graph_objects as go
import pandas as pd
import datetime
import pytz
import re
import time
import math
import requests
import calendar
import base64
import io
import unicodedata
import hashlib
import hmac
import secrets as secrets_lib
import smtplib
from email.mime.text import MIMEText
from weasyprint import HTML
import fitz


_PBKDF2_ITERATIONS = 200_000

def hacher_mot_de_passe(mot_de_passe: str, sel: bytes = None) -> str:
    """Retourne 'sel_hex$hash_hex' à stocker dans st.secrets."""
    if sel is None:
        sel = secrets_lib.token_bytes(16)
    h = hashlib.pbkdf2_hmac("sha256", mot_de_passe.encode("utf-8"), sel, _PBKDF2_ITERATIONS)
    return f"{sel.hex()}${h.hex()}"

def verifier_mot_de_passe(mot_de_passe_saisi: str, hash_stocke: str) -> bool:
    """Compare en temps constant le mot de passe saisi au hash stocké (format 'sel_hex$hash_hex')."""
    try:
        sel_hex, hash_hex = hash_stocke.split("$", 1)
        sel = bytes.fromhex(sel_hex)
        h_saisi = hashlib.pbkdf2_hmac("sha256", mot_de_passe_saisi.encode("utf-8"), sel, _PBKDF2_ITERATIONS)
        return hmac.compare_digest(h_saisi.hex(), hash_hex)
    except Exception:
        return False

# ==========================================
# SÉCURITÉ : LIMITATION DES TENTATIVES (ANTI-BRUTEFORCE)
# ==========================================
_MAX_TENTATIVES = 5
_DUREE_BLOCAGE_SECONDES = 60  # 5 minutes

def tentative_bloquee(cle: str) -> int:
    """Retourne le nb de secondes restant avant déblocage (0 si non bloqué)."""
    blocage_jusqu_a = st.session_state.get(f"blocage_{cle}", 0)
    reste = blocage_jusqu_a - time.time()
    return max(0, int(reste))

def enregistrer_echec(cle: str):
    nb = st.session_state.get(f"echecs_{cle}", 0) + 1
    st.session_state[f"echecs_{cle}"] = nb
    if nb >= _MAX_TENTATIVES:
        st.session_state[f"blocage_{cle}"] = time.time() + _DUREE_BLOCAGE_SECONDES
        st.session_state[f"echecs_{cle}"] = 0

def reinitialiser_echecs(cle: str):
    st.session_state[f"echecs_{cle}"] = 0
    st.session_state[f"blocage_{cle}"] = 0

# ==========================================
# ENVOI D'E-MAILS (SMTP) — utilisé pour la vérification Visiteur (code OTP)
# et pour les relances automatiques d'échéances
# ==========================================
def envoyer_email(destinataire: str, sujet: str, corps_texte: str) -> tuple:
    """Envoie un e-mail texte simple via SMTP. Retourne (succes: bool, erreur: str|None)."""
    try:
        cfg = st.secrets["smtp"]
    except Exception:
        return False, "Configuration SMTP absente (section [smtp] manquante dans secrets.toml)."
    try:
        msg = MIMEText(corps_texte, "plain", "utf-8")
        msg["Subject"] = sujet
        msg["From"] = cfg.get("expediteur", cfg.get("utilisateur"))
        msg["To"] = destinataire
        with smtplib.SMTP(cfg["host"], int(cfg.get("port", 587)), timeout=15) as serveur:
            serveur.starttls()
            serveur.login(cfg["utilisateur"], cfg["mot_de_passe"])
            serveur.sendmail(msg["From"], [destinataire], msg.as_string())
        return True, None
    except Exception as e:
        return False, str(e)

# ==========================================
# JOURNAL D'AUDIT (traçabilité des actions de modification)
# ==========================================
def journaliser_action(utilisateur: str, action: str, details: str = ""):
    try:
        now_str = datetime.datetime.now(TZ).strftime("%d/%m/%Y %H:%M:%S")
        sheets_append("AuditLog", [now_str, utilisateur or "inconnu", action, details])
    except Exception:
        pass  # la journalisation ne doit jamais bloquer l'action métier principale

def utilisateur_courant() -> str:
    """Identifie l'utilisateur actif (Admin / identifiant responsable / e-mail visiteur),
    pour le journal d'audit. Robuste même si appelé avant que `role` n'existe encore."""
    try:
        if role == "Admin" and password_correct:
            return "Admin"
        if role == "Responsable" and st.session_state.get("responsable_connecte"):
            return f"Responsable:{st.session_state.get('responsable_actif','?')}"
        if role == "Visiteur" and st.session_state.get("email_visiteur"):
            return f"Visiteur:{st.session_state.get('email_visiteur')}"
    except NameError:
        pass
    return "inconnu"

def afficher_apercu_pdf(pdf_bytes, hauteur=800, cle=None):
    """Aperçu PDF paginé : ne rend (PyMuPDF, 130 DPI) que la page actuellement affichée,
    au lieu de toutes les pages d'un coup — beaucoup plus léger pour les rapports longs."""
    try:
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        nb_pages = len(doc)
        if cle is None:
            cle = hashlib.md5(pdf_bytes[:4096]).hexdigest()[:10]
        etat_key = f"page_apercu_{cle}"
        if etat_key not in st.session_state:
            st.session_state[etat_key] = 0
        st.session_state[etat_key] = max(0, min(st.session_state[etat_key], nb_pages - 1))

        if nb_pages > 1:
            c_prev, c_mid, c_next = st.columns([1, 3, 1])
            with c_prev:
                if st.button("⬅️ Précédente", key=f"prev_{cle}", use_container_width=True, disabled=(st.session_state[etat_key] <= 0)):
                    st.session_state[etat_key] -= 1
                    st.rerun()
            with c_mid:
                st.markdown(f"<p style='text-align:center;margin-top:6px;color:#475569;font-weight:600;'>Page {st.session_state[etat_key] + 1} / {nb_pages}</p>", unsafe_allow_html=True)
            with c_next:
                if st.button("Suivante ➡️", key=f"next_{cle}", use_container_width=True, disabled=(st.session_state[etat_key] >= nb_pages - 1)):
                    st.session_state[etat_key] += 1
                    st.rerun()

        page = doc[st.session_state[etat_key]]
        pix = page.get_pixmap(dpi=130)
        img_bytes = pix.tobytes("png")
        st.image(img_bytes, use_container_width=True)
        if nb_pages > 1:
            st.caption(f"Page {st.session_state[etat_key] + 1} / {nb_pages}")
        doc.close()
    except Exception as e:
        st.error(f"Impossible d'afficher l'aperçu du PDF : {e}")
        st.info("Vous pouvez tout de même télécharger le rapport ci-dessous.")

def afficher_apercu_pdf_grille(pdf_bytes, colonnes=2, largeur_colonne=380, cle=None, taille_lot=6):
    """Aperçu PDF en grille, chargé par lots (au lieu de rendre toutes les pages en une fois) :
    seules les `taille_lot` premières pages sont rendues, avec un bouton pour en charger davantage."""
    try:
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        nb_pages = len(doc)
        if cle is None:
            cle = hashlib.md5(pdf_bytes[:4096]).hexdigest()[:10]
        etat_key = f"nb_pages_affichees_{cle}"
        if etat_key not in st.session_state:
            st.session_state[etat_key] = min(taille_lot, nb_pages)

        nb_a_afficher = min(st.session_state[etat_key], nb_pages)
        cols = st.columns(colonnes)
        for i in range(nb_a_afficher):
            page = doc[i]
            pix = page.get_pixmap(dpi=110)
            img_bytes = pix.tobytes("png")
            col = cols[i % colonnes]
            with col:
                st.image(img_bytes, width=largeur_colonne)
                st.caption(f"Page {i + 1} / {nb_pages}")

        if nb_a_afficher < nb_pages:
            if st.button(f"⬇️ Afficher {min(taille_lot, nb_pages - nb_a_afficher)} page(s) de plus ({nb_a_afficher}/{nb_pages})", key=f"plus_{cle}", use_container_width=True):
                st.session_state[etat_key] = min(st.session_state[etat_key] + taille_lot, nb_pages)
                st.rerun()
        doc.close()
    except Exception as e:
        st.error(f"Impossible d'afficher l'aperçu du PDF : {e}")
        st.info("Vous pouvez tout de même télécharger le rapport ci-dessous.")


def generer_rapport_equipements_pdf(df_exigences, site_filtre):
    installations = [
        "Installations électriques",
        "Equipements de levage",
        "Sécurité incendie",
        "Installations de gaz",
        "Appareil pression de gaz"
    ]
    
    df_eq = df_exigences[df_exigences.iloc[:, 0].astype(str).str.strip().str.lower() == "equipement"]
    
    df_eq = df_eq[df_eq.iloc[:, 1].astype(str).str.strip().str.upper() == site_filtre.upper()]
    logo_url = "https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcR6q1BtDSDgVnJZFo0hOBfQJoDS6OYiub-qfQ&s"
    
    html_content = f"""
    <html>
    <head>
    <style>
        @page {{
            size: A4 portrait;
            margin: 20mm 15mm;
            @bottom-right {{
                content: "Page " counter(page) " / " counter(pages);
                font-family: 'Helvetica Neue', Helvetica, Arial, sans-serif;
                font-size: 9pt;
                color: #64748B;
            }}
        }}
        body {{
            font-family: 'Helvetica Neue', Helvetica, Arial, sans-serif;
            color: #1E293B;
            margin: 0;
            padding: 0;
            font-size: 10pt;
        }}
        .page {{
            page-break-after: always;
        }}
        .page:last-child {{
            page-break-after: avoid;
        }}
        .page-header {{
            display: flex;
            align-items: center;
            gap: 12px;
            margin-bottom: 22px;
            padding-bottom: 10px;
            border-bottom: 1px solid #E2E8F0;
        }}
        .page-header img {{
            height: 30px;
        }}
        .page-header-text {{
            font-size: 9.5pt;
            color: #1E3A8A;
            font-weight: 700;
            text-transform: uppercase;
            letter-spacing: 0.4px;
        }}
        .header-title {{
            text-align: center;
            font-size: 17pt;
            font-weight: bold;
            color: #1E3A8A;
            margin-bottom: 22px;
            text-transform: uppercase;
            border-bottom: 2px solid #1E3A8A;
            padding-bottom: 14px;
            line-height: 1.35;
        }}
        .meta-info {{
            margin-bottom: 28px;
            background-color: #F8FAFC;
            border: 1px solid #E2E8F0;
            padding: 16px 18px;
            border-radius: 6px;
            line-height: 2.1;
            font-size: 11pt;
        }}
        .meta-info .dots {{
            color: #94A3B8;
            letter-spacing: 1px;
        }}
        .category-title {{
            font-size: 14pt;
            color: #0EA5E9;
            font-weight: bold;
            margin-top: 10px;
            margin-bottom: 15px;
            border-left: 4px solid #0EA5E9;
            padding-left: 8px;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            margin-bottom: 30px;
        }}
        th, td {{
            border: 1px solid #CBD5E1;
            padding: 10px;
            text-align: left;
        }}
        th {{
            background-color: #1E3A8A;
            color: white;
            font-weight: bold;
            text-transform: uppercase;
            font-size: 9pt;
        }}
        .col-sub {{ width: 60%; }}
        .col-nb {{ width: 20%; text-align: center; }}
        .col-chk {{ width: 20%; text-align: center; }}
        .td-center {{ text-align: center; }}

        .checkbox-box {{
            display: inline-block;
            width: 14px;
            height: 14px;
            border: 1px solid #475569;
            border-radius: 2px;
            margin-top: 3px;
        }}
        .signature-section {{
            margin-top: 40px;
            width: 100%;
            border-top: 1px dashed #CBD5E1;
            padding-top: 15px;
        }}
        .signature-title {{
            font-weight: bold;
            text-decoration: underline;
            margin-bottom: 60px;
        }}
        .empty-message {{
            color: #94A3B8;
            font-size: 11pt;
        }}
    </style>
    </head>
    <body>
    """

    def bloc_entete():
        return f"""
        <div class="page-header">
            <img src="{logo_url}"/>
            <span class="page-header-text">Tunisie Profilés d'Aluminium — Direction Maintenance &amp; TN</span>
        </div>
        <div class="header-title">Rapport d'inspection réglementaire — Site {site_filtre.upper()}</div>
        <div class="meta-info">
            <strong>Inspecteur technique :</strong> <span class="dots">…………………………………………………………………</span><br>
            <strong>Accompagnant :</strong> <span class="dots">……………………………………………………………………………</span><br>
            <strong>Date :</strong> <span class="dots">………………………………………………………………………………………</span>
        </div>
        """

    def bloc_signature():
        return """
        <div class="signature-section">
            <p class="signature-title">Signature :</p>
        </div>
        """

    if df_eq.empty:
        html_content += f"""
        <div class="page">
            {bloc_entete()}
            <p class="empty-message">Aucun équipement enregistré pour ce site.</p>
            {bloc_signature()}
        </div>
        """
    else:
        pages_generees = 0
        for installation in installations:
            df_ins = df_eq[df_eq.get("Installation", pd.Series(dtype=str)).astype(str).str.strip() == installation]
            if df_ins.empty:
                continue
            lignes_html = ""
            for _, row in df_ins.iterrows():
                sous_eq = row.get("Sous_equipement", "")
                try:
                    nombre = int(float(row.get("Nombre", 0) or 0))
                except (ValueError, TypeError):
                    nombre = row.get("Nombre", "")
                lignes_html += f"""
                <tr>
                    <td class="col-sub">{sous_eq}</td>
                    <td class="col-nb td-center">{nombre}</td>
                    <td class="col-chk td-center"><span class="checkbox-box"></span></td>
                </tr>"""

            html_content += f"""
            <div class="page">
                {bloc_entete()}
                <div class="category-title">{installation}</div>
                <table>
                    <thead>
                        <tr>
                            <th class="col-sub">Sous-équipements</th>
                            <th class="col-nb">Nombre</th>
                            <th class="col-chk">Case à cocher</th>
                        </tr>
                    </thead>
                    <tbody>{lignes_html}</tbody>
                </table>
                {bloc_signature()}
            </div>
            """
            pages_generees += 1

        if pages_generees == 0:
            html_content += f"""
            <div class="page">
                {bloc_entete()}
                <p class="empty-message">Aucun équipement enregistré pour ce site.</p>
                {bloc_signature()}
            </div>
            """

    html_content += """
    </body>
    </html>
    """

    return HTML(string=html_content).write_pdf()

# -------------------------------------------------------------------------------
# Calcul du tableau + génération PDF / Excel
# --------------------------------------------------------------------------------

def construire_calendrier_controle(df_rapports: pd.DataFrame, annee_reference: int = None) -> pd.DataFrame:
    if annee_reference is None:
        annee_reference = datetime.date.today().year
    if df_rapports.empty:
        return pd.DataFrame()

    col_ins       = [c for c in df_rapports.columns if "ins" in c.lower()]
    col_date      = [c for c in df_rapports.columns if "date" in c.lower()
                      and "reelle" not in c.lower() and "réelle" not in c.lower()
                      and "prochaine" not in c.lower() and "planifi" not in c.lower()]
    col_site      = [c for c in df_rapports.columns if "site" in c.lower()]
    col_reelle    = [c for c in df_rapports.columns if "reelle" in c.lower() or "réelle" in c.lower()]
    col_prochaine = [c for c in df_rapports.columns if "prochaine" in c.lower()]

    if not (col_ins and col_date and col_site):
        return pd.DataFrame()

    df = df_rapports.copy()
    df["_date_brute"]  = pd.to_datetime(df[col_date[0]], dayfirst=True, errors="coerce")
    df["_date_reelle"] = pd.to_datetime(df[col_reelle[0]], dayfirst=True, errors="coerce") if col_reelle else pd.NaT
    df["_date_prochaine_manuelle"] = (pd.to_datetime(df[col_prochaine[0]], dayfirst=True, errors="coerce")
                                       if col_prochaine else pd.NaT)
    df["_date"] = df["_date_reelle"].combine_first(df["_date_brute"])
    df = df.dropna(subset=["_date"])
    df["_site"] = df[col_site[0]].astype(str).str.strip().str.upper()
    df["_installation"] = df[col_ins[0]].astype(str).str.strip()

    col_realisation = f"Réalisation {annee_reference}"
    lignes = []

    for site in ("MEG", "SGB"):
        df_site = df[df["_site"] == site]
        for installation in ORDRE_CI:
            attendu = round(12 / PERIODICITE.get(installation, 12)) 
            df_grp = df_site[df_site["_installation"] == installation].sort_values("_date")

            dates_realisees = sorted(df_grp["_date"].dropna().unique())[-attendu:] if not df_grp.empty else []

            nb_realisees_annee = df_grp[df_grp["_date"].dt.year == annee_reference]["_date"].nunique()
            pct = round(min(nb_realisees_annee, attendu) / attendu * 100) if attendu else 0
            realisation_txt = f"{pct}%" if nb_realisees_annee > 0 else "A planifier"

            periodicite_mois = PERIODICITE.get(installation, 12)
            if not df_grp.empty and df_grp["_date_prochaine_manuelle"].notna().any():
                dates_planifiees = sorted(df_grp["_date_prochaine_manuelle"].dropna().unique())[-attendu:]
            else:
                dates_planifiees = [pd.Timestamp(d) + pd.DateOffset(months=periodicite_mois) for d in dates_realisees]

            lignes.append({
                "Site": site,
                "CI": CI_CODES.get(installation, ""),
                "Installation": installation,
                "Nbr visite/An": attendu,
                "Nbr jour": NB_JOURS_VISITE.get(site, {}).get(installation, "-"),
                "Dates réalisées": (" | ".join(pd.Timestamp(d).strftime("%d/%m/%Y") for d in dates_realisees)
                                     if len(dates_realisees) else "-"),
                col_realisation: realisation_txt,
                "Dates planifiées": "-",
                "Nbr visites réalisées en 2026": min(nb_realisees_annee, attendu), 
                "_dates_planifiees": list(dates_planifiees),
            })

    # Repère, sur l'ensemble du calendrier (tous sites/installations confondus), la SEULE
    # date planifiée la plus proche de la date du jour (= prochaine échéance à réaliser en priorité)
    # pour y accoler un petit symbole. Toutes les autres dates restent affichées sans symbole.
    aujourd_hui = pd.Timestamp(datetime.date.today())
    idx_ligne_proche, idx_date_proche, ecart_min = None, None, None
    for idx_ligne, ligne in enumerate(lignes):
        for idx_date, d in enumerate(ligne["_dates_planifiees"]):
            ecart = abs((pd.Timestamp(d) - aujourd_hui).days)
            if ecart_min is None or ecart < ecart_min:
                ecart_min, idx_ligne_proche, idx_date_proche = ecart, idx_ligne, idx_date

    for idx_ligne, ligne in enumerate(lignes):
        dates_planifiees = ligne.pop("_dates_planifiees")
        if len(dates_planifiees):
            ligne["Dates planifiées"] = " | ".join(
                (f"★ {pd.Timestamp(d).strftime('%d/%m/%Y')}"
                 if (idx_ligne == idx_ligne_proche and idx_date == idx_date_proche)
                 else pd.Timestamp(d).strftime("%d/%m/%Y"))
                for idx_date, d in enumerate(dates_planifiees)
            )
        else:
            ligne["Dates planifiées"] = "-"

    return pd.DataFrame(lignes)


# ==========================================
# NOTIFICATIONS / RELANCES D'ÉCHÉANCES
# ==========================================
# ⚠️ Streamlit n'exécute du code que lorsqu'une page est ouverte/interagie : il n'y a pas
# de tâche planifiée en arrière-plan ici. Le bouton de relance (onglet Statistiques) doit
# donc être déclenché manuellement par un administrateur, ou automatisé de l'extérieur
# (ex. une tâche planifiée GitHub Actions / cron qui appelle un petit script Python
# réutilisant ces mêmes fonctions, indépendamment de l'app Streamlit).
def extraire_echeances_proches(df_calendrier: pd.DataFrame, jours_horizon: int = 30) -> list:
    """Parcourt la colonne 'Dates planifiées' du calendrier et retourne la liste des échéances
    en retard ou à venir dans les `jours_horizon` prochains jours : [{Site, Installation, Date, Jours}]."""
    if df_calendrier.empty or "Dates planifiées" not in df_calendrier.columns:
        return []
    aujourd_hui = pd.Timestamp(datetime.date.today())
    resultats = []
    for _, row in df_calendrier.iterrows():
        brut = str(row.get("Dates planifiées", "") or "")
        if brut in ("", "-"):
            continue
        for morceau in brut.split("|"):
            morceau = morceau.replace("★", "").strip()
            try:
                d = pd.to_datetime(morceau, format="%d/%m/%Y")
            except Exception:
                continue
            jours = (d - aujourd_hui).days
            if jours <= jours_horizon:
                resultats.append({"Site": row.get("Site", ""), "Installation": row.get("Installation", ""),
                                   "Date": d.strftime("%d/%m/%Y"), "Jours": int(jours)})
    resultats.sort(key=lambda r: r["Jours"])
    return resultats

def construire_message_relance(echeances: list) -> str:
    if not echeances:
        return "Aucune échéance de contrôle réglementaire en retard ou à venir dans les 30 prochains jours."
    lignes = ["Echéances de contrôle réglementaire à surveiller :", ""]
    for e in echeances:
        if e["Jours"] < 0:
            statut = f"⚠️ EN RETARD de {abs(e['Jours'])} jour(s)"
        elif e["Jours"] == 0:
            statut = "📅 Aujourd'hui"
        else:
            statut = f"A venir dans {e['Jours']} jour(s)"
        lignes.append(f"[{e['Site']}] {e['Installation']} — prévue le {e['Date']} ({statut})")
    lignes.append("")
    lignes.append("Merci de vérifier la planification correspondante dans le Tableau de Bord Réglementaire.")
    return "\n".join(lignes)

def calculer_taux_realisation(df_calendrier: pd.DataFrame) -> dict:
    if df_calendrier.empty or "Nbr visites réalisées en 2026" not in df_calendrier.columns:
        return {"MEG": 0.0, "SGB": 0.0, "Global": 0.0}

    resultat = {}
    for site in ("MEG", "SGB"):
        df_site = df_calendrier[df_calendrier["Site"] == site]
        attendu_total = df_site["Nbr visite/An"].sum()
        realise_total = df_site["Nbr visites réalisées en 2026"].sum()
        resultat[site] = round(realise_total / attendu_total * 100, 1) if attendu_total else 0.0

    attendu_global = df_calendrier["Nbr visite/An"].sum()
    realise_global = df_calendrier["Nbr visites réalisées en 2026"].sum()
    resultat["Global"] = round(realise_global / attendu_global * 100, 1) if attendu_global else 0.0
    return resultat


def generer_calendrier_controle_pdf(df_calendrier: pd.DataFrame, annee_reference: int) -> bytes:
    """Génère le PDF du calendrier de contrôle (paysage, une ligne par installation, Site fusionné)."""
    import math

    col_realisation = f"Réalisation {annee_reference}"
    logo_url = "https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcR6q1BtDSDgVnJZFo0hOBfQJoDS6OYiub-qfQ&s"

    lignes_html = ""
    for site in df_calendrier["Site"].unique():
        df_site = df_calendrier[df_calendrier["Site"] == site].reset_index(drop=True)
        for i, row in df_site.iterrows():
            pct_val = row[col_realisation]
            if pct_val == "A planifier" or pd.isna(pct_val):
                couleur = "#94A3B8"
            else:
                try:
                    pct_num = int(round(float(str(pct_val).replace("%", "").strip() or 0)))
                except (ValueError, TypeError):
                    pct_num = 0
                couleur = "#16A34A" if pct_num >= 80 else "#F97316" if pct_num >= 50 else "#EF4444"

            lignes_html += "<tr>"
            if i == 0:
                lignes_html += f"<td rowspan='{len(df_site)}' class='site-cell'>{site}</td>"
            lignes_html += f"""
                <td class="ci-cell">{row['CI']}</td>
                <td>{row['Installation']}</td>
                <td class="center">{row['Nbr visite/An']}</td>
                <td class="center">{row['Nbr jour']}</td>
                <td class="center">{row['Dates réalisées']}</td>
                <td class="center" style="color:{couleur};font-weight:700;">{pct_val}</td>
                <td class="center planifiee">{row['Dates planifiées']}</td>
            </tr>"""

    # ----- Bloc graphiques : taux de réalisation par site (barres horizontales) + taux global (anneau) -----
    taux = calculer_taux_realisation(df_calendrier)
    taux_meg, taux_sgb, taux_global = taux.get("MEG", 0.0), taux.get("SGB", 0.0), taux.get("Global", 0.0)

    couleurs_site = {"MEG": "#2563EB", "SGB": "#059669"}
    barres_html = ""
    for site_lbl, val in (("MEG", taux_meg), ("SGB", taux_sgb)):
        barres_html += f"""
        <div class="bar-row">
            <div class="bar-label">{site_lbl}</div>
            <div class="bar-track">
                <div class="bar-fill" style="width:{val}%; background:{couleurs_site[site_lbl]};"></div>
            </div>
            <div class="bar-value">{val}%</div>
        </div>"""

    r_anneau = 68
    circonf = 2 * math.pi * r_anneau
    dash_val = circonf * (taux_global / 100)

    html_content = f"""
    <html><head><style>
        @page {{ size: A4 landscape; margin: 10mm 12mm; }}
        * {{ box-sizing: border-box; }}
        body {{ font-family: 'Helvetica Neue', Helvetica, Arial, sans-serif; color:#1E293B; font-size:9pt; }}
        .header {{ display:flex; align-items:center; gap:12px; margin-bottom:10px; border-bottom:2px solid #1E3A8A; padding-bottom:6px; }}
        .header img {{ height:28px; }}
        .header-texts {{ display:flex; flex-direction:column; }}
        .header-title {{ font-size:14pt; font-weight:800; color:#1E3A8A; text-transform:uppercase; }}
        .header-company {{ font-size:9pt; font-weight:700; color:#334155; }}
        .header-entite {{ font-size:8.3pt; font-weight:700; color:#0EA5E9; text-transform:uppercase; letter-spacing:0.3px; }}
        .subtitle {{ font-size:9pt; color:#64748B; margin:0 0 10px 0; }}
        table {{ width:100%; border-collapse:collapse; margin-bottom:14px; }}
        th, td {{ border:1px solid #CBD5E1; padding:8px 9px; text-align:left; }}
        th {{ background:#1E3A8A; color:white; font-size:8.5pt; text-transform:uppercase; }}
        .site-cell {{ font-weight:800; text-align:center; background:#F1F5F9; color:#1E3A8A; }}
        .ci-cell {{ font-weight:700; text-align:center; }}
        .center {{ text-align:center; }}
        .planifiee {{ font-weight:700; }}
        tr:nth-child(even) td:not(.site-cell) {{ background:#F8FAFC; }}

        .charts-title {{ font-size:11pt; font-weight:800; color:#1E3A8A; margin:10px 0 8px 0; text-align:center; }}
        .charts-wrap {{ display:flex; gap:16px; align-items:stretch; }}
        .charts-left {{ flex:1.5; background:#F8FAFC; border:1px solid #E2E8F0; border-radius:10px; padding:12px 16px; }}
        .charts-left-title {{ font-size:9.5pt; font-weight:700; color:#0F172A; margin-bottom:10px; text-align:center; }}
        .bar-row {{ display:flex; align-items:center; margin-bottom:12px; }}
        .bar-label {{ width:55px; font-weight:700; color:#0F172A; }}
        .bar-track {{ flex:1; height:18px; background:#E2E8F0; border-radius:5px; overflow:hidden; }}
        .bar-fill {{ height:100%; border-radius:5px; }}
        .bar-value {{ width:50px; text-align:right; font-weight:800; color:#0F172A; }}
        .charts-right {{ flex:1; background:#F8FAFC; border:1px solid #E2E8F0; border-radius:10px; padding:8px 16px; text-align:center; }}
        .charts-right-title {{ font-size:9.5pt; font-weight:700; color:#0F172A; margin-bottom:4px; }}
    </style></head>
    <body>
        <div class="header">
            <img src="{logo_url}"/>
            <div class="header-texts">
                <span class="header-title">Calendrier de contrôle réglementaire</span>
                <span class="header-company">Tunisie Profilés d'Aluminium</span>
                <span class="header-entite">DMTN - BT</span>
            </div>
        </div>
        <p class="subtitle">Année : {annee_reference}</p>
        <table>
            <thead><tr>
                <th>Sites</th><th>CI</th><th>Les installations</th><th>Nbr visite/An</th><th>Nbr jour</th>
                <th>Les visites réalisées</th><th>Réalisation {annee_reference}</th><th>Les visites planifiées</th>
            </tr></thead>
            <tbody>{lignes_html}</tbody>
        </table>

        <div class="charts-title">Taux de réalisation des contrôles — {annee_reference}</div>
        <div class="charts-wrap">
            <div class="charts-left">
                <div class="charts-left-title">Taux de réalisation par site</div>
                {barres_html}
            </div>
            <div class="charts-right">
                <div class="charts-right-title">Taux global {annee_reference}</div>
                <svg width="150" height="150" viewBox="0 0 180 180">
                    <circle cx="90" cy="90" r="{r_anneau}" fill="none" stroke="#E2E8F0" stroke-width="16"/>
                    <circle cx="90" cy="90" r="{r_anneau}" fill="none" stroke="#1E3A8A" stroke-width="16"
                            stroke-linecap="round"
                            stroke-dasharray="{dash_val:.1f} {circonf:.1f}"
                            transform="rotate(-90 90 90)"/>
                    <text x="90" y="104" text-anchor="middle" font-size="46" font-weight="800" fill="#0F172A">{taux_global}%</text>
                </svg>
            </div>
        </div>
    </body></html>
    """
    return HTML(string=html_content).write_pdf()


def generer_calendrier_controle_excel(df_calendrier: pd.DataFrame, annee_reference: int) -> bytes:
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, DoughnutChart, Reference
    from openpyxl.chart.marker import DataPoint
    from openpyxl.chart.label import DataLabelList

    df_export = df_calendrier[[c for c in df_calendrier.columns if not c.startswith("_")]]
    taux = calculer_taux_realisation(df_calendrier)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_export.to_excel(writer, index=False, sheet_name="Calendrier", startrow=1)
        ws = writer.sheets["Calendrier"]

        nb_cols = len(df_export.columns)
        ws.cell(row=1, column=1, value=f"Calendrier de contrôle réglementaire — Année {annee_reference}")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=nb_cols)
        ws.cell(row=1, column=1).font = Font(size=14, bold=True, color="1E3A8A")
        ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

        header_row = 2
        header_fill = PatternFill("solid", fgColor="1E3A8A")
        header_font = Font(color="FFFFFF", bold=True)
        thin = Side(style="thin", color="CBD5E1")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col_idx in range(1, nb_cols + 1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        for row_idx in range(header_row + 1, header_row + 1 + len(df_export)):
            for col_idx in range(1, nb_cols + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        nb_installations = len(ORDRE_CI)
        for i in range(0, len(df_export), nb_installations):
            r1 = header_row + 1 + i
            r2 = header_row + nb_installations + i
            ws.merge_cells(start_row=r1, start_column=1, end_row=r2, end_column=1)
            ws.cell(row=r1, column=1).font = Font(bold=True, color="1E3A8A")

        largeurs = [10, 6, 26, 12, 10, 26, 18, 26]
        for i, largeur in enumerate(largeurs, start=1):
            ws.column_dimensions[get_column_letter(i)].width = largeur

        # ================= FEUILLE "Taux de réalisation" (graphiques) =================
        ws_g = writer.book.create_sheet("Taux de réalisation")
        ws_g["A1"] = f"Taux de réalisation des contrôles — {annee_reference}"
        ws_g.merge_cells("A1:D1")
        ws_g["A1"].font = Font(size=14, bold=True, color="1E3A8A")
        ws_g["A1"].alignment = Alignment(horizontal="center")

        # --- Données pour la barre horizontale (par site) ---
        ws_g["A3"] = "Site"; ws_g["B3"] = "Taux (%)"
        ws_g["A3"].font = header_font; ws_g["A3"].fill = header_fill
        ws_g["B3"].font = header_font; ws_g["B3"].fill = header_fill
        ws_g["A4"] = "MEG"; ws_g["B4"] = taux.get("MEG", 0.0)
        ws_g["A5"] = "SGB"; ws_g["B5"] = taux.get("SGB", 0.0)

        bar = BarChart()
        bar.type = "bar"  # barres horizontales
        bar.title = f"Taux de réalisation par site — {annee_reference}"
        bar.y_axis.title = None
        bar.x_axis.title = "Taux (%)"
        bar.x_axis.scaling.min = 0
        bar.x_axis.scaling.max = 100
        data_bar = Reference(ws_g, min_col=2, min_row=3, max_row=5)
        cats_bar = Reference(ws_g, min_col=1, min_row=4, max_row=5)
        bar.add_data(data_bar, titles_from_data=True)
        bar.set_categories(cats_bar)
        serie_bar = bar.series[0]
        pt_meg = DataPoint(idx=0); pt_meg.graphicalProperties.solidFill = "2563EB"
        pt_sgb = DataPoint(idx=1); pt_sgb.graphicalProperties.solidFill = "059669"
        serie_bar.data_points = [pt_meg, pt_sgb]
        serie_bar.dLbls = DataLabelList()
        serie_bar.dLbls.showVal = True
        bar.height, bar.width = 8.5, 16
        bar.legend = None
        ws_g.add_chart(bar, "D3")

        # --- Données pour l'anneau (taux global) ---
        ws_g["A8"] = "Répartition"; ws_g["B8"] = "Valeur"
        ws_g["A8"].font = header_font; ws_g["A8"].fill = header_fill
        ws_g["B8"].font = header_font; ws_g["B8"].fill = header_fill
        ws_g["A9"] = "Réalisé"; ws_g["B9"] = taux.get("Global", 0.0)
        ws_g["A10"] = "Restant"; ws_g["B10"] = round(100 - taux.get("Global", 0.0), 1)

        doughnut = DoughnutChart()
        doughnut.title = f"Taux global {annee_reference} : {taux.get('Global', 0.0)}%"
        data_don = Reference(ws_g, min_col=2, min_row=9, max_row=10)
        cats_don = Reference(ws_g, min_col=1, min_row=9, max_row=10)
        doughnut.add_data(data_don, titles_from_data=False)
        doughnut.set_categories(cats_don)
        serie_don = doughnut.series[0]
        pt_real = DataPoint(idx=0); pt_real.graphicalProperties.solidFill = "1E3A8A"
        pt_reste = DataPoint(idx=1); pt_reste.graphicalProperties.solidFill = "E2E8F0"
        serie_don.data_points = [pt_real, pt_reste]
        doughnut.height, doughnut.width = 8.5, 10
        ws_g.add_chart(doughnut, "D21")

        ws_g.column_dimensions["A"].width = 14
        ws_g.column_dimensions["B"].width = 12

    output.seek(0)
    return output.getvalue()

    for ins in installations:
        # Filtrer par installation parmi les équipements du site
        df_ins = df_eq[df_eq.iloc[:, 2].astype(str).str.strip() == ins]
        
        html_content += f"""
        <div class="page">
            <div class="page-header">
                <img src="{logo_url}"/>
                <div class="page-header-text">Tunisie Profilés d'Aluminium — Direction Maintenance &amp; TN</div>
            </div>
            <div class="header-title" style="border-bottom: none; padding-bottom: 0;">Rapport d'Inspection Réglementaire</div>
            <div class="header-title">Site {site_filtre.upper()}</div>
            
            <div class="meta-info">
                <strong>Inspecteur technique :</strong> ............................................................<br>
                <strong>Accompagnant :</strong> ........................................................................<br>
                <strong>Date :</strong> .......................................................................................
            </div>
            
            <div class="category-title">{ins}</div>
            
            <table>
                <thead>
                    <tr>
                        <th class="col-sub">Sous-équipements</th>
                        <th class="col-nb">Nombre</th>
                        <th class="col-chk">Case à cocher</th>
                    </tr>
                </thead>
                <tbody>
        """
        
        if not df_ins.empty:
            for _, row in df_ins.iterrows():
                sous_eq = row.iloc[3] if pd.notna(row.iloc[3]) else "-"
                nombre = row.iloc[4] if pd.notna(row.iloc[4]) else "0"
                html_content += f"""
                    <tr>
                        <td>{sous_eq}</td>
                        <td class="td-center">{nombre}</td>
                        <td class="td-center"><span class="checkbox-box"></span></td>
                    </tr>
                """
        else:
            html_content += """
                <tr>
                    <td colspan="3" style="text-align:center; color:#94A3B8; font-style: italic;">Aucun équipement enregistré pour cette instalaltion sur ce site</td>
                </tr>
            """
            
        html_content += """
                </tbody>
            </table>
            
            <div class="signature-section">
                <div class="signature-title">Signature :</div>
            </div>
        </div>
        """
        
    html_content += """
    </body>
    </html>
    """
    
    return HTML(string=html_content).write_pdf()


def generer_export_global_excel() -> bytes:
    """Génère un classeur Excel unique regroupant toutes les données de l'application
    (rapports, planification, exigences, points de réserve, suivi des actions), un onglet
    par jeu de données — utile pour un export/archivage complet en un clic."""
    from openpyxl.styles import Font, PatternFill, Alignment

    def _ecrire_feuille(writer, df, nom_feuille):
        df_export = df.copy() if df is not None else pd.DataFrame()
        if df_export.empty:
            df_export = pd.DataFrame({"Info": ["Aucune donnée disponible"]})
        df_export.to_excel(writer, index=False, sheet_name=nom_feuille[:31])
        ws = writer.sheets[nom_feuille[:31]]
        header_fill = PatternFill("solid", fgColor="1E3A8A")
        header_font = Font(color="FFFFFF", bold=True)
        for col_idx in range(1, len(df_export.columns) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for col_idx, col_name in enumerate(df_export.columns, start=1):
            largeur = min(max(len(str(col_name)), df_export[col_name].astype(str).str.len().max() if len(df_export) else 0) + 4, 45)
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = largeur

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        _ecrire_feuille(writer, df_rapports, "Rapports")
        _ecrire_feuille(writer, df_planning, "Planning")
        _ecrire_feuille(writer, lire_exigences(), "Exigences")
        _ecrire_feuille(writer, lire_points_reserve(), "PointsReserve")
        _ecrire_feuille(writer, lire_points_reserve_nature(), "PointsReserveNature")
        _ecrire_feuille(writer, lire_suivi_encours(), "SuiviActions")
        _ecrire_feuille(writer, lire_actions_realisees(), "ActionsRealisees")
    output.seek(0)
    return output.getvalue()


def generer_rapport_kpi_pdf(kpi_data, df_reserve, df_nature, carto_b64, logo_url):
    """
    Génère un rapport PDF premium regroupant tous les KPI de l'onglet KPI :
    Taux de réalisation 2026, Taux de respect de délai,
    cartographie du taux de non-conformité, et répartition par site et par pilote.
    """
    date_str = datetime.date.today().strftime('%d/%m/%Y')

    def barre(pct, couleur):
        pct = max(0, min(100, pct))
        return f"""<div style="background:#E2E8F0;border-radius:6px;height:14px;width:100%;overflow:hidden;">
            <div style="background:{couleur};height:100%;width:{pct}%;"></div></div>"""

    k1 = kpi_data["kpi1"]; k2 = kpi_data["kpi2"]

    # ---- Palettes de couleurs (cohérentes avec le tableau de bord Streamlit) ----
    COULEURS_SITE = {"SGB": "#1E3A8A", "MEG": "#0EA5E9"}
    COULEURS_NATURE = {
        "Technique": "#10B981", "Sécurité": "#F97316", "Organisation": "#84CC16",
        "Règlementation": "#EAB308", "Documentation": "#EC4899", "Energétique": "#64748B",
    }
    COULEURS_PILOTE = {
        "Maintenance": "#FACC15", "HSE": "#F97316", "BT": "#EF4444",
        "Chef service BT": "#3B82F6", "DMTN": "#A855F7", "RH": "#92400E", "DG": "#22C55E",
    }

    # ---- Helpers SVG (donut chart et bar chart horizontal, sans dépendance externe) ----
    def _polar(cx, cy, r, angle_deg):
        a = math.radians(angle_deg - 90)
        return (cx + r * math.cos(a), cy + r * math.sin(a))

    def _donut_path(cx, cy, r_out, r_in, a0, a1):
        p0o = _polar(cx, cy, r_out, a0)
        p1o = _polar(cx, cy, r_out, a1)
        p1i = _polar(cx, cy, r_in, a1)
        p0i = _polar(cx, cy, r_in, a0)
        large = 1 if (a1 - a0) > 180 else 0
        return (f"M {p0o[0]:.2f} {p0o[1]:.2f} "
                f"A {r_out:.2f} {r_out:.2f} 0 {large} 1 {p1o[0]:.2f} {p1o[1]:.2f} "
                f"L {p1i[0]:.2f} {p1i[1]:.2f} "
                f"A {r_in:.2f} {r_in:.2f} 0 {large} 0 {p0i[0]:.2f} {p0i[1]:.2f} Z")

    def _donut_chart(data, color_map, titre="", size=190, show_pct_labels=True):
        """data: dict {label: valeur numérique}. Retourne (svg, legend_html).
        Les pourcentages des petites parts sont affichés à l'extérieur (avec un
        trait de rappel) pour rester lisibles ; les grandes parts gardent le
        pourcentage centré à l'intérieur de l'anneau.
        show_pct_labels=False : n'affiche aucun pourcentage sur le graphique lui-même
        (utile quand la légende à côté les affiche déjà, pour éviter la redondance)."""
        data = {k: v for k, v in data.items() if v and v > 0}
        total = sum(data.values())
        if not data or not total:
            return "", ""
        pad = 34 if show_pct_labels else 14  # marge latérale pour les étiquettes extérieures
        cx, cy = size / 2 + pad, size / 2 + 18
        r_out, r_in = size * 0.40, size * 0.40 * 0.58
        angle = 0.0
        slices, labels = "", ""
        for label, val in data.items():
            pct = val / total * 100
            a1 = angle + pct / 100 * 360
            color = color_map.get(label, "#94A3B8")
            slices += f'<path d="{_donut_path(cx,cy,r_out,r_in,angle,a1)}" fill="{color}" stroke="#ffffff" stroke-width="1.5"/>'
            if show_pct_labels:
                mid = (angle + a1) / 2
                if pct >= 6:
                    # part assez grande : pourcentage centré, en blanc, à l'intérieur de l'anneau
                    lx, ly = _polar(cx, cy, (r_out + r_in) / 2, mid)
                    labels += (f'<text x="{lx:.1f}" y="{ly:.1f}" font-size="11" font-weight="700" '
                               f'fill="#ffffff" text-anchor="middle" dominant-baseline="middle">{pct:.1f}%</text>')
                else:
                    # petite part : trait de rappel + pourcentage à l'extérieur, dans la couleur de la part
                    p0 = _polar(cx, cy, r_out, mid)
                    p1 = _polar(cx, cy, r_out + 12, mid)
                    anchor = "start" if p1[0] >= cx else "end"
                    tx = p1[0] + (4 if anchor == "start" else -4)
                    labels += (f'<line x1="{p0[0]:.1f}" y1="{p0[1]:.1f}" x2="{p1[0]:.1f}" y2="{p1[1]:.1f}" '
                               f'stroke="{color}" stroke-width="1.2"/>'
                               f'<text x="{tx:.1f}" y="{p1[1]:.1f}" font-size="9.5" font-weight="700" '
                               f'fill="{color}" text-anchor="{anchor}" dominant-baseline="middle">{pct:.1f}%</text>')
            angle = a1
        titre_svg = (f'<text x="{cx:.1f}" y="16" font-size="12.5" font-weight="700" fill="#0F172A" '
                     f'text-anchor="middle">{titre}</text>') if titre else ""
        w_total = size + 2 * pad
        svg = (f'<svg viewBox="0 0 {w_total} {size+22}" width="{w_total}" height="{size+22}" '
               f'xmlns="http://www.w3.org/2000/svg">{titre_svg}{slices}{labels}</svg>')
        legend = "".join(
            f'<div style="display:flex;align-items:center;margin-bottom:9px;">'
            f'<span style="width:11px;height:11px;min-width:11px;border-radius:3px;'
            f'background:{color_map.get(l,"#94A3B8")};display:inline-block;margin-right:8px;"></span>'
            f'<span style="font-size:10pt;color:#334155;white-space:nowrap;">{l}</span>'
            f'<span style="font-size:10pt;font-weight:800;color:#0F172A;white-space:nowrap;margin-left:14px;">{(v/total*100):.1f}%</span>'
            f'</div>'
            for l, v in data.items()
        )
        return svg, legend

    def _bar_list_html(data_pct, color_map):
        """data_pct: dict {label: pourcentage}, déjà trié décroissant.
        Présentation HTML (et non SVG) : libellé complet + barre + pourcentage en gras
        bien séparés, pour rester lisibles quel que soit le nombre de pilotes ou la
        longueur de leur nom (ex. « Chef service BT »)."""
        if not data_pct:
            return ""
        rows = ""
        for label, pct in data_pct.items():
            color = color_map.get(label, "#F59E0B")
            rows += f"""
            <div style="display:flex;align-items:center;gap:12px;margin-bottom:11px;">
                <div style="flex:0 0 150px;font-size:9.5pt;font-weight:600;color:#334155;
                    white-space:nowrap;overflow:hidden;text-overflow:ellipsis;">{label}</div>
                <div style="flex:1;background:#E2E8F0;border-radius:5px;height:16px;overflow:hidden;">
                    <div style="height:100%;width:{pct}%;background:{color};border-radius:5px;"></div>
                </div>
                <div style="flex:0 0 58px;text-align:right;font-size:10.5pt;font-weight:800;color:#0F172A;">{pct:.1f}%</div>
            </div>"""
        return rows

    # ---- Section 1 : Actions de contrôle — par site et par installation (source : PointsReserve) ----
    df_r = df_reserve.copy() if (df_reserve is not None and not df_reserve.empty) else pd.DataFrame()
    if not df_r.empty and "Nombre" in df_r.columns:
        df_r["Nombre"] = pd.to_numeric(df_r["Nombre"], errors="coerce").fillna(0)

    site_donut_svg, site_donut_legend = "", ""
    if not df_r.empty and "Site" in df_r.columns:
        site_donut_svg, site_donut_legend = _donut_chart(
            df_r.groupby("Site")["Nombre"].sum().to_dict(), COULEURS_SITE, "Répartition par site", size=210)

    def _ins_donut(site):
        if df_r.empty or "Installation" not in df_r.columns or "Site" not in df_r.columns:
            return "", ""
        d = df_r[df_r["Site"] == site].groupby("Installation")["Nombre"].sum().to_dict()
        return _donut_chart(d, COULEURS_INS, site, size=185, show_pct_labels=False)

    meg_ins_svg, meg_ins_legend = _ins_donut("MEG")
    sgb_ins_svg, sgb_ins_legend = _ins_donut("SGB")

    # ---- Section 2 : Répartition par site — Nature et Pilote (source : PointsReserveNature) ----
    df_n = df_nature.copy() if (df_nature is not None and not df_nature.empty) else pd.DataFrame()
    if not df_n.empty and "Nombre" in df_n.columns:
        df_n["Nombre"] = pd.to_numeric(df_n["Nombre"], errors="coerce").fillna(0)

    def _nature_donut(site):
        if df_n.empty or "Nature" not in df_n.columns or "Site" not in df_n.columns:
            return "", ""
        d = df_n[df_n["Site"] == site].groupby("Nature")["Nombre"].sum().to_dict()
        return _donut_chart(d, COULEURS_NATURE, f"{site} — % par nature", size=185, show_pct_labels=False)

    def _pilote_bar(site):
        if df_n.empty or "Pilote" not in df_n.columns or "Site" not in df_n.columns:
            return ""
        sub = df_n[df_n["Site"] == site]
        total = sub["Nombre"].sum()
        compte = {}
        for _, row in sub.iterrows():
            for e in str(row.get("Pilote", "")).split("+"):
                e = e.strip()
                if not e:
                    continue
                compte[e] = compte.get(e, 0) + row["Nombre"]
        if not compte or not total:
            return ""
        pct_dict = {k: round(v / total * 100, 1) for k, v in sorted(compte.items(), key=lambda x: -x[1])}
        return _bar_list_html(pct_dict, COULEURS_PILOTE)

    sgb_nature_svg, sgb_nature_legend = _nature_donut("SGB")
    meg_nature_svg, meg_nature_legend = _nature_donut("MEG")
    sgb_pilote_svg = _pilote_bar("SGB")
    meg_pilote_svg = _pilote_bar("MEG")


    carto_html = ""
    if carto_b64:
        carto_html = f"""
        <div class="page">
            <div class="category-title">Taux de non-conformité des sites</div>
            <p style="font-size:10pt;color:#475569;margin-bottom:12px;">
            Cartographie de synthèse du taux de non-conformité par site et par installation,
            établie lors de la campagne de contrôle réglementaire 2026.</p>
            <div style="text-align:center;">
                <img src="data:image/png;base64,{carto_b64}" style="max-width:100%;max-height:150mm;width:auto;
                    border-radius:8px;border:1px solid #E2E8F0;"/>
            </div>
        </div>"""

    html_content = f"""
    <html><head><style>
        @page {{ size: A4 landscape; margin: 15mm 18mm;
            @bottom-right {{ content: "Page " counter(page) " / " counter(pages);
                font-family:'Helvetica Neue',Helvetica,Arial,sans-serif; font-size:9pt; color:#64748B; }} }}
        body {{ font-family:'Helvetica Neue',Helvetica,Arial,sans-serif; color:#1E293B; margin:0; padding:0; font-size:10pt; }}
        .page {{ page-break-after: always; }}
        .page:last-child {{ page-break-after: avoid; }}
        .logo-box {{ text-align:center; margin-bottom:8px; }}
        .logo-box img {{ height:58px; }}
        .header-title {{ text-align:center; font-size:20pt; font-weight:bold; color:#1E3A8A; margin:6px 0 4px 0;
            text-transform:uppercase; letter-spacing:0.5px; }}
        .header-sub {{ text-align:center; font-size:11pt; color:#64748B; margin-bottom:22px; }}
        .meta-info {{ background:#F8FAFC; border:1px solid #E2E8F0; border-radius:6px; padding:14px 16px;
            font-size:10pt; line-height:1.7; margin-bottom:25px; }}
        .category-title {{ font-size:15pt; color:#0EA5E9; font-weight:bold; border-left:4px solid #0EA5E9;
            padding-left:10px; margin:10px 0 16px 0; }}
        .kpi-card {{ background:#F8FAFC; border:1px solid #E2E8F0; border-left:5px solid #1E3A8A; border-radius:8px;
            padding:18px; margin-bottom:18px; }}
        .kpi-title {{ font-size:13pt; font-weight:700; color:#0F172A; margin:0 0 6px 0; }}
        .kpi-desc {{ font-size:9.5pt; color:#475569; margin:0 0 12px 0; line-height:1.5; }}
        .kpi-value {{ font-size:26pt; font-weight:800; color:#1E3A8A; margin:0 0 10px 0; }}
        table {{ width:100%; border-collapse:collapse; margin-bottom:20px; }}
        th, td {{ border:1px solid #CBD5E1; padding:8px 10px; text-align:left; font-size:9pt; }}
        th {{ background:#1E3A8A; color:white; text-transform:uppercase; font-size:8.5pt; font-weight:bold; }}
    </style></head><body>

    <div class="page">
        <div class="logo-box"><img src="{logo_url}"/></div>
        <div class="header-title">Rapport KPI</div>
        <div class="header-title">Contrôle Réglementaire</div>
        <div class="header-sub">Tunisie Profilés d'Aluminium — Direction Maintenance &amp; TN</div>
        <div class="meta-info">
            <b>Date d'édition :</b> {date_str}<br>
            <b>Objet :</b> Synthèse des indicateurs de performance du suivi de conformité réglementaire —
            taux de réalisation, respect des délais, non-conformités et actions de contrôle.
        </div>

        <div class="category-title">Indicateurs de performance</div>

       





        <div style="display:flex;gap:20px;align-items:stretch;">
        <div class="kpi-card" style="border-left-color:#0EA5E9;flex:1;">
            <p class="kpi-title">1. Taux de réalisation 2026</p>
            <p class="kpi-desc">Proportion des visites réalisées dont l'écart entre la date réelle de contrôle
            et l'échéance théorique initiale du cycle n'excède pas 1 mois, par rapport au nombre total
            de visites réalisées.</p>
            <p class="kpi-value">{k1['taux']}%</p>
            {barre(k1['taux'], '#10B981')}
            <p style="font-size:9pt;color:#64748B;margin-top:8px;">{k1['realises']} réalisés / {k1['restants']} non réalisés
            — sur {k1['total']} visites planifiées</p>
        </div>

         <div class="kpi-card" style="flex:1;">
            <p class="kpi-title">2. Taux de respect de délai de visite</p>
            <p class="kpi-desc">Proportion des contrôles réglementaires dont l'échéance théorique est comprise
            entre le 01/01/2026 et le 31/12/2026, effectivement réalisés (date réelle de visite enregistrée)
            par rapport au nombre total de contrôles dus sur cette période.</p>
            <p class="kpi-value">{k2['taux']}%</p>
            {barre(k2['taux'], '#0EA5E9')}
            <p style="font-size:9pt;color:#64748B;margin-top:8px;">{k2['respectes']} respectés / {k2['respectes']} réalisés</p>
        </div>
        </div>

    </div>

    {carto_html}

    <div class="page">
        <div class="category-title">Répartition par site et par installation</div>
        <p style="font-size:10pt;color:#475569;margin-bottom:15px;">
        Répartition des actions de contrôle relevées, par site et par installation.</p>

        <div style="display:flex;justify-content:center;gap:30px;align-items:center;margin-bottom:10px;">
            <div>
                {site_donut_svg if site_donut_svg else "<p style='color:#94A3B8;font-size:9pt;'>Aucune donnée</p>"}
            </div>
            <div>{site_donut_legend}</div>
        </div>

        <p style="font-weight:700;font-size:12pt;color:#0F172A;text-align:center;margin:20px 0 12px 0;">
        Répartition par installation</p>

        <div style="display:flex;gap:40px;">
            <div style="flex:1;">
                <p style="font-weight:700;font-size:11pt;color:#0F172A;text-align:center;margin:0 0 10px 0;">MEG</p>
                <div style="display:flex;justify-content:center;align-items:center;gap:20px;">
                    <div style="flex:0 0 auto;">
                        {meg_ins_svg if meg_ins_svg else "<p style='color:#94A3B8;font-size:9pt;'>Aucune donnée MEG</p>"}
                    </div>
                    <div style="flex:0 0 180px;">{meg_ins_legend}</div>
                </div>
            </div>
            <div style="flex:1;">
                <p style="font-weight:700;font-size:11pt;color:#0F172A;text-align:center;margin:0 0 10px 0;">SGB</p>
                <div style="display:flex;justify-content:center;align-items:center;gap:20px;">
                    <div style="flex:0 0 auto;">
                        {sgb_ins_svg if sgb_ins_svg else "<p style='color:#94A3B8;font-size:9pt;'>Aucune donnée SGB</p>"}
                    </div>
                    <div style="flex:0 0 180px;">{sgb_ins_legend}</div>
                </div>
            </div>
        </div>
    </div>

    <div class="page">
        <div class="category-title">Répartition par site : Nature et Pilote</div>
        <p style="font-size:10pt;color:#475569;margin-bottom:15px;">
        Répartition des actions de contrôle relevées, par nature et par pilote, pour chaque site.</p>

        <p style="font-weight:700;font-size:12pt;color:#0F172A;margin:10px 0 12px 0;">SGB</p>
        <div style="display:flex;align-items:center;gap:25px;margin-bottom:25px;">
            <div style="flex:0 0 auto;">
                {sgb_nature_svg if sgb_nature_svg else "<p style='color:#94A3B8;font-size:9pt;'>Aucune donnée</p>"}
            </div>
            <div style="flex:0 0 150px;">{sgb_nature_legend}</div>
            <div style="flex:1;">
                <p style="font-size:9.5pt;font-weight:700;color:#64748B;margin:0 0 10px 0;text-transform:uppercase;">% par pilote</p>
                {sgb_pilote_svg if sgb_pilote_svg else "<p style='color:#94A3B8;font-size:9pt;'>Aucune donnée</p>"}
            </div>
        </div>

        <p style="font-weight:700;font-size:12pt;color:#0F172A;margin:10px 0 12px 0;">MEG</p>
        <div style="display:flex;align-items:center;gap:25px;">
            <div style="flex:0 0 auto;">
                {meg_nature_svg if meg_nature_svg else "<p style='color:#94A3B8;font-size:9pt;'>Aucune donnée</p>"}
            </div>
            <div style="flex:0 0 150px;">{meg_nature_legend}</div>
            <div style="flex:1;">
                <p style="font-size:9.5pt;font-weight:700;color:#64748B;margin:0 0 10px 0;text-transform:uppercase;">% par pilote</p>
                {meg_pilote_svg if meg_pilote_svg else "<p style='color:#94A3B8;font-size:9pt;'>Aucune donnée</p>"}
            </div>
        </div>
    </div>

    </body></html>
    """


    return HTML(string=html_content).write_pdf()


def generer_rapport_pilote_pdf(pilote_choisi, df_filtre, logo_url):
    """
    Génère un rapport PDF (format paysage) listant, pour un pilote donné, toutes les actions de
    la codification (classeur externe) qui le concernent — une page par installation
    (= un onglet du classeur source), sous forme de fiche de suivi terrain :
    Equipement | Actions | Responsable | Etat (Immédiat/Sous-traitant*/Planifié*) | Réalisation (O/N) | Observation.
    df_filtre doit contenir les colonnes : Installation, Désignation, Observation, Code, Nature.
    """
    date_str = datetime.date.today().strftime('%d/%m/%Y')
    installations = list(dict.fromkeys(df_filtre["Installation"].tolist()))  # ordre stable, sans doublons
    total_general = len(df_filtre)

    # Détermine le(s) site(s) présent(s) dans le périmètre filtré, afin d'afficher le bon
    # sous-pilote (ex : Maintenance -> Chafik ABID pour SGB, Saber BEN CHAABEN pour MEG).
    sites_presents = []
    if "Site" in df_filtre.columns:
        sites_presents = sorted({
            str(s).strip().upper() for s in df_filtre["Site"].dropna().unique().tolist() if str(s).strip()
        })

    def _nom_pour_site(site_val):
        return nom_pour_pilote_site(pilote_choisi, site_val)

    if len(sites_presents) == 1:
        nom_responsable = _nom_pour_site(sites_presents[0])
    else:
        # Plusieurs sites (ou site inconnu) dans le même rapport : on garde le libellé générique.
        nom_responsable = _nom_pour_site(None)

    # Nom du sous-pilote par installation (une installation appartient toujours à un seul site).
    nom_par_installation = {}
    for ins in installations:
        site_ins = None
        if "Site" in df_filtre.columns:
            valeurs_site = df_filtre.loc[df_filtre["Installation"] == ins, "Site"].dropna().unique().tolist()
            if len(valeurs_site) == 1:
                site_ins = valeurs_site[0]
        nom_par_installation[ins] = _nom_pour_site(site_ins)

    html_content = f"""
    <html>
    <head>
    <style>
        @page {{
            size: A4 landscape;
            margin: 15mm 12mm;
            @bottom-right {{
                content: "Page " counter(page) " / " counter(pages);
                font-family: 'Helvetica Neue', Helvetica, Arial, sans-serif;
                font-size: 9pt;
                color: #64748B;
            }}
        }}
        body {{
            font-family: 'Helvetica Neue', Helvetica, Arial, sans-serif;
            color: #1E293B;
            margin: 0;
            padding: 0;
            font-size: 10pt;
        }}
        .page {{ page-break-after: always; }}
        .page:last-child {{ page-break-after: avoid; }}
        .header-title {{
            text-align: center;
            font-size: 16pt;
            font-weight: bold;
            color: #1E3A8A;
            margin-bottom: 14px;
            text-transform: uppercase;
            border-bottom: 2px solid #1E3A8A;
            padding-bottom: 8px;
        }}
        .page-header {{
            display: flex;
            align-items: center;
            gap: 12px;
            margin-bottom: 12px;
            padding-bottom: 8px;
            border-bottom: 1px solid #E2E8F0;
        }}
        .page-header img {{ height: 32px; }}
        .page-header-text {{
            font-size: 9.5pt;
            color: #64748B;
            font-weight: 600;
            text-transform: uppercase;
            letter-spacing: 0.4px;
        }}
        .meta-info {{
            margin-bottom: 14px;
            background-color: #F8FAFC;
            border: 1px solid #E2E8F0;
            padding: 10px 15px;
            border-radius: 6px;
            line-height: 1.7;
            font-size: 10.5pt;
        }}
        .category-title {{
            font-size: 13pt;
            color: #0EA5E9;
            font-weight: bold;
            margin-top: 6px;
            margin-bottom: 10px;
            border-left: 4px solid #0EA5E9;
            padding-left: 8px;
        }}
        table {{ width: 100%; border-collapse: collapse; margin-bottom: 14px; }}
        th, td {{ border: 1px solid #CBD5E1; padding: 6px; text-align: left; font-size: 8.5pt; vertical-align: middle; }}
        th {{
            background-color: #1406BE; color: #FFFFFF; font-weight: bold;
            text-align: center; font-size: 8.5pt;
        }}
        .col-equip  {{ width: 13%; }}
        .col-action {{ width: 22%; }}
        .col-resp   {{ width: 10%; text-align: center; }}
        .col-etat   {{ width: 7%; text-align: center; }}
        .col-real   {{ width: 9%; text-align: center; }}
        .col-obs    {{ width: 25%; }}
        .td-chk {{ text-align: center; }}
        .checkbox-box {{
            display: inline-block;
            width: 13px;
            height: 13px;
            border: 1.5px solid #1E293B;
            border-radius: 3px;
        }}
        .footnote {{
            font-size: 8pt;
            color: #475569;
            margin-top: 4px;
        }}
        .total-badge {{
            display: inline-block; background:#0EA5E9; color:white; font-weight:700;
            padding:3px 12px; border-radius:12px; font-size:9pt; margin-left:8px;
        }}
        .table-synthese th {{
            background-color: #1406BE; color: #FFFFFF; font-weight: bold;
            text-align: center; font-size: 9pt;
        }}
        .table-synthese td {{
            font-size: 9.5pt;
        }}
        .table-synthese .ligne-total td {{
            font-weight: bold;
            background-color: #F1F5F9;
        }}
        .synthese-cadre {{
            border: 1.5px solid #1E3A8A;
            border-radius: 6px;
            padding: 12px 15px;
            margin-top: 10px;
            min-height: 260px;
        }}
        .synthese-titre {{
            font-size: 11pt;
            font-weight: bold;
            color: #1E3A8A;
            margin-bottom: 10px;
        }}
    </style>
    </head>
    <body>
    """

    for ins in installations:
        d_ins = df_filtre[df_filtre["Installation"] == ins]
        html_content += f"""
        <div class="page">
            <div class="page-header">
                <img src="{logo_url}"/>
                <div class="page-header-text">Tunisie Profilés d'Aluminium — Direction Maintenance &amp; TN</div>
            </div>
            <div class="header-title" style="border-bottom: none; padding-bottom: 0;">Plan d'actions - Contrôle réglementaire</div>
            <div class="meta-info">
                <strong>Sous-pilote :</strong> {nom_par_installation.get(ins, nom_responsable)}<br>
                <strong>Installation :</strong> {ins}<br>
                <strong>Date d'édition :</strong> {date_str}
            </div>
            <div class="category-title">{ins} <span class="total-badge">{len(d_ins)} action(s)</span></div>
            <table>
                <thead>
                    <tr>
                        <th class="col-equip" rowspan="2">Equipement</th>
                        <th class="col-action" rowspan="2">Actions</th>
                        <th class="col-resp" rowspan="2">Responsable</th>
                        <th colspan="3">Etat de suivi</th>
                        <th class="col-real" rowspan="2">Réalisation<br>(O/N)</th>
                        <th class="col-obs" rowspan="2">Suivi d'avancement *</th>
                    </tr>
                    <tr>
                        <th class="col-etat">Immédiat</th>
                        <th class="col-etat">Sous-traitant*</th>
                        <th class="col-etat">Planifié*</th>
                    </tr>
                </thead>
                <tbody>
        """
        if not d_ins.empty:
            for equip, span, observation in _lignes_avec_rowspan(d_ins):
                html_content += "<tr>"
                if span is not None:
                    html_content += f'<td rowspan="{span}">{equip}</td>'
                html_content += f"""
                        <td>{observation}</td>
                        <td class="col-resp"></td>
                        <td class="td-chk"><span class="checkbox-box"></span></td>
                        <td class="td-chk"><span class="checkbox-box"></span></td>
                        <td class="td-chk"><span class="checkbox-box"></span></td>
                        <td></td>
                        <td></td>
                    </tr>
                """
        else:
            html_content += """
                <tr><td colspan="8" style="text-align:center;color:#94A3B8;font-style:italic;">Aucune action</td></tr>
            """
        html_content += """
                </tbody>
            </table>
            <div class="footnote">(*) Suivi d'avancement : Date de réalisation, Besoin PDR, Lancement DA, Nom de sous-traitant...</div>
        </div>
        """

    footnote_synthese = ('<div class="footnote">(*) Suivi d\'avancement : Date de réalisation, Besoin PDR, '
                          'Lancement DA, Nom de sous-traitant...</div>')

    lignes_synthese_tableau = ""
    for ins in installations:
        nb_ins = len(df_filtre[df_filtre["Installation"] == ins])
        lignes_synthese_tableau += f"""
                    <tr>
                        <td>{ins}</td>
                        <td style="text-align:center;">{nb_ins}</td>
                        <td style="text-align:center;"></td>
                    </tr>
        """

    html_content += f"""
    <div class="page">
        <div class="page-header">
            <img src="{logo_url}"/>
            <div class="page-header-text">Tunisie Profilés d'Aluminium — Direction Maintenance &amp; TN</div>
        </div>
        <div class="header-title">Plan d'actions - Contrôle réglementaire</div>
        <div class="meta-info">
            <strong>Sous-pilote :</strong> {nom_responsable}<br>
            <strong>Total des actions :</strong> {total_general} action(s)<br>
            <strong>Date d'édition :</strong> {date_str}
        </div>

        <table class="table-synthese">
            <thead>
                <tr>
                    <th style="width:50%;">Installation</th>
                    <th style="width:25%;">Nombre d'actions</th>
                    <th style="width:25%;">Taux de réalisation</th>
                </tr>
            </thead>
            <tbody>
                {lignes_synthese_tableau}
                <tr class="ligne-total">
                    <td>Total</td>
                    <td style="text-align:center;">{total_general}</td>
                    <td style="text-align:center;"></td>
                </tr>
            </tbody>
        </table>

        <div class="synthese-cadre">
            <div class="synthese-titre">Synthèse / Observations et remarques</div>
        </div>
    </div>
    </body>
    </html>
    """

    return HTML(string=html_content).write_pdf()


st.set_page_config(
    page_title="Contrôle Réglementaire",
    page_icon="🛡️",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ==========================================
# INITIALISATION
# ==========================================
if "email_visiteur"   not in st.session_state: st.session_state.email_visiteur   = None
if "heartbeat_actif"  not in st.session_state: st.session_state.heartbeat_actif  = False
if "cal_mois"         not in st.session_state: st.session_state.cal_mois         = datetime.date.today().month
if "cal_annee"        not in st.session_state: st.session_state.cal_annee        = datetime.date.today().year
if "jour_selectionne" not in st.session_state: st.session_state.jour_selectionne = None

tab3 = None
TZ       = pytz.timezone('Africa/Tunis')
SHEET_ID = "1ZK6VWg_gcCO70nt6DTyYogDeNeQUgovFmwWQufMVO-M"
URL_GOOGLE_SHEET = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/edit?gid=0#gid=0"
# Classeurs externes "Classification des actions CR 2026" : un classeur par site, un onglet par
# installation à l'intérieur, colonnes Désignation | Observation | Code (T/S/E/D/O/R).
# Utilisés pour le rapport PDF par pilote ainsi que pour le suivi des actions.
CODIF_SHEET_ID_PAR_SITE = {
    "MEG": "1AF65P1sQPKM6JN7_r2mck-UrrZqz7tn5QmpnJD1iICA",
    "SGB": "1bD6LUxs_nGgamVsC9DAmGffScUo5wgMhxjvPBuBxzx0",
}
# Conservé par compatibilité (ancien classeur unique) — non utilisé si les IDs par site ci-dessus sont renseignés.
CODIF_SHEET_ID = "119hyynlCiIUzf-17iiSkcPnaEr2oCiFC"
SEUIL_EN_LIGNE_SECONDES = 90
calendar.setfirstweekday(0)

PERIODICITE = {
    "Installations électriques": 6,
    "Equipements de levage":     12,
    "Sécurité incendie":         12,
    "Installations de gaz":      12,
    "Appareil pression de gaz":  12,
}
COULEURS_INS = {
    "Installations électriques": "#2a78d6",
    "Equipements de levage":     "#1baf7a",
    "Sécurité incendie":         "#e34948",
    "Installations de gaz":      "#eda100",
    "Appareil pression de gaz":  "#4a3aa7",
}
# Couleur d'accent par site, pour une identification visuelle rapide dans les badges,
# titres et graphes comparatifs (MEG = bleu, SGB = vert).
COULEUR_SITE = {
    "MEG": {"principale": "#2563EB", "claire": "#DBEAFE", "fonce": "#1E3A8A"},
    "SGB": {"principale": "#059669", "claire": "#D1FAE5", "fonce": "#065F46"},
}
def badge_site(site: str) -> str:
    """Retourne un petit badge HTML coloré selon le site (MEG bleu / SGB vert)."""
    c = COULEUR_SITE.get(str(site).upper(), {"principale": "#64748B", "claire": "#F1F5F9", "fonce": "#334155"})
    return (f'<span style="background:{c["claire"]};color:{c["fonce"]};padding:2px 10px;'
            f'border-radius:999px;font-weight:700;font-size:12px;">{site}</span>')
# Code interne (CI) de chaque type d'installation — adapte librement les libellés
CI_CODES = {
    "Installations de gaz":       "B1",
    "Installations électriques":  "B2",
    "Equipements de levage":      "B3",
    "Appareil pression de gaz":   "A1",
    "Sécurité incendie":          "A2",
}
# Ordre d'affichage dans le tableau (reprend celui de ta capture)
ORDRE_CI = [
    "Installations de gaz",
    "Installations électriques",
    "Equipements de levage",
    "Appareil pression de gaz",
    "Sécurité incendie",
]

# Durée de visite en jours, par site et par installation.
# ⚠️ Cette donnée n'existe dans aucune feuille actuelle de ton Google Sheet :
# soit tu la fixes ici "en dur", soit (mieux) tu ajoutes une colonne "Nbr_jour"
# dans l'onglet Exigences et tu remplaces ce dict par une lecture de cette colonne.
NB_JOURS_VISITE = {
    "SGB": {"Installations de gaz": 1, "Installations électriques": 3, "Equipements de levage": 6,
            "Appareil pression de gaz": 1, "Sécurité incendie": 2},
    "MEG": {"Installations de gaz": 1, "Installations électriques": 3, "Equipements de levage": 6,
            "Appareil pression de gaz": 1, "Sécurité incendie": 1},
}
MOIS_FR = ["","Janvier","Février","Mars","Avril","Mai","Juin",
           "Juillet","Août","Septembre","Octobre","Novembre","Décembre"]
SOUS_EQUIPEMENTS = {
    "Installations électriques": [],
    "Equipements de levage": ["Transpalette","Table élévatrice","Potence","Pont roulant",
                               "Plateforme de travail","Nacelle","Gerbeur","Chariot élévateur","Palan électrique","Ascenseur"],
    "Sécurité incendie": [],
    "Installations de gaz": ["Industrielle","Chaudière"],
    "Appareil pression de gaz": []
}

LUCID_CARTOGRAPHIE_URL = "https://lucid.app/lucidspark/088f02a4-bdb7-4c79-8e28-64e05fc773c3/edit?beaconFlowId=69403DCAA7251095&invitationId=inv_16e69b3a-177f-4fb1-922e-fd6c28f294d5&page=0_0"

def _charger_cartographie_b64():
    """Charge l'image de cartographie du taux de non-conformité en base64 (fichier local)."""
    try:
        with open("Cartographie.png", "rb") as f:
            return base64.b64encode(f.read()).decode("ascii")
    except Exception:
        return None

# ==========================================
# STYLE
# ==========================================
st.html("""<style>
    [data-testid="stVVerticalBlockBorderBordered"]{background-color:#FFFFFF!important;border:1px solid #E2E8F0!important;border-left:5px solid #1E3A8A!important;border-radius:12px!important;box-shadow:0 4px 15px rgba(0,0,0,0.02)!important;padding:20px!important;}
    .stSelectbox label p{color:#475569!important;font-weight:600!important;font-size:13px!important;}
    div[data-baseweb="select"]{background-color:#F8FAFC!important;border:1px solid #CBD5E1!important;border-radius:8px!important;}
    div[data-baseweb="select"]>div{border:none!important;background-color:transparent!important;}
    div[data-baseweb="select"]:hover{border-color:#0EA5E9!important;background-color:#FFFFFF!important;box-shadow:0 0 0 3px rgba(14,165,233,0.12)!important;}
    div[data-baseweb="select"] span{color:#0F172A!important;font-weight:500!important;}
    div[data-testid="stTabs"] [data-baseweb="tab-list"],
    [data-baseweb="tab-list"],
    div[role="tablist"]{
        gap:12px!important;
        flex-wrap:wrap!important;
        border-bottom:none!important;
    }
    /* Supprime la barre de soulignement native (rouge/couleur du theme) */
    div[data-testid="stTabs"] [data-baseweb="tab-list"] > div:not([role="tab"]),
    div[data-testid="stTabs"] [data-baseweb="tab-highlight"],
    div[data-testid="stTabs"] [data-baseweb="tab-highlight-bar"],
    div[data-testid="stTabs"] [data-baseweb="tab-border"]{
        display:none!important;
        height:0!important;
        background:transparent!important;
        box-shadow:none!important;
        border:none!important;
    }
    div[data-testid="stTabs"] button,
    button[data-baseweb="tab"],
    [role="tab"]{
        font-size:17px!important;
        font-weight:700!important;
        color:#475569!important;
        background-color:#FFFFFF!important;
        padding:16px 32px!important;
        margin-right:0!important;
        border-radius:14px!important;
        border:2px solid #CBD5E1!important;
        box-shadow:0 2px 6px rgba(15,23,42,0.05)!important;
        transition:all 0.18s ease-in-out!important;
    }
    div[data-testid="stTabs"] button p,
    button[data-baseweb="tab"] p,
    [role="tab"] p,
    div[data-testid="stTabs"] button div,
    button[data-baseweb="tab"] div,
    [role="tab"] div{
        font-size:17px!important;
        font-weight:700!important;
        color:inherit!important;
    }
    div[data-testid="stTabs"] button:hover,
    button[data-baseweb="tab"]:hover,
    [role="tab"]:hover{
        color:#FFFFFF!important;
        background-color:#3B82F6!important;
        border:2px solid #1E3A8A!important;
        box-shadow:0 6px 14px rgba(30,58,138,0.25)!important;
    }
    div[data-testid="stTabs"] button[aria-selected="true"],
    button[data-baseweb="tab"][aria-selected="true"],
    [role="tab"][aria-selected="true"]{
        color:#FFFFFF!important;
        background-color:#1E3A8A!important;
        border:2px solid #1E3A8A!important;
        box-shadow:0 6px 16px rgba(14,165,233,0.35)!important;
    }

    /* Centrage global des titres de section (paragraphes en gras 1.2rem) */
    p[style*="font-size:1.2rem"],
    div[data-testid="stMarkdownContainer"] p[style*="font-size:1.2rem"]{
        text-align:center!important;
        width:100%!important;
        display:block!important;
    }

    /* ================= RESPONSIVE / MOBILE ================= */
    @media (max-width: 768px){
        h1{font-size:1.7rem!important;}
        div[data-testid="stTabs"] button,
        button[data-baseweb="tab"],
        [role="tab"]{
            font-size:13px!important;
            padding:10px 14px!important;
        }
        div[data-testid="stTabs"] button p,
        button[data-baseweb="tab"] p,
        [role="tab"] p{
            font-size:13px!important;
        }
        /* Les colonnes Streamlit passent en pile verticale sur petit écran */
        div[data-testid="stHorizontalBlock"]{
            flex-wrap:wrap!important;
        }
        div[data-testid="column"]{
            min-width:100%!important;
            flex:1 1 100%!important;
        }
        div[data-testid="stMetricValue"]{font-size:22px!important;}
    }
</style>""")

st.markdown("""<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
    html,body,[data-testid="stAppViewContainer"],[data-testid="stSidebarView"]{font-family:'Inter',sans-serif!important;background-color:#F8FAFC!important;}
    [data-testid="stForm"],.stCornerRadius{background-color:#FFFFFF!important;border:1px solid #E2E8F0!important;border-radius:12px!important;}
    .stButton>button{background-color:#1E3A8A!important;color:white!important;border-radius:8px!important;border:none!important;font-weight:500!important;padding:10px 24px!important;}

    /* Champ e-mail (page de connexion visiteur) : fond gris, coins arrondis, style aligné sur la maquette */
    div[data-testid="stTextInput"] input{
        background-color:#F1F5F9!important;
        border:1px solid #E2E8F0!important;
        border-radius:8px!important;
        padding:12px 16px!important;
        font-size:14px!important;
        color:#334155!important;
    }
    div[data-testid="stTextInput"] input:focus{
        border-color:#0EA5E9!important;
        box-shadow:0 0 0 3px rgba(14,165,233,0.12)!important;
        background-color:#FFFFFF!important;
    }
    div[data-testid="stTextInput"] input::placeholder{color:#94A3B8!important;}

    /* === FIX ROBUSTE : empêche les boutons de s'écraser/se casser verticalement ===
       Cause réelle : les colonnes Streamlit rétrécissent (flex-shrink) au lieu de
       passer à la ligne quand il n'y a pas assez de place horizontale.
       Solution : autoriser le retour à la ligne (flex-wrap) + bloquer le rétrécissement
       des colonnes + forcer le texte des boutons sur une seule ligne (nowrap). */

    div[data-testid="stHorizontalBlock"]{
        flex-wrap:wrap!important;
        row-gap:10px!important;
        column-gap:10px!important;
    }
    div[data-testid="column"]{
        flex:0 1 auto!important;
        width:auto!important;
        min-width:max-content!important;
    }
    .stButton, .stDownloadButton{
        width:auto!important;
    }
    .stButton>button, .stDownloadButton>button,
    div[data-testid="stButton"] button,
    div[data-testid="stDownloadButton"] button,
    button[kind="primary"], button[kind="secondary"],
    button[data-testid^="baseButton"]{
        white-space:nowrap!important;
        width:auto!important;
        min-width:unset!important;
        height:auto!important;
        line-height:1.3!important;
        font-size:14px!important;
        padding:10px 18px!important;
        display:inline-flex!important;
        align-items:center!important;
        justify-content:center!important;
    }
    /* Les libellés des boutons sont parfois dans un <p> ou <div> imbriqué
       qui possède son propre comportement de retour à la ligne : on le neutralise. */
    .stButton>button *, .stDownloadButton>button *,
    div[data-testid="stButton"] button *,
    div[data-testid="stDownloadButton"] button *,
    button[kind="primary"] *, button[kind="secondary"] *{
        white-space:nowrap!important;
    }
    .stDownloadButton>button{
        background-color:#16A34A!important;
        color:white!important;
        border-radius:8px!important;
        border:none!important;
        font-weight:600!important;
    }
    .stDownloadButton>button:hover{background-color:#15803D!important;}
    button[kind="primary"]{
        background-color:#1E3A8A!important;
        color:white!important;
        border-radius:8px!important;
        border:none!important;
        font-weight:600!important;
    }
    button[kind="primary"]:hover{background-color:#1D4ED8!important;}
</style>""", unsafe_allow_html=True)

# ==========================================
# API GOOGLE SHEETS
# ==========================================
def obtenir_access_token():
    try:
        import jwt as pyjwt
        private_key  = st.secrets["connections"]["gsheets"]["private_key"]
        client_email = st.secrets["connections"]["gsheets"]["client_email"]
        now = int(time.time())
        payload = {"iss":client_email,"scope":"https://www.googleapis.com/auth/spreadsheets",
                   "aud":"https://oauth2.googleapis.com/token","exp":now+3600,"iat":now}
        token_jwt = pyjwt.encode(payload, private_key, algorithm="RS256")
        resp = requests.post("https://oauth2.googleapis.com/token",
            data={"grant_type":"urn:ietf:params:oauth:grant-type:jwt-bearer","assertion":token_jwt},timeout=15)
        return resp.json()["access_token"] if resp.status_code==200 else None
    except Exception:
        return None

def sheets_append(onglet, valeurs):
    token = obtenir_access_token()
    if not token: return False,"Token invalide"
    try:
        url = f"https://sheets.googleapis.com/v4/spreadsheets/{SHEET_ID}/values/{onglet}!A:Z:append"
        resp = requests.post(url,headers={"Authorization":f"Bearer {token}","Content-Type":"application/json"},
            params={"valueInputOption":"RAW","insertDataOption":"INSERT_ROWS"},json={"values":[valeurs]},timeout=15)
        return (True,"") if resp.status_code==200 else (False,resp.text)
    except Exception as e:
        return False,str(e)

def sheets_lire(onglet, plage="A:Z"):
    token = obtenir_access_token()
    if not token: return pd.DataFrame()
    try:
        url = f"https://sheets.googleapis.com/v4/spreadsheets/{SHEET_ID}/values/{onglet}!{plage}"
        resp = requests.get(url,headers={"Authorization":f"Bearer {token}"},timeout=15)
        if resp.status_code!=200: return pd.DataFrame()
        valeurs = resp.json().get("values",[])
        if len(valeurs)<=1: return pd.DataFrame()
        entetes = [str(c).strip() for c in valeurs[0]]
        nb_col = len(entetes)
        # Complète les lignes trop courtes avec des cellules vides (Google Sheets omet les cellules vides en fin de ligne)
        lignes = [ (r + [""]*(nb_col-len(r)))[:nb_col] for r in valeurs[1:] ]
        df = pd.DataFrame(lignes,columns=entetes)
        # Nettoie les espaces superflus dans toutes les valeurs texte pour éviter les non-correspondances silencieuses
        for c in df.columns:
            if df[c].dtype == object:
                df[c] = df[c].astype(str).str.strip()
        return df
    except Exception:
        return pd.DataFrame()

def _obtenir_token_scope(scope):
    """Comme obtenir_access_token(), mais permet de demander un scope OAuth différent
    (ex: Drive en lecture) avec le même compte de service."""
    try:
        import jwt as pyjwt
        private_key  = st.secrets["connections"]["gsheets"]["private_key"]
        client_email = st.secrets["connections"]["gsheets"]["client_email"]
        now = int(time.time())
        payload = {"iss":client_email,"scope":scope,
                   "aud":"https://oauth2.googleapis.com/token","exp":now+3600,"iat":now}
        token_jwt = pyjwt.encode(payload, private_key, algorithm="RS256")
        resp = requests.post("https://oauth2.googleapis.com/token",
            data={"grant_type":"urn:ietf:params:oauth:grant-type:jwt-bearer","assertion":token_jwt},timeout=15)
        return resp.json()["access_token"] if resp.status_code==200 else None
    except Exception:
        return None


def codif_charger_classeur(sheet_id):
    """Télécharge le classeur de codification via l'API Google Drive.
    Essaie d'abord l'export Excel (/export?mimeType=...) qui fonctionne pour un Google Sheets
    natif (cas des classeurs MEG/SGB), puis se rabat sur le téléchargement direct (alt=media),
    utile pour un fichier .xlsx binaire jamais converti en Sheets natif.
    Retourne (dict {nom_onglet: DataFrame_brut_sans_entete}, message_erreur détaillé)."""
    token = _obtenir_token_scope("https://www.googleapis.com/auth/drive.readonly")
    if not token:
        return None, "Impossible d'obtenir un jeton d'accès Google (scope Drive)."
    headers = {"Authorization": f"Bearer {token}"}
    try:
        email = st.secrets["connections"]["gsheets"]["client_email"]
    except Exception:
        email = "(voir le champ client_email de vos secrets)"

    # 0) Vérification préalable : le compte de service voit-il seulement les métadonnées du fichier ?
    #    Permet de distinguer "fichier introuvable / ID erroné" de "fichier vu mais export refusé".
    meta_msg = ""
    try:
        resp_meta = requests.get(
            f"https://www.googleapis.com/drive/v3/files/{sheet_id}",
            headers=headers,
            params={"fields": "id,name,mimeType,driveId", "supportsAllDrives": "true"},
            timeout=15,
        )
        if resp_meta.status_code == 200:
            meta_msg = f" (fichier vu : {resp_meta.json().get('name','?')})"
        else:
            meta_msg = f" [métadonnées: HTTP {resp_meta.status_code} — {resp_meta.text[:200]}]"
    except Exception as e:
        meta_msg = f" [métadonnées: erreur réseau — {e}]"

    dernier_status = None
    dernier_texte = ""

    # 1) Tentative : export natif (Google Sheets -> xlsx). C'est la bonne méthode pour un
    #    classeur Google Sheets classique (ex: liens docs.google.com/spreadsheets/.../edit).
    try:
        url_export = f"https://www.googleapis.com/drive/v3/files/{sheet_id}/export"
        resp = requests.get(
            url_export, headers=headers,
            params={
                "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "supportsAllDrives": "true",
            },
            timeout=30,
        )
        if resp.status_code == 200:
            classeur = pd.read_excel(io.BytesIO(resp.content), sheet_name=None, header=None, engine="openpyxl")
            return classeur, None
        dernier_status, dernier_texte = resp.status_code, resp.text
    except Exception as e:
        dernier_status, dernier_texte = None, str(e)

    # 2) Repli : téléchargement direct (fichier .xlsx binaire jamais converti en Sheets natif).
    try:
        url = f"https://www.googleapis.com/drive/v3/files/{sheet_id}"
        resp = requests.get(url, headers=headers, params={"alt": "media", "supportsAllDrives": "true"}, timeout=30)
        if resp.status_code == 200:
            classeur = pd.read_excel(io.BytesIO(resp.content), sheet_name=None, header=None, engine="openpyxl")
            return classeur, None
        dernier_status, dernier_texte = resp.status_code, resp.text
    except Exception as e:
        return None, f"Erreur inattendue lors de la lecture du fichier Excel : {e}{meta_msg}"

    if dernier_status in (403, 404):
        return None, (f"Accès refusé au fichier de codification (ID {sheet_id}){meta_msg}. "
                       f"Vérifiez qu'il est bien partagé (en lecture) avec le compte de service : {email}. "
                       f"Détail Google : HTTP {dernier_status} — {str(dernier_texte)[:300]}")
    return None, f"Erreur API Google Drive ({dernier_status}){meta_msg} : {str(dernier_texte)[:300]}"


# Noms EXACTS des colonnes (Désignation | Observation | Code) pour chaque type d'installation,
# tels que mis à jour dans les classeurs de codification MEG et SGB (structure identique sur les
# deux sites, seul le nom de l'onglet change : "... - MEG" / "... - SGB").
# Format : (mots-clés identifiant le type d'installation à partir du nom de l'onglet,
#           nom exact colonne désignation, nom exact colonne observation, nom exact colonne code)
COLONNES_CODIF_PAR_INSTALLATION = [
    (["electrique"],           "Désignation", "Observation",                    "C"),
    (["levage"],                "Rapport",     "Organes examines NC",            "C"),
    (["pression"],               "Rapport",     "Points examines non conforme",   "C"),
    (["gaz"],                   "Rapport",     "Points examines non conforme",   "C"),
    (["incendie"],               "Local",       "Observation",                    "C"),
]


def _sans_accents(texte):
    """Retire les accents d'une chaîne pour faciliter les comparaisons insensibles aux accents."""
    return "".join(c for c in unicodedata.normalize("NFD", str(texte)) if unicodedata.category(c) != "Mn")


def _normaliser(texte):
    return _sans_accents(texte).strip().lower()


def _colonnes_attendues_pour_onglet(onglet):
    """Détermine, à partir du nom de l'onglet (ex: 'Installation électrique - MEG' ou
    'Sécurité incendie - SGB'), les noms exacts des colonnes Désignation/Observation/Code à
    utiliser pour cet onglet. Retourne None si le type d'installation n'est pas reconnu
    (repli sur la détection générique par mots-clés)."""
    if not onglet:
        return None
    # On retire un éventuel suffixe de site ("- MEG" / "- SGB") pour ne garder que le type
    # d'installation, puis on teste les mots-clés dans un ordre où les plus spécifiques
    # ("pression") sont vérifiés avant les plus génériques ("gaz").
    base = re.sub(r"\s*-\s*(meg|sgb)\s*$", "", onglet, flags=re.IGNORECASE).strip()
    base_norm = _normaliser(base)
    for mots_cles, nom_desig, nom_obs, nom_code in COLONNES_CODIF_PAR_INSTALLATION:
        if all(mc in base_norm for mc in mots_cles):
            return (nom_desig, nom_obs, nom_code)
    return None


def _trouver_colonne_exacte(colonnes, nom_attendu):
    """Recherche une colonne correspondant exactement (insensible aux accents/casse) au nom
    attendu ; à défaut, se rabat sur une correspondance partielle."""
    cible = _normaliser(nom_attendu)
    for c in colonnes:
        if _normaliser(c) == cible:
            return c
    for c in colonnes:
        if cible in _normaliser(c):
            return c
    return None


def _detecter_entete_et_nettoyer_codif(valeurs, onglet=None):
    """Prend les lignes brutes (liste de listes) d'un onglet du classeur de codification et
    retourne un DataFrame propre avec les colonnes Désignation | Observation | Code.
    Cherche automatiquement la ligne d'en-tête, en tolérant les différents intitulés utilisés
    selon les onglets (ex: 'Désignation'/'Rapport' pour l'équipement,
    'Observation'/'Organes examinés NC'/'Problème' pour l'action), et complète (forward-fill)
    les cellules d'équipement fusionnées verticalement dans la feuille source."""
    if not valeurs:
        return pd.DataFrame()

    MOTS_CLES_EQUIP = ["désignation", "désignation", "équipement", "equipement", "rapport"]
    MOTS_CLES_OBS   = ["observation", "organe", "examin", "problème", "probleme", "action"]

    # Colonnes exactes attendues pour ce type d'installation (déduites du nom de l'onglet),
    # sinon repli sur la détection générique par mots-clés (ancien comportement).
    colonnes_attendues = _colonnes_attendues_pour_onglet(onglet)

    idx_entete = None
    for i, ligne in enumerate(valeurs):
        cellules = [str(c).strip().lower() for c in ligne]
        if colonnes_attendues:
            nom_desig, nom_obs, _nom_code = colonnes_attendues
            cible_desig = _normaliser(nom_desig)
            cible_obs = _normaliser(nom_obs)
            a_equip = any(_normaliser(c) == cible_desig for c in cellules)
            a_obs   = any(_normaliser(c) == cible_obs for c in cellules)
        else:
            a_equip = any(any(mc in c for mc in MOTS_CLES_EQUIP) for c in cellules)
            a_obs   = any(any(mc in c for mc in MOTS_CLES_OBS) for c in cellules)
        if a_equip and a_obs:
            idx_entete = i
            break
    if idx_entete is None:
        return pd.DataFrame()

    entetes = [str(c).strip() for c in valeurs[idx_entete]]
    nb_col = len(entetes)
    lignes = valeurs[idx_entete + 1:]
    lignes = [(list(r) + [""] * (nb_col - len(r)))[:nb_col] for r in lignes]
    df = pd.DataFrame(lignes, columns=entetes)

    def _trouver_colonne(colonnes, groupes_mots_cles):
        """Cherche la colonne correspondant au groupe de mots-clés le plus spécifique possible
        (on essaie groupe par groupe, du plus spécifique au plus générique, et on ne retombe sur
        un groupe générique — ex: 'rapport', 'action' — que si aucune colonne plus spécifique
        n'a matché, pour éviter de capturer par erreur une autre colonne du même onglet)."""
        for groupe in groupes_mots_cles:
            for c in colonnes:
                if any(mc in c.lower() for mc in groupe):
                    return c
        return None

    if colonnes_attendues:
        nom_desig, nom_obs, nom_code = colonnes_attendues
        col_desig = _trouver_colonne_exacte(df.columns, nom_desig)
        col_obs   = _trouver_colonne_exacte(df.columns, nom_obs)
        col_code  = _trouver_colonne_exacte(df.columns, nom_code)
    else:
        col_desig = _trouver_colonne(df.columns, [
            ["désignation", "désignation", "équipement", "equipement"],
            ["rapport"],
        ])
        col_obs = _trouver_colonne(df.columns, [
            ["observation"],
            ["organe", "examin"],
            ["problème", "probleme"],
            ["action"],
        ])
        col_code  = next((c for c in df.columns if c.strip().lower() in ("c", "code")), None)
    if not (col_desig and col_obs and col_code):
        return pd.DataFrame()

    df = df[[col_desig, col_obs, col_code]].copy()
    df.columns = ["Désignation", "Observation", "Code"]
    for c in df.columns:
        df[c] = df[c].astype(str).str.strip()
        df[c] = df[c].replace("nan", "")
    df["Désignation"] = df["Désignation"].replace("", pd.NA).ffill().fillna("")
    df["Code"] = df["Code"].str.upper()
    df = df[(df["Observation"] != "") & (df["Code"] != "")]
    df = df[df["Code"].isin(NATURE_PILOTE.keys())]
    return df.reset_index(drop=True)


def _codes_pour_pilote(pilote_choisi):
    """Retourne la liste des codes (T,S,E,D,O,R) dont le champ Pilote (potentiellement
    combiné avec '+') contient l'entité choisie."""
    codes = []
    for code, (_, pilote_str) in NATURE_PILOTE.items():
        entites = [e.strip() for e in pilote_str.split("+") if e.strip()]
        if pilote_choisi in entites:
            codes.append(code)
    return codes


@st.cache_data(ttl=300, show_spinner=False)
def codif_charger_toutes_actions():
    """Charge et combine les classeurs de codification des deux sites (MEG et SGB).
    Retourne (DataFrame combiné [Site, Installation, Désignation, Observation, Code], message_erreur_ou_None)."""
    frames, erreurs = [], []
    for site, sheet_id in CODIF_SHEET_ID_PAR_SITE.items():
        classeur, err = codif_charger_classeur(sheet_id)
        if err:
            erreurs.append(f"{site} : {err}")
            continue
        if not classeur:
            continue
        for onglet, df_brut in classeur.items():
            valeurs = df_brut.fillna("").astype(str).values.tolist()
            d = _detecter_entete_et_nettoyer_codif(valeurs, onglet)
            if not d.empty:
                d["Installation"] = onglet
                d["Site"] = site
                frames.append(d)
    if not frames:
        return pd.DataFrame(), (" / ".join(erreurs) if erreurs else "Aucune donnée trouvée dans les classeurs de codification.")
    return pd.concat(frames, ignore_index=True), (" / ".join(erreurs) if erreurs else None)


def _cle_action(row):
    """Clé unique identifiant une action précise, utilisée pour repérer les actions déjà réalisées."""
    return "||".join(str(row.get(c, "")).strip().upper() for c in
                      ["Site", "Installation", "Désignation", "Observation", "Code"])


def lire_actions_realisees():
    """Lit l'onglet ActionsRealisees : Site | Installation | Désignation | Observation | Code | Pilote | Responsable | DateRealisation."""
    return sheets_lire("ActionsRealisees", "A:H")


def marquer_actions_realisees(df_lignes, responsable_nom):
    """Enregistre chaque action cochée comme réalisée (une ligne par action) dans l'onglet ActionsRealisees."""
    date_str = datetime.datetime.now(TZ).strftime("%d/%m/%Y %H:%M")
    ok_total = True
    for _, row in df_lignes.iterrows():
        ok, _msg = sheets_append("ActionsRealisees", [
            row.get("Site", ""), row.get("Installation", ""), row.get("Désignation", ""),
            row.get("Observation", ""), row.get("Code", ""), row.get("Pilote", ""),
            responsable_nom, date_str
        ])
        ok_total = ok_total and ok
    if ok_total and not df_lignes.empty:
        journaliser_action(utilisateur_courant(), "Action marquée réalisée",
                            f"{len(df_lignes)} action(s), pilote={responsable_nom}")
    return ok_total


def lire_suivi_encours():
    """Lit l'onglet SuiviActions : Site | Installation | Désignation | Observation | Code | Pilote |
    Statut | Type | Commentaire | Responsable | DateMaJ.
    Ne garde que la dernière saisie connue pour chaque action (une action peut être mise à jour
    plusieurs fois : seule la ligne la plus récente fait foi)."""
    df = sheets_lire("SuiviActions", "A:K")
    if df.empty:
        return df
    df["Cle"] = df.apply(_cle_action, axis=1)
    df = df.drop_duplicates(subset="Cle", keep="last").reset_index(drop=True)
    return df


def enregistrer_statut_en_cours(row, type_suivi, commentaire, responsable_nom):
    """Enregistre (nouvelle ligne d'historique) le statut « En cours » d'une action, avec son type
    de suivi (Immédiat / Sous-traitance / Planifié) et un commentaire libre facultatif."""
    date_str = datetime.datetime.now(TZ).strftime("%d/%m/%Y %H:%M")
    ok, _msg = sheets_append("SuiviActions", [
        row.get("Site", ""), row.get("Installation", ""), row.get("Désignation", ""),
        row.get("Observation", ""), row.get("Code", ""), row.get("Pilote", ""),
        "En cours", type_suivi, commentaire, responsable_nom, date_str
    ])
    if ok:
        journaliser_action(utilisateur_courant(), "Action mise à jour (En cours)",
                            f"{row.get('Code','')} — {type_suivi} — pilote={responsable_nom}")
    return ok


def _lignes_avec_rowspan(d_ins):
    """Regroupe les lignes consécutives ayant le même Equipement (Désignation) pour permettre
    une fusion de cellules (rowspan) dans le tableau PDF.
    Retourne une liste de tuples (equipement_ou_None, rowspan_ou_None, observation)."""
    valeurs = d_ins["Désignation"].tolist()
    obs = d_ins["Observation"].tolist()
    lignes = []
    i, n = 0, len(valeurs)
    while i < n:
        j = i
        while j < n and valeurs[j] == valeurs[i]:
            j += 1
        span = j - i
        lignes.append((valeurs[i], span, obs[i]))
        for k in range(i + 1, j):
            lignes.append((None, None, obs[k]))
        i = j
    return lignes



    ok, _ = sheets_ecrire_cellule_v2(onglet, cellule, valeur)
    return ok

def sheets_ecrire_cellule_v2(onglet, cellule, valeur):
    """Écrit une valeur dans une cellule précise. Retourne (ok, message_erreur)."""
    token = obtenir_access_token()
    if not token: return False, "Token invalide"
    try:
        url = f"https://sheets.googleapis.com/v4/spreadsheets/{SHEET_ID}/values/{onglet}!{cellule}"
        resp = requests.put(url,headers={"Authorization":f"Bearer {token}","Content-Type":"application/json"},
            params={"valueInputOption":"RAW"},json={"values":[[valeur]]},timeout=15)
        if resp.status_code == 200:
            return True, ""
        return False, f"HTTP {resp.status_code}: {resp.text[:200]}"
    except Exception as e:
        return False, str(e)

def sheets_trouver_ligne_email(onglet, email):
    token = obtenir_access_token()
    if not token: return None
    try:
        url = f"https://sheets.googleapis.com/v4/spreadsheets/{SHEET_ID}/values/{onglet}!A:A"
        resp = requests.get(url,headers={"Authorization":f"Bearer {token}"},timeout=15)
        if resp.status_code!=200: return None
        for i,row in enumerate(resp.json().get("values",[])):
            if row and row[0]==email: return i+1
        return None
    except Exception:
        return None

def ecrire_log(email):
    maintenant = datetime.datetime.now(TZ).strftime("%d/%m/%Y %H:%M")
    return sheets_append("Logs",[maintenant,email])

def mettre_a_jour_presence(email):
    maintenant = datetime.datetime.now(TZ).strftime("%d/%m/%Y %H:%M:%S")
    ligne = sheets_trouver_ligne_email("Presence",email)
    if ligne:
        token = obtenir_access_token()
        if token:
            url = f"https://sheets.googleapis.com/v4/spreadsheets/{SHEET_ID}/values/Presence!B{ligne}:C{ligne}"
            requests.put(url,headers={"Authorization":f"Bearer {token}","Content-Type":"application/json"},
                params={"valueInputOption":"RAW"},json={"values":[[maintenant,"En ligne"]]},timeout=15)
    else:
        sheets_append("Presence",[email,maintenant,"En ligne"])

def visiteur_deja_verifie(email: str) -> bool:
    """Retourne True si cet e-mail figure déjà dans l'onglet 'Presence', c.-à-d. qu'il
    s'agit d'un visiteur déjà connu (première vérification OTP déjà effectuée par le passé).
    Dans ce cas, on ne renvoie plus de code : l'accès est direct, mais on continue à
    tracer sa présence/historique comme pour n'importe quelle connexion."""
    return sheets_trouver_ligne_email("Presence", email) is not None

def lire_presence():
    df = sheets_lire("Presence","A:C")
    if df.empty: return pd.DataFrame(columns=["Email","Derniere_activite","Statut","Activite"])
    maintenant = datetime.datetime.now(TZ)
    resultats = []
    for _,row in df.iterrows():
        email=row.get("Email",""); derniere=row.get("Derniere_activite","")
        try:
            dt=TZ.localize(datetime.datetime.strptime(derniere,"%d/%m/%Y %H:%M:%S"))
            delta=(maintenant-dt).total_seconds()
            if delta<SEUIL_EN_LIGNE_SECONDES: statut="🟢 En ligne"; activite="Actif maintenant"
            elif delta<300:
                m=int(delta//60); statut="🟡 Récemment actif"; activite=f"Actif il y a {m} min" if m>0 else "Actif il y a quelques secondes"
            else:
                m=int(delta//60); h=int(m//60); statut="🔴 Hors ligne"
                activite=f"Vu il y a {h}h{m%60:02d}" if h>0 else f"Vu il y a {m} min"
        except Exception:
            statut="⚪ Inconnu"; activite=derniere
        resultats.append({"Email":email,"Dernière activité":derniere,"Statut":statut,"Activité":activite})
    return pd.DataFrame(resultats)

def lire_logs():
    return sheets_lire("Logs","A:B")
def lire_audit_log():
    """Lit l'onglet AuditLog : Date | Utilisateur | Action | Détails.
    (Onglet à créer manuellement dans le Google Sheet s'il n'existe pas encore.)"""
    return sheets_lire("AuditLog", "A:D")
def lire_exigences():
    """Lit l'onglet Exigences."""
    return sheets_lire("Exigences", "A:F")

def ecrire_contrat(lien_pdf):
    """Met à jour ou crée la ligne du contrat dans l'onglet Exigences."""
    df = lire_exigences()
    token = obtenir_access_token()
    if not token:
        return False, "Token invalide"

    if not df.empty and "Type" in df.columns:
        ligne_contrat = df[df["Type"] == "Contrat"]
        if not ligne_contrat.empty:
            num_ligne = ligne_contrat.index[0] + 2
            url = f"https://sheets.googleapis.com/v4/spreadsheets/{SHEET_ID}/values/Exigences!F{num_ligne}"
            resp = requests.put(url,
                headers={"Authorization": f"Bearer {token}", "Content-Type": "application/json"},
                params={"valueInputOption": "RAW"},
                json={"values": [[lien_pdf]]}, timeout=15)
            ok = (resp.status_code == 200)
            if ok:
                journaliser_action(utilisateur_courant(), "Contrat mis à jour", lien_pdf)
            return ok, ""

    ok, msg = sheets_append("Exigences", ["Contrat", "", "", "", "", lien_pdf])
    if ok:
        journaliser_action(utilisateur_courant(), "Contrat créé", lien_pdf)
    return ok, msg


def supprimer_contrat():
    """Vide le lien du contrat."""
    df = lire_exigences()
    token = obtenir_access_token()
    if not token or df.empty or "Type" not in df.columns:
        return False
    ligne_contrat = df[df["Type"] == "Contrat"]
    if ligne_contrat.empty:
        return False
    num_ligne = ligne_contrat.index[0] + 2
    url = f"https://sheets.googleapis.com/v4/spreadsheets/{SHEET_ID}/values/Exigences!F{num_ligne}"
    resp = requests.put(url,
        headers={"Authorization": f"Bearer {token}", "Content-Type": "application/json"},
        params={"valueInputOption": "RAW"},
        json={"values": [[""]]}, timeout=15)
    ok = resp.status_code == 200
    if ok:
        journaliser_action(utilisateur_courant(), "Contrat supprimé")
    return ok


def ajouter_equipement(site, installation, sous_eq, nombre):
    """Ajoute une ligne équipement dans Exigences."""
    ok, msg = sheets_append("Exigences", ["Equipement", site, installation, sous_eq, str(nombre), ""])
    if ok:
        journaliser_action(utilisateur_courant(), "Equipement ajouté", f"{site} / {installation} / {sous_eq} x{nombre}")
    return ok, msg


def supprimer_equipement_ligne(num_ligne_sheet):
    """Vide une ligne équipement (remplace par des cellules vides)."""
    token = obtenir_access_token()
    if not token: return False
    url = f"https://sheets.googleapis.com/v4/spreadsheets/{SHEET_ID}/values/Exigences!A{num_ligne_sheet}:F{num_ligne_sheet}"
    resp = requests.put(url,
        headers={"Authorization": f"Bearer {token}", "Content-Type": "application/json"},
        params={"valueInputOption": "RAW"},
        json={"values": [["", "", "", "", "", ""]]}, timeout=15)
    ok = resp.status_code == 200
    if ok:
        journaliser_action(utilisateur_courant(), "Equipement supprimé", f"ligne {num_ligne_sheet}")
    return ok


# ==========================================
# Actions de contrôle (onglet dédié "PointsReserve")
# ==========================================
def lire_points_reserve():
    """Lit l'onglet PointsReserve : Site | Installation | Sous_equipement | Nombre."""
    return sheets_lire("PointsReserve", "A:D")


def ajouter_point_reserve(site, installation, sous_eq, nombre):
    """Ajoute une ligne dans l'onglet PointsReserve."""
    ok, msg = sheets_append("PointsReserve", [site, installation, sous_eq, str(nombre)])
    if ok:
        journaliser_action(utilisateur_courant(), "Point de réserve ajouté", f"{site} / {installation} / {sous_eq} x{nombre}")
    return ok, msg


# ==========================================
# Actions de contrôle PAR NATURE (onglet dédié "PointsReserveNature")
# Table KPI : Site | Installation | Nombre de actions | Nature | Pilote
# Le pilote est déduit automatiquement du code de la nature.
# ==========================================
NATURE_PILOTE = {
    "T": ("Technique",      "Maintenance"),
    "S": ("Sécurité",       "HSE"),
    "E": ("Energétique",    "BT + Maintenance"),
    "D": ("Documentation",  "BT + HSE"),
    "O": ("Organisation",   "DMTN + Chef service BT"),
    "R": ("Règlementation", "BT + HSE + RH + DG"),
}

# Nom du sous-pilote (personne responsable) associé à chaque entité de pilotage.
# Les entités non listées ici (ex: DMTN) affichent l'entité elle-même à défaut de nom connu.
SOUS_PILOTE_NOMS = {
    "Maintenance":     "Saber BEN CHAABEN",
    "HSE":             "Montassar MEHRABI",
    "BT":              "Aïcha BELLAKHAL",
    "Chef service BT": "Aïcha BELLAKHAL",
    "RH":              "Aïcha BELLAKHAL",
    "DG":              "Aïcha BELLAKHAL",
}

def nom_pour_pilote_site(pilote, site=None):
    """Renvoie le nom du responsable à afficher pour un pilote donné, en tenant compte du site
    lorsque plusieurs responsables se partagent le même pilote (ex : Maintenance -> Saber pour MEG,
    Chafik pour SGB)."""
    if pilote == "Maintenance":
        site_norm = (site or "").strip().upper()
        if site_norm == "MEG":
            return "Saber BEN CHAABEN"
        if site_norm == "SGB":
            return "Chafik ABID"
        return "Maintenance (Saber BEN CHAABEN — MEG / Chafik ABID — SGB)"
    return SOUS_PILOTE_NOMS.get(pilote, pilote)

def lire_points_reserve_nature():
    """Lit l'onglet PointsReserveNature : Site | Installation | Nombre | Nature | Pilote."""
    return sheets_lire("PointsReserveNature", "A:E")


def ajouter_point_reserve_nature(site, installation, nombre, code_nature):
    """Ajoute une ligne dans l'onglet PointsReserveNature. Le pilote est calculé depuis le code de la nature."""
    nature_nom, pilote = NATURE_PILOTE[code_nature]
    ok, msg = sheets_append("PointsReserveNature", [site, installation, str(nombre), nature_nom, pilote])
    if ok:
        journaliser_action(utilisateur_courant(), "Point de réserve (nature) ajouté", f"{site} / {installation} — {nature_nom} x{nombre}")
    return ok, msg


def supprimer_ligne_generique(onglet, num_ligne_sheet, nb_colonnes):
    """Vide une ligne (remplace par des cellules vides) dans un onglet donné, sur nb_colonnes colonnes (A..)."""
    token = obtenir_access_token()
    if not token: return False
    derniere_col = chr(ord('A') + nb_colonnes - 1)
    url = f"https://sheets.googleapis.com/v4/spreadsheets/{SHEET_ID}/values/{onglet}!A{num_ligne_sheet}:{derniere_col}{num_ligne_sheet}"
    resp = requests.put(url,
        headers={"Authorization": f"Bearer {token}", "Content-Type": "application/json"},
        params={"valueInputOption": "RAW"},
        json={"values": [[""] * nb_colonnes]}, timeout=15)
    ok = resp.status_code == 200
    if ok:
        journaliser_action(utilisateur_courant(), "Ligne supprimée", f"onglet={onglet}, ligne={num_ligne_sheet}")
    return ok
# ==========================================
# CHARGEMENT DONNÉES
# ==========================================
@st.cache_data(ttl=30)
def charger_donnees_sheet(nom_onglet):
    try:
        base_url = URL_GOOGLE_SHEET.split("/edit")[0]
        df = pd.read_csv(f"{base_url}/gviz/tq?tqx=out:csv&sheet={nom_onglet}")
        return df.dropna(how='all')
    except Exception:
        return pd.DataFrame()

df_rapports = charger_donnees_sheet("Rapports")
df_planning = charger_donnees_sheet("Planning")

# ==========================================
# SIDEBAR
# ==========================================
with st.sidebar:
    st.markdown("<br>",unsafe_allow_html=True)
    _,c2,_=st.columns([1,4,1])
    with c2:
        st.image("https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcR6q1BtDSDgVnJZFo0hOBfQJoDS6OYiub-qfQ&s",use_container_width=True)
    st.markdown("""<div style="text-align:center;margin-top:15px;margin-bottom:25px;">
        <h3 style="font-size:1.15rem;font-weight:700;margin-bottom:4px;color:#0F172A;">Tunisie Profilés d'Aluminium</h3>
        <p style="font-size:0.85rem;color:#64748B;margin:0;font-weight:500;text-transform:uppercase;letter-spacing:0.5px;">Direction Maintenance & TN</p>
    </div>""",unsafe_allow_html=True)
    st.divider()
    st.markdown("<p style='font-weight:600;color:#334155;margin-bottom:0;'>🔐 Espace sécurisé</p>",unsafe_allow_html=True)
    role=st.selectbox("Profil :",["Visiteur","Responsable","Admin"],label_visibility="collapsed")
    password_correct=False

    ADMIN_HASH_DEFAUT = "69d097edbe715222649bfc7296f14331$ed69ae3c1298c0254a053fa6650745b588637332095855e2541c1aeb784fff98"
    RESPONSABLES={
        "SABER": {"hash":"df08cf138fa54315e428e454a2bfcd91$2a12dbc095764565b0fe5682001a39eb738dc01c1b8c4fbedead2a9d170b3a13","nom":"Saber BEN CHAABEN","entites":["Maintenance"],"site":"MEG"},
        "HSE":   {"hash":"316b35609d01dcc04bd1cc0610bc6768$a3835ace3975737dcda29c8c83ff6e7cbab3a223a40ec464ea4d6dfd4cf01da1",  "nom":"Montassar MEHRABI","entites":["HSE"],"site":None},
        "AICHA": {"hash":"4c990dd6d323e5c52ed2175e4ebc9580$5bc468e84a0d1323ee52a1d6b1e0281d27c9ddf184b417290fceeacc10ba6a99","nom":"Aïcha BELLAKHAL",  "entites":["BT","Chef service BT","RH","DG"],"site":None},
        "CHAFIK": {"hash":"a69216e51a745c746dce04d1d9a9ea3f$9263b18f3fdfb31fabd1a912dc8009fe054b2e0b1818e2c1f090ca07494f150f","nom":"Chafik ABID",    "entites":["Maintenance"],"site":"SGB"},
    }
    # Surcharge par st.secrets si configuré (prioritaire sur les valeurs par défaut ci-dessus)
    try:
        ADMIN_HASH = st.secrets["auth"]["admin_hash"]
    except Exception:
        ADMIN_HASH = ADMIN_HASH_DEFAUT
    try:
        _hash_secrets = st.secrets["auth"]["responsables"]
        for _id, _h in _hash_secrets.items():
            if _id in RESPONSABLES:
                RESPONSABLES[_id]["hash"] = _h
    except Exception:
        pass

    if "responsable_connecte" not in st.session_state: st.session_state.responsable_connecte=False
    if "responsable_actif" not in st.session_state: st.session_state.responsable_actif=None

    if role=="Admin":
        secondes_restantes = tentative_bloquee("admin")
        if secondes_restantes > 0:
            st.error(f"🔒 Trop de tentatives échouées. Réessaie dans {secondes_restantes} s.")
        else:
            password=st.text_input("Code d'accès :",type="password",placeholder="•••")
            if password and verifier_mot_de_passe(password, ADMIN_HASH):
                password_correct=True
                reinitialiser_echecs("admin")
                st.success("Accès administrateur validé")
                if "responsable_log_enregistre" not in st.session_state:
                    st.session_state.responsable_log_enregistre=True
                    now_str=datetime.datetime.now(TZ).strftime("%d/%m/%Y %H:%M")
                    sheets_append("Logs",[now_str,"responsable@admin"])
            elif password:
                enregistrer_echec("admin")
                st.error("Code d'accès incorrect")
    elif role=="Responsable":
        secondes_restantes = tentative_bloquee("responsable")
        if secondes_restantes > 0:
            st.error(f"🔒 Trop de tentatives échouées. Réessaie dans {secondes_restantes} s.")
        else:
            identifiant=st.text_input("Identifiant :",placeholder="Identifiant").strip().upper()
            mdp_resp=st.text_input("Mot de passe :",type="password",placeholder="•••")
            if identifiant and mdp_resp:
                compte=RESPONSABLES.get(identifiant)
                if compte and verifier_mot_de_passe(mdp_resp, compte["hash"]):
                    st.session_state.responsable_connecte=True
                    st.session_state.responsable_actif=identifiant
                    reinitialiser_echecs("responsable")
                    st.success(f"Accès responsable validé : {compte['nom']}")
                    if st.session_state.get("responsable_dernier_log")!=identifiant:
                        st.session_state.responsable_dernier_log=identifiant
                        now_str=datetime.datetime.now(TZ).strftime("%d/%m/%Y %H:%M")
                        sheets_append("Logs",[now_str,f"{identifiant.lower()}@responsable"])
                else:
                    st.session_state.responsable_connecte=False
                    st.session_state.responsable_actif=None
                    enregistrer_echec("responsable")
                    st.error("Identifiant ou mot de passe incorrect")
            else:
                st.session_state.responsable_connecte=False
                st.session_state.responsable_actif=None

# ==========================================
# CONTRÔLE D'ACCÈS
# ==========================================
def format_email_valide(email):
    return re.match(r"[^@]+@[^@]+\.[^@]+",email) is not None

acces_autorise=(role=="Admin" and password_correct) or (role=="Responsable" and st.session_state.responsable_connecte) or (role=="Visiteur" and st.session_state.email_visiteur)

# ==========================================
# EN-TÊTE (toujours affiché en premier, connecté ou non)
# ==========================================
# --> Renseignez ici le nom (ou le chemin) de votre fichier image PNG déjà importé
MAINTENANCE_ICON_PATH = "unnamed.png"

col_titre, col_icone = st.columns([5,1])
with col_titre:
    st.markdown("""<div class="app-header-block" style="width:100%;margin:10px auto 0 auto;">
    <h1 style="text-align:center;font-size:2.6rem;font-weight:800;color:#0F172A;margin:0 0 6px 0;letter-spacing:-1px;line-height:1.2;">Tableau de Bord Réglementaire</h1>
    <p style="text-align:center;font-size:1.05rem;color:#64748B;margin:0 auto;font-weight:400;line-height:1.5;max-width:800px;">L'amélioration continue.. Notre trajectoire..</p>
</div>""",unsafe_allow_html=True)
with col_icone:
    st.image(MAINTENANCE_ICON_PATH, use_container_width=True)

if not acces_autorise and role=="Visiteur":
    # ---- Vérification réelle de l'adresse e-mail par code à usage unique (OTP) ----
    # Étape 1 : l'utilisateur saisit son e-mail et reçoit un code à 6 chiffres (valable 10 min).
    # Étape 2 : il doit saisir ce code pour prouver qu'il possède bien cette boîte e-mail.
    if "otp_code" not in st.session_state: st.session_state.otp_code = None
    if "otp_email" not in st.session_state: st.session_state.otp_email = None
    if "otp_expire_a" not in st.session_state: st.session_state.otp_expire_a = 0
    if "otp_dernier_envoi" not in st.session_state: st.session_state.otp_dernier_envoi = 0

    otp_deja_envoye = st.session_state.otp_code is not None and time.time() < st.session_state.otp_expire_a

    if not otp_deja_envoye:
        st.markdown("""<div style="margin-bottom:14px;">
            <p style="color:#0F172A;font-size:15px;font-weight:700;margin:0 0 6px 0;">Adresse e-mail :</p>
            <p style="color:#64748B;font-size:13.5px;margin:0;line-height:1.5;">Lors de votre première connexion, un code de vérification à 6 chiffres vous sera envoyé par e-mail. Les fois suivantes, l'accès est direct.</p>
        </div>""",unsafe_allow_html=True)
        email_saisi=st.text_input("Adresse e-mail :",placeholder="exemple@domain.com",label_visibility="collapsed",key="otp_email_input")
        blocage_otp = tentative_bloquee("otp")
        if blocage_otp > 0:
            st.error(f"🔒 Trop de tentatives échouées. Réessaie dans {blocage_otp} s.")
        elif st.button("Continuer",type="primary"):
            if not format_email_valide(email_saisi):
                st.error("Veuillez saisir une adresse e-mail valide.")
            elif visiteur_deja_verifie(email_saisi):
                # ---- Visiteur déjà connu (déjà vérifié par le passé) : accès direct, sans nouveau code ----
                # On garde malgré tout la traçabilité (log + présence) pour l'historique.
                st.session_state.email_visiteur=email_saisi
                with st.spinner("Enregistrement de votre accès..."):
                    succes,erreur=ecrire_log(email_saisi)
                    mettre_a_jour_presence(email_saisi)
                if succes:
                    st.success("✅ Bon retour ! Accès accordé.")
                    st.rerun()
                else:
                    st.error(f"❌ Erreur d'enregistrement : {erreur}")
                    st.stop()
            elif time.time() - st.session_state.otp_dernier_envoi < 60:
                st.warning("Merci de patienter avant de redemander un code.")
            else:
                code = f"{secrets_lib.randbelow(1_000_000):06d}"
                succes_envoi, erreur_envoi = envoyer_email(
                    email_saisi,
                    "Votre code d'accès — Tableau de Bord Réglementaire",
                    f"Votre code de vérification est : {code}\n\nCe code est valable 10 minutes.\nSi vous n'êtes pas à l'origine de cette demande, ignorez cet e-mail."
                )
                if succes_envoi:
                    st.session_state.otp_code = code
                    st.session_state.otp_email = email_saisi
                    st.session_state.otp_expire_a = time.time() + 600
                    st.session_state.otp_dernier_envoi = time.time()
                    st.success("✅ Code envoyé. Vérifiez votre boîte e-mail (et vos spams).")
                    st.rerun()
                elif "Configuration SMTP absente" in (erreur_envoi or ""):
                    # Repli si l'administrateur n'a pas encore configuré le SMTP : on ne bloque
                    # pas l'accès, mais on ne peut pas garantir que l'adresse est vérifiée.
                    st.session_state.email_visiteur=email_saisi
                    with st.spinner("Enregistrement de votre accès..."):
                        succes,erreur=ecrire_log(email_saisi)
                        mettre_a_jour_presence(email_saisi)
                    st.warning("⚠️ Vérification par code indisponible (SMTP non configuré côté administrateur) — accès accordé sans confirmation de l'e-mail.")
                    st.rerun()
                else:
                    st.error(f"❌ Impossible d'envoyer le code : {erreur_envoi}")
    else:
        st.markdown(f"""<div style="margin-bottom:14px;">
            <p style="color:#0F172A;font-size:15px;font-weight:700;margin:0 0 6px 0;">Code de vérification :</p>
            <p style="color:#64748B;font-size:13.5px;margin:0;line-height:1.5;">Un code a été envoyé à {st.session_state.otp_email}. Saisissez-le ci-dessous.</p>
        </div>""",unsafe_allow_html=True)
        code_saisi = st.text_input("Code :",placeholder="123456",label_visibility="collapsed",max_chars=6,key="otp_code_input")
        c_valider, c_renvoyer = st.columns(2)
        with c_valider:
            if st.button("Valider le code",type="primary",use_container_width=True):
                if code_saisi and hmac.compare_digest(code_saisi.strip(), st.session_state.otp_code):
                    email_confirme = st.session_state.otp_email
                    st.session_state.email_visiteur=email_confirme
                    st.session_state.otp_code = None
                    st.session_state.otp_email = None
                    reinitialiser_echecs("otp")
                    with st.spinner("Enregistrement de votre accès..."):
                        succes,erreur=ecrire_log(email_confirme)
                        mettre_a_jour_presence(email_confirme)
                    if succes:
                        st.success("✅ Accès accordé. Bienvenue !")
                        st.rerun()
                    else:
                        st.error(f"❌ Erreur d'enregistrement : {erreur}")
                        st.stop()
                else:
                    enregistrer_echec("otp")
                    st.error("Code incorrect.")
        with c_renvoyer:
            if st.button("↻ Renvoyer / changer d'e-mail",use_container_width=True):
                st.session_state.otp_code = None
                st.session_state.otp_email = None
                st.rerun()

# ==========================================
# HEARTBEAT
# ==========================================
if role=="Admin" and password_correct: email_actif="responsable@admin"
elif role=="Responsable" and st.session_state.responsable_connecte: email_actif=f"{st.session_state.responsable_actif.lower()}@responsable"
elif role=="Visiteur" and st.session_state.email_visiteur: email_actif=st.session_state.email_visiteur
else: email_actif=None

if acces_autorise and email_actif:
    if "last_heartbeat" not in st.session_state: st.session_state.last_heartbeat=0
    now_ts=time.time()
    if now_ts-st.session_state.last_heartbeat>30:
        mettre_a_jour_presence(email_actif)
        st.session_state.last_heartbeat=now_ts
    st.markdown("""<script>setTimeout(function(){window.parent.document.querySelector('[data-testid="stApp"]').click();},30000);</script>""",unsafe_allow_html=True)

# ==========================================
# CONTENU PRINCIPAL
# ==========================================
if acces_autorise:
    val_total=len(df_rapports) if not df_rapports.empty else 0

    # ---- Contrôles réalisés en 2026 / total des contrôles suivis (même dédup que les KPI) ----
    col_reelle_hdr = [c for c in df_rapports.columns if "reelle" in c.lower() or "réelle" in c.lower()]
    col_ins_hdr    = [c for c in df_rapports.columns if "ins" in c.lower()]
    col_site_hdr   = [c for c in df_rapports.columns if "site" in c.lower()]
    col_label_hdr  = [c for c in df_rapports.columns if "equip" in c.lower() or "label" in c.lower() or "nom" in c.lower()]
    col_date_hdr   = [c for c in df_rapports.columns if "date" in c.lower() and "reelle" not in c.lower() and "réelle" not in c.lower() and "prochaine" not in c.lower() and "planifi" not in c.lower()]

    # ---- Contrôles 2026 réalisés PAR SITE (SGB / MEG), PAR CATÉGORIE d'installation (et non par équipement) ----
    # Périodicité : Installations électriques = 2 campagnes/an, les 4 autres catégories = 1 campagne/an
    # => 2 + 1 + 1 + 1 + 1 = 6 campagnes attendues par site sur l'année 2026
    def _nb_campagnes_attendues_hdr(installation):
        return round(12 / PERIODICITE.get(installation, 12))

    TOTAL_CATEGORIES_PAR_SITE = sum(_nb_campagnes_attendues_hdr(ins) for ins in PERIODICITE.keys())  # = 6

    nb_ctrl_site = {"SGB": 0, "MEG": 0}
    if not df_rapports.empty and col_ins_hdr and col_date_hdr:
        df_hdr = df_rapports.copy()
        df_hdr["_date_brute"]  = pd.to_datetime(df_hdr[col_date_hdr[0]], dayfirst=True, errors='coerce')
        df_hdr["_date_reelle"] = pd.to_datetime(df_hdr[col_reelle_hdr[0]], dayfirst=True, errors='coerce') if col_reelle_hdr else pd.NaT
        df_hdr = df_hdr.dropna(subset=["_date_brute"])
        df_realises_hdr = df_hdr[df_hdr["_date_reelle"].notna() & (df_hdr["_date_reelle"].dt.year == 2026)]
        if col_site_hdr and col_ins_hdr:
            for site_h in ("SGB", "MEG"):
                df_site_h = df_realises_hdr[df_realises_hdr[col_site_hdr[0]].astype(str).str.strip().str.upper() == site_h]
                total_site = 0
                for ins_h in PERIODICITE.keys():
                    attendu_h = _nb_campagnes_attendues_hdr(ins_h)
                    df_grp_h  = df_site_h[df_site_h[col_ins_hdr[0]].astype(str).str.strip() == ins_h]
                    nb_camp_h = df_grp_h["_date_brute"].nunique() if not df_grp_h.empty else 0
                    total_site += min(nb_camp_h, attendu_h)
                nb_ctrl_site[site_h] = total_site

    def _pct_et_couleur(nb, total):
        pct = round(nb/total*100) if total>0 else 0
        couleur = "#10B981" if pct>=80 else "#F97316" if pct>=50 else "#EF4444"
        return pct, couleur

    pct_sgb, couleur_sgb = _pct_et_couleur(nb_ctrl_site["SGB"], TOTAL_CATEGORIES_PAR_SITE)
    pct_meg, couleur_meg = _pct_et_couleur(nb_ctrl_site["MEG"]-1, TOTAL_CATEGORIES_PAR_SITE)

    k1,k2=st.columns(2)
    with k1:
        st.markdown(f"""<div style="background:white;padding:22px;border-radius:12px;box-shadow:0 4px 6px -1px rgba(0,0,0,0.05);border-left:5px solid #1E3A8A;height:118px;box-sizing:border-box;display:flex;flex-direction:column;justify-content:center;">
            <p style="margin:0;font-size:12px;color:#64748B;font-weight:600;text-transform:uppercase;letter-spacing:0.5px;">Total Rapports Archivés</p>
            <p style="margin:8px 0 0 0;font-size:34px;color:#0F172A;font-weight:700;line-height:1;">{val_total}</p></div>""",unsafe_allow_html=True)
    with k2:
        st.markdown(f"""<div style="background:white;padding:22px;border-radius:12px;box-shadow:0 4px 6px -1px rgba(0,0,0,0.05);border-left:5px solid #0EA5E9;height:118px;box-sizing:border-box;display:flex;flex-direction:column;justify-content:center;gap:8px;">
            <p style="margin:0;font-size:12px;color:#64748B;font-weight:600;text-transform:uppercase;letter-spacing:0.5px;">Contrôles Réalisés 2026</p>
            <div style="display:flex;align-items:center;gap:8px;">
                <span>{badge_site("SGB")}</span>
                <div style="flex:1;height:8px;background:#E2E8F0;border-radius:4px;overflow:hidden;">
                    <div style="width:{pct_sgb}%;height:100%;background:{COULEUR_SITE['SGB']['principale']};border-radius:4px;transition:width 0.6s ease-in-out;"></div>
                </div>
                <span style="font-size:11px;color:{couleur_sgb};font-weight:700;white-space:nowrap;width:70px;text-align:right;">{nb_ctrl_site["SGB"]}/{TOTAL_CATEGORIES_PAR_SITE} ({pct_sgb}%)</span>
            </div>
            <div style="display:flex;align-items:center;gap:8px;">
                <span>{badge_site("MEG")}</span>
                <div style="flex:1;height:8px;background:#E2E8F0;border-radius:4px;overflow:hidden;">
                    <div style="width:{pct_meg}%;height:100%;background:{COULEUR_SITE['MEG']['principale']};border-radius:4px;transition:width 0.6s ease-in-out;"></div>
                </div>
                <span style="font-size:11px;color:{couleur_meg};font-weight:700;white-space:nowrap;width:70px;text-align:right;">{nb_ctrl_site["MEG"]-1}/{TOTAL_CATEGORIES_PAR_SITE} ({pct_meg}%)</span>
            </div></div>""",unsafe_allow_html=True)

    st.markdown("<br>",unsafe_allow_html=True)

    afficher_suivi_actions = (role == "Admin" and password_correct) or (role == "Responsable" and st.session_state.responsable_connecte)

    liste_onglets = ["📋 Rapports CR","📅 Planification","📌 Exigences"]
    if role == "Admin" and password_correct:
        liste_onglets.append("👥 Statistiques")
    liste_onglets.append("📊 KPI")
    if afficher_suivi_actions:
        liste_onglets.append("✅ Suivi des actions")
    onglets = st.tabs(liste_onglets)
    tab1, tab2, tab_exigences = onglets[0], onglets[1], onglets[2]
    tab3 = None
    tab_kpi = None
    tab_suivi = None
    _idx = 3
    if role == "Admin" and password_correct:
        tab3 = onglets[_idx]; _idx += 1
    tab_kpi = onglets[_idx]; _idx += 1
    if afficher_suivi_actions:
        tab_suivi = onglets[_idx]; _idx += 1

    def convertir_lien(url):
        try:
            if "drive.google.com" in str(url) and "/file/d/" in str(url):
                file_id = str(url).split('/file/d/')[1].split('/')[0]
                return f"https://drive.google.com/file/d/{file_id}/preview"
        except Exception: pass
        return url
    
    # ---- ONGLET 1 : RAPPORTS ----
    with tab1:
        st.markdown("""<style>
            .filter-title{text-align:center!important;font-weight:600;color:#1E293B;margin-top:0;margin-bottom:15px;width:100%;}
            div[data-testid="stSelectbox"] label p{text-align:center!important;width:100%;display:block;}
        </style>""",unsafe_allow_html=True)
        with st.container(border=True):
            st.markdown("<p class='filter-title'>Filtres de recherche</p>",unsafe_allow_html=True)
            c1,c2,c3,c4=st.columns(4)
            with c1: f_site =st.selectbox("Site",["Tous","SGB","MEG"])
            with c2: f_annee=st.selectbox("Année",["Tous","2025","2026"])
            with c3: f_ins  =st.selectbox("Installation",["Tous"]+list(SOUS_EQUIPEMENTS.keys()))
            with c4:
                opts=["Tous"]+SOUS_EQUIPEMENTS[f_ins] if f_ins!="Tous" else ["Tous"]+[i for sub in SOUS_EQUIPEMENTS.values() for i in sub]
                f_sous_eq=st.selectbox("Sous-équipement",opts)

        st.markdown("<br><p style='font-size:1.2rem;font-weight:700;color:#0F172A;margin-bottom:10px;'>📂 Documents rattachés</p>",unsafe_allow_html=True)
        df_f=df_rapports.copy()
        col_site=[c for c in df_f.columns if "site" in c.lower()]
        col_ex  =[c for c in df_f.columns if "exerc" in c.lower() or "ann" in c.lower()]
        col_ins =[c for c in df_f.columns if "ins" in c.lower()]
        col_seq =[c for c in df_f.columns if "sous" in c.lower()]
        col_lien=[c for c in df_f.columns if "lien" in c.lower() or "pdf" in c.lower()]
        col_date=[c for c in df_f.columns if "date" in c.lower() or "contr" in c.lower()]
        if not df_f.empty:
            if f_site !="Tous" and col_site: df_f=df_f[df_f[col_site[0]].astype(str).str.strip()==f_site]
            if f_annee!="Tous" and col_ex:   df_f=df_f[pd.to_numeric(df_f[col_ex[0]],errors='coerce')==int(f_annee)]
            if f_ins  !="Tous" and col_ins:  df_f=df_f[df_f[col_ins[0]].astype(str).str.strip()==f_ins]
            if f_sous_eq!="Tous" and col_seq:df_f=df_f[df_f[col_seq[0]].astype(str).str.strip()==f_sous_eq]
            if col_lien: df_f[col_lien[0]]=df_f[col_lien[0]].apply(convertir_lien)
            if col_date: df_f[col_date[0]]=pd.to_datetime(df_f[col_date[0]],dayfirst=True,errors='coerce')
        if not df_f.empty:
            col_reelle_doc=[c for c in df_f.columns if "reelle" in c.lower() or "réelle" in c.lower()]
            if col_reelle_doc: df_f=df_f.drop(columns=col_reelle_doc)
            col_planifiee_doc=[c for c in df_f.columns if "planifi" in c.lower()]
            if col_planifiee_doc: df_f=df_f.drop(columns=col_planifiee_doc)
            col_prochaine_doc=[c for c in df_f.columns if "prochaine" in c.lower()]
            if col_prochaine_doc: df_f=df_f.drop(columns=col_prochaine_doc)
            st.dataframe(df_f,column_config={
                (col_lien[0] if col_lien else "Lien PDF"):st.column_config.LinkColumn("Action",display_text="Voir le rapport"),
                (col_ex[0]   if col_ex   else "Année"):st.column_config.NumberColumn("Année",format="%d"),
                (col_date[0] if col_date else "Date"):    st.column_config.DateColumn("Date de dernier contrôle",format="DD/MM/YYYY"),
            },hide_index=True,use_container_width=True)
        else:
            st.warning("Aucun rapport ne correspond aux critères sélectionnés.")

        st.markdown("<br><hr style='border-color:#E2E8F0;'><p style='font-size:1.2rem;font-weight:700;color:#0F172A;'>📊 Gestion des rapports</p>",unsafe_allow_html=True)
        if not df_rapports.empty:
            col_sc=[c for c in df_rapports.columns if "site" in c.lower()]
            col_cc=[c for c in df_rapports.columns if "ins" in c.lower()]
            if col_sc and col_cc:
                df_s=df_rapports[col_sc[0]].value_counts().reset_index(); df_s.columns=['Site','Nombre']
                df_c=df_rapports[col_cc[0]].value_counts().reset_index(); df_c.columns=['Domaine','Nombre']
                g1,g2=st.columns(2)
                with g1:
                    fig=px.pie(df_s,values='Nombre',names='Site',hole=0.6,color_discrete_sequence=['#1E3A8A','#0EA5E9','#94A3B8'])
                    fig.update_traces(textposition='inside',textinfo='percent+label')
                    fig.update_layout(margin=dict(t=10,b=10,l=10,r=10),height=220,showlegend=False,paper_bgcolor='rgba(0,0,0,0)',plot_bgcolor='rgba(0,0,0,0)')
                    st.plotly_chart(fig,use_container_width=True,config={'displayModeBar':False})
                with g2:
                    fig2=px.bar(df_c.sort_values('Nombre'),x='Nombre',y='Domaine',orientation='h',text='Nombre',color_discrete_sequence=['#1E3A8A'])
                    fig2.update_traces(textposition='outside',cliponaxis=False)
                    fig2.update_layout(margin=dict(t=5,b=5,l=10,r=40),height=220,xaxis_title=None,yaxis_title=None,paper_bgcolor='rgba(0,0,0,0)',plot_bgcolor='rgba(0,0,0,0)',transition_duration=500,transition_easing="cubic-in-out")
                    fig2.update_xaxes(showgrid=True,gridcolor='#E2E8F0')
                    st.plotly_chart(fig2,use_container_width=True,config={'displayModeBar':False})
        if role=="Admin" and password_correct:
            with st.expander("🛠️ Panneau d'administration"):
                st.markdown(f"[Ouvrir le Google Sheets]({URL_GOOGLE_SHEET})")

    # ---- ONGLET 2 : PLANNING ----
    with tab2:
        st.markdown("<br><p style='font-size:1.2rem;font-weight:700;color:#0F172A;'>📅 Prochaines échéances</p>",unsafe_allow_html=True)

        # ---- FILTRES ÉCHÉANCES ----
        with st.container(border=True):
            st.markdown("<p style='font-weight:600;color:#1E293B;margin:0 0 10px 0;font-size:13px;'>🔍 Filtrer les échéances</p>",unsafe_allow_html=True)
            fc1,fc2,fc3=st.columns(3)
            with fc1: f_ech_site=st.selectbox("Site",["Tous","SGB","MEG"],key="f_ech_site")
            with fc2: f_ech_ins =st.selectbox("Installation",["Tous"]+list(PERIODICITE.keys()),key="f_ech_ins")
            with fc3:
                opts_seq=["Tous"]+SOUS_EQUIPEMENTS.get(f_ech_ins,[]) if f_ech_ins!="Tous" else ["Tous"]+[i for sub in SOUS_EQUIPEMENTS.values() for i in sub]
                f_ech_seq=st.selectbox("Sous-équipement",opts_seq,key="f_ech_seq")

        if not df_rapports.empty:
            col_ins_r  =[c for c in df_rapports.columns if "ins" in c.lower()]
            col_date_r =[c for c in df_rapports.columns if "date" in c.lower() and "reelle" not in c.lower() and "réelle" not in c.lower() and "prochaine" not in c.lower()]
            col_site_r =[c for c in df_rapports.columns if "site" in c.lower()]
            col_label_r=[c for c in df_rapports.columns if "equip" in c.lower() or "label" in c.lower() or "nom" in c.lower()]
            col_reelle =[c for c in df_rapports.columns if "reelle" in c.lower() or "réelle" in c.lower()]
            col_prochaine_r=[c for c in df_rapports.columns if "prochaine" in c.lower()]

            if col_ins_r and col_date_r:
                df_ech=df_rapports.copy()
                # Identifiant stable = numéro de ligne réel dans le Sheet (header=ligne1, donc +2)
                df_ech["_ligne_sheet"]=df_ech.index+2
                df_ech["_date_brute"]=pd.to_datetime(df_ech[col_date_r[0]],dayfirst=True,errors='coerce')

                if col_reelle:
                    df_ech["_date_reelle"]=pd.to_datetime(df_ech[col_reelle[0]],dayfirst=True,errors='coerce')
                else:
                    df_ech["_date_reelle"]=pd.NaT

                if col_prochaine_r:
                    df_ech["_prochaine_manuelle"]=pd.to_datetime(df_ech[col_prochaine_r[0]],dayfirst=True,errors='coerce')
                else:
                    df_ech["_prochaine_manuelle"]=pd.NaT

                # Date de dernière visite = date réelle si dispo, sinon date planifiée initiale
                df_ech["_date"]=df_ech["_date_reelle"].combine_first(df_ech["_date_brute"])
                df_ech=df_ech.dropna(subset=["_date"])

                # Déduplication
                cles=[]
                if col_site_r:  cles.append(col_site_r[0])
                cles.append(col_ins_r[0])
                if col_label_r: cles.append(col_label_r[0])
                df_ech=df_ech.sort_values("_date_brute",ascending=True)
                df_ech=df_ech.drop_duplicates(subset=cles,keep="last")

                today_dt=pd.Timestamp.today().normalize()

                def calc_prochaine(row):
                    # Priorité à une échéance saisie manuellement par le responsable,
                    # sinon calcul automatique selon la périodicité de l'installation
                    if pd.notna(row["_prochaine_manuelle"]):
                        return row["_prochaine_manuelle"]
                    mois=PERIODICITE.get(str(row[col_ins_r[0]]).strip(),12)
                    return row["_date"]+pd.DateOffset(months=mois)

                df_ech["Prochaine échéance"]=df_ech.apply(calc_prochaine,axis=1)
                df_ech["Jours restants"]=(df_ech["Prochaine échéance"]-today_dt).dt.days
                df_ech["Statut"]=df_ech["Jours restants"].apply(
                    lambda j:"⚠️ Dépassé" if j<0 else "🔴 Urgent" if j<30 else "🟡 Proche" if j<90 else "🟢 OK")

                cols_affich=[]
                if col_site_r:  cols_affich.append(col_site_r[0])
                if col_label_r: cols_affich.append(col_label_r[0])
                cols_affich+=[col_ins_r[0],"_date_reelle","Prochaine échéance","Jours restants","Statut","_ligne_sheet"]
                df_show=df_ech[cols_affich].sort_values("Prochaine échéance")

                # ---- APPLICATION DES FILTRES ----
                df_show_filtre=df_show.copy()
                if f_ech_site!="Tous" and col_site_r:
                    df_show_filtre=df_show_filtre[df_show_filtre[col_site_r[0]].astype(str).str.strip()==f_ech_site]
                if f_ech_ins!="Tous" and col_ins_r:
                    df_show_filtre=df_show_filtre[df_show_filtre[col_ins_r[0]].astype(str).str.strip()==f_ech_ins]
                if f_ech_seq!="Tous" and col_label_r:
                    df_show_filtre=df_show_filtre[df_show_filtre[col_label_r[0]].astype(str).str.strip().str.contains(f_ech_seq,case=False,na=False)]

                # KPIs statut
                nb_depasse=len(df_show_filtre[df_show_filtre["Statut"]=="⚠️ Dépassé"])
                nb_urgent =len(df_show_filtre[df_show_filtre["Statut"]=="🔴 Urgent"])
                nb_proche =len(df_show_filtre[df_show_filtre["Statut"]=="🟡 Proche"])
                nb_ok     =len(df_show_filtre[df_show_filtre["Statut"]=="🟢 OK"])

                kf1,kf2,kf3,kf4=st.columns(4)
                with kf1:
                    st.markdown(f"""<div style="background:#FEF2F2;padding:10px;border-radius:8px;border-left:3px solid #EF4444;margin-bottom:12px;">
                        <p style="margin:0;font-size:10px;color:#7F1D1D;font-weight:600;text-transform:uppercase;">⚠️ Dépassé</p>
                        <p style="margin:2px 0 0 0;font-size:22px;color:#991B1B;font-weight:700;">{nb_depasse}</p></div>""",unsafe_allow_html=True)
                with kf2:
                    st.markdown(f"""<div style="background:#FFF1F0;padding:10px;border-radius:8px;border-left:3px solid #e34948;margin-bottom:12px;">
                        <p style="margin:0;font-size:10px;color:#7F1D1D;font-weight:600;text-transform:uppercase;">🔴 Urgent</p>
                        <p style="margin:2px 0 0 0;font-size:22px;color:#e34948;font-weight:700;">{nb_urgent}</p></div>""",unsafe_allow_html=True)
                with kf3:
                    st.markdown(f"""<div style="background:#FFFBEB;padding:10px;border-radius:8px;border-left:3px solid #eda100;margin-bottom:12px;">
                        <p style="margin:0;font-size:10px;color:#78350F;font-weight:600;text-transform:uppercase;">🟡 Proche</p>
                        <p style="margin:2px 0 0 0;font-size:22px;color:#eda100;font-weight:700;">{nb_proche}</p></div>""",unsafe_allow_html=True)
                with kf4:
                    st.markdown(f"""<div style="background:#F0FDF4;padding:10px;border-radius:8px;border-left:3px solid #10B981;margin-bottom:12px;">
                        <p style="margin:0;font-size:10px;color:#064E3B;font-weight:600;text-transform:uppercase;">🟢 OK</p>
                        <p style="margin:2px 0 0 0;font-size:22px;color:#10B981;font-weight:700;">{nb_ok}</p></div>""",unsafe_allow_html=True)

                left_col,right_col=st.columns([1.5,1])

                # ---- COLONNE GAUCHE : TABLEAU ----
                with left_col:
                    if role!="Admin" or not password_correct:
                        # Visiteur : lecture seule
                        cols_visiteur=[]
                        if col_site_r:  cols_visiteur.append(col_site_r[0])
                        if col_label_r: cols_visiteur.append(col_label_r[0])
                        cols_visiteur+=[col_ins_r[0],"_date_reelle","Prochaine échéance","Jours restants","Statut"]
                        st.dataframe(df_show_filtre[cols_visiteur],column_config={
                            "_date_reelle":       st.column_config.DateColumn("📅 Date de dernière visite",format="DD/MM/YYYY"),
                            "Prochaine échéance": st.column_config.DateColumn("⏭️ Prochaine échéance",format="DD/MM/YYYY"),
                            "Jours restants":     st.column_config.NumberColumn("Jours restants",format="%d j"),
                        },hide_index=True,use_container_width=True)
                    else:
                        # Admin : édition de la date de dernière visite ET de la prochaine échéance
                        st.markdown("""<div style='background:#EFF6FF;border-left:4px solid #2a78d6;padding:10px 14px;border-radius:6px;margin-bottom:10px;'>
                            <p style='margin:0;font-size:12px;color:#1e40af;font-weight:600;'>✏️ Mode administrateur</p>
                        </div>""",unsafe_allow_html=True)
                        cols_resp=[]
                        if col_site_r:  cols_resp.append(col_site_r[0])
                        if col_label_r: cols_resp.append(col_label_r[0])
                        cols_resp+=[col_ins_r[0],"_date_reelle","Prochaine échéance","Jours restants","Statut","_ligne_sheet"]
                        df_editable=df_show_filtre[cols_resp].copy()
                        df_editable["_date_reelle"]=pd.to_datetime(df_editable["_date_reelle"],errors='coerce')
                        df_editable["Prochaine échéance"]=pd.to_datetime(df_editable["Prochaine échéance"],errors='coerce')
                        edited_df=st.data_editor(df_editable,column_config={
                            "_date_reelle":       st.column_config.DateColumn("✅ Date de dernière visite",format="DD/MM/YYYY",help="Saisissez ici la date réelle du dernier contrôle effectué"),
                            "Prochaine échéance": st.column_config.DateColumn("⏭️ Prochaine échéance",format="DD/MM/YYYY",help="Calculée automatiquement, modifiable si besoin"),
                            "Jours restants":     st.column_config.NumberColumn("Jours restants",format="%d j"),
                            "_ligne_sheet":       None,
                        },disabled=[c for c in df_editable.columns if c not in ("_date_reelle","Prochaine échéance")],hide_index=True,use_container_width=True,key="editor_dates_reelles")

                        if not col_prochaine_r:
                            st.caption("ℹ️ Pour que la « Prochaine échéance » modifiée soit conservée après actualisation, ajoutez une colonne « Prochaine_echeance » dans l'onglet « Rapports » du Google Sheet.")

                        if st.button("💾 Sauvegarder les modifications",type="primary"):
                            with st.spinner("Mise à jour dans Google Sheets..."):
                                nb_maj=0
                                erreurs=[]
                                for idx,row_edit in edited_df.iterrows():
                                    num_ligne_sheet=int(row_edit["_ligne_sheet"])

                                    # -- Date de dernière visite --
                                    nouvelle_date=row_edit["_date_reelle"]
                                    ancienne_date=df_editable.loc[idx,"_date_reelle"]
                                    dates_diff=False
                                    if pd.isna(nouvelle_date) and pd.isna(ancienne_date): dates_diff=False
                                    elif pd.isna(nouvelle_date)!=pd.isna(ancienne_date): dates_diff=True
                                    elif not pd.isna(nouvelle_date) and nouvelle_date!=ancienne_date: dates_diff=True
                                    if dates_diff and not pd.isna(nouvelle_date):
                                        if col_reelle:
                                            num_col=df_rapports.columns.tolist().index(col_reelle[0])+1
                                        else:
                                            num_col=len(df_rapports.columns)+1
                                        lettre_col=chr(64+num_col)
                                        date_str=nouvelle_date.strftime("%d/%m/%Y")
                                        ok, msg = sheets_ecrire_cellule_v2("Rapports",f"{lettre_col}{num_ligne_sheet}",date_str)
                                        if ok:
                                            nb_maj+=1
                                        else:
                                            erreurs.append(f"Ligne {num_ligne_sheet}: {msg}")

                                    # -- Prochaine échéance (surcharge manuelle) --
                                    if col_prochaine_r:
                                        nouvelle_prochaine=row_edit["Prochaine échéance"]
                                        ancienne_prochaine=df_editable.loc[idx,"Prochaine échéance"]
                                        prochaine_diff=False
                                        if pd.isna(nouvelle_prochaine) and pd.isna(ancienne_prochaine): prochaine_diff=False
                                        elif pd.isna(nouvelle_prochaine)!=pd.isna(ancienne_prochaine): prochaine_diff=True
                                        elif not pd.isna(nouvelle_prochaine) and nouvelle_prochaine!=ancienne_prochaine: prochaine_diff=True
                                        if prochaine_diff and not pd.isna(nouvelle_prochaine):
                                            num_col_p=df_rapports.columns.tolist().index(col_prochaine_r[0])+1
                                            lettre_col_p=chr(64+num_col_p)
                                            prochaine_str=nouvelle_prochaine.strftime("%d/%m/%Y")
                                            ok_p, msg_p = sheets_ecrire_cellule_v2("Rapports",f"{lettre_col_p}{num_ligne_sheet}",prochaine_str)
                                            if ok_p:
                                                nb_maj+=1
                                            else:
                                                erreurs.append(f"Ligne {num_ligne_sheet} (prochaine échéance): {msg_p}")
                            if nb_maj>0:
                                st.success(f"✅ {nb_maj} modification(s) enregistrée(s) !")
                            if erreurs:
                                st.error("❌ Erreurs : " + " | ".join(erreurs))
                            if nb_maj>0 or erreurs:
                                st.cache_data.clear()
                                st.rerun()
                            else:
                                st.info("Aucune modification détectée.")

                # ---- COLONNE DROITE : CALENDRIER ----
                with right_col:
                    m_view=st.session_state.cal_mois
                    a_view=st.session_state.cal_annee

                    nav1,nav2,nav3=st.columns([1,3,1])
                    with nav1:
                        if st.button("◀",key="prev_month"):
                            if st.session_state.cal_mois==1: st.session_state.cal_mois=12; st.session_state.cal_annee-=1
                            else: st.session_state.cal_mois-=1
                            st.session_state.jour_selectionne=None; st.rerun()
                    with nav2:
                        st.markdown(f"<p style='text-align:center;font-weight:600;font-size:14px;margin:0;padding-top:4px;'>{MOIS_FR[m_view]} {a_view}</p>",unsafe_allow_html=True)
                    with nav3:
                        if st.button("▶",key="next_month"):
                            if st.session_state.cal_mois==12: st.session_state.cal_mois=1; st.session_state.cal_annee+=1
                            else: st.session_state.cal_mois+=1
                            st.session_state.jour_selectionne=None; st.rerun()

                    m_view=st.session_state.cal_mois
                    a_view=st.session_state.cal_annee

                    # Événements du mois (sur df_ech complet, pas filtré)
                    evenements={}; details_evt={}
                    for _,row in df_ech.iterrows():
                        d=row["Prochaine échéance"]
                        if pd.notna(d) and d.month==m_view and d.year==a_view:
                            j=d.day; ins=str(row[col_ins_r[0]]).strip()
                            col_c=COULEURS_INS.get(ins,"#94a3b8")
                            evenements.setdefault(j,[]).append(col_c)
                            details_evt.setdefault(j,[]).append(row)

                    # En-tête jours
                    jours_abbr=["Lu","Ma","Me","Je","Ve","Sa","Di"]
                    cols_hdr=st.columns(7)
                    for i,j in enumerate(jours_abbr):
                        with cols_hdr[i]:
                            st.markdown(f"<p style='text-align:center;font-size:10px;color:#94a3b8;font-weight:500;margin:0;padding:2px 0;'>{j}</p>",unsafe_allow_html=True)

                    # CSS boutons calendrier
                    st.markdown("""<style>
                        section[data-testid="stSidebar"] ~ div button[kind="secondary"]{min-height:28px!important;height:28px!important;width:28px!important;border-radius:50%!important;font-size:11px!important;padding:0!important;line-height:1!important;}
                    </style>""",unsafe_allow_html=True)

                    today_dt2=datetime.date.today()
                    for semaine in calendar.monthcalendar(a_view,m_view):
                        cols_sem=st.columns(7)
                        for i,jour in enumerate(semaine):
                            with cols_sem[i]:
                                if jour==0:
                                    st.markdown("<div style='height:28px;'></div>",unsafe_allow_html=True)
                                else:
                                    evts    =evenements.get(jour,[])
                                    is_today=(jour==today_dt2.day and m_view==today_dt2.month and a_view==today_dt2.year)
                                    is_sel  =(jour==st.session_state.jour_selectionne)
                                    has_evt =len(evts)>0

                                    if has_evt:
                                        bg=evts[0]
                                        outline="outline:2.5px solid #0F172A;outline-offset:1px;" if is_sel else ""
                                        st.markdown(f"<div style='width:28px;height:28px;border-radius:50%;background:{bg};color:white;font-size:11px;font-weight:700;display:flex;align-items:center;justify-content:center;margin:auto;margin-bottom:-30px;position:relative;z-index:1;pointer-events:none;{outline}'>{jour}</div>",unsafe_allow_html=True)
                                        if st.button("​",key=f"cal_{a_view}_{m_view}_{jour}",help=f"{len(evts)} contrôle(s)"):
                                            st.session_state.jour_selectionne=jour; st.rerun()
                                    elif is_today:
                                        st.markdown(f"<div style='width:28px;height:28px;border-radius:50%;background:#1E3A8A;color:white;font-size:11px;font-weight:600;display:flex;align-items:center;justify-content:center;margin:auto;'>{jour}</div>",unsafe_allow_html=True)
                                    else:
                                        st.markdown(f"<div style='width:28px;height:28px;display:flex;align-items:center;justify-content:center;margin:auto;font-size:11px;color:#334155;'>{jour}</div>",unsafe_allow_html=True)

                    # Légende
                    st.markdown("<div style='margin-top:12px;border-top:1px dashed #E2E8F0;padding-top:8px;'></div>",unsafe_allow_html=True)
                    inss_du_mois={str(r[col_ins_r[0]]).strip() for evts_list in details_evt.values() for r in evts_list}
                    for ins,couleur in COULEURS_INS.items():
                        opacity="1" if ins in inss_du_mois else "0.3"
                        st.markdown(f"""<div style='display:flex;align-items:center;gap:8px;margin-bottom:5px;opacity:{opacity};'>
                            <span style='width:10px;height:10px;border-radius:2px;background:{couleur};display:inline-block;flex-shrink:0;'></span>
                            <span style='font-size:11px;color:#475569;'>{ins}</span>
                        </div>""",unsafe_allow_html=True)

                # ---- DÉTAIL JOUR SÉLECTIONNÉ ----
                jour_sel=st.session_state.jour_selectionne
                if jour_sel and jour_sel in details_evt:
                    nb_c=len(details_evt[jour_sel])
                    st.markdown(f"""<div style='margin-top:20px;padding:12px 16px;background:#F0F9FF;border-left:4px solid #0EA5E9;border-radius:8px;margin-bottom:12px;'>
                        <p style='margin:0;font-size:14px;font-weight:600;color:#0C4A6E;'>📋 {nb_c} contrôle(s) planifié(s) le {jour_sel} {MOIS_FR[m_view]} {a_view}</p>
                    </div>""",unsafe_allow_html=True)
                    nb_cols=min(nb_c,4)
                    card_cols=st.columns(nb_cols)
                    for idx,row_ctrl in enumerate(details_evt[jour_sel]):
                        with card_cols[idx%nb_cols]:
                            c_ins  =str(row_ctrl[col_ins_r[0]]).strip()
                            c_site =str(row_ctrl[col_site_r[0]]).strip()  if col_site_r  else ""
                            c_label=str(row_ctrl[col_label_r[0]]).strip() if col_label_r else ""
                            c_date =row_ctrl["_date_brute"]
                            c_reel =row_ctrl["_date_reelle"]
                            c_next =row_ctrl["Prochaine échéance"]
                            c_jours=int(row_ctrl["Jours restants"])
                            c_stat =row_ctrl["Statut"]
                            c_col  =COULEURS_INS.get(c_ins,"#94a3b8")
                            date_fmt=c_date.strftime("%d/%m/%Y") if pd.notna(c_date) else "—"
                            reel_fmt=c_reel.strftime("%d/%m/%Y") if pd.notna(c_reel) else None
                            next_fmt=c_next.strftime("%d/%m/%Y") if pd.notna(c_next) else "—"
                            j_txt  =f"⚠️ {abs(c_jours)}j de retard" if c_jours<0 else f"Dans {c_jours} j"
                            if reel_fmt:
                                date_ctrl_html=(
                                    "<p style='margin:0 0 2px 0;font-size:10px;color:#94a3b8;'>Date réelle visite</p>"
                                    f"<p style='margin:0 0 6px 0;font-size:11px;color:#059669;font-weight:600;'>✅ {reel_fmt}</p>"
                                    "<p style='margin:0 0 2px 0;font-size:10px;color:#94a3b8;'>Date planifiée initiale</p>"
                                    f"<p style='margin:0 0 6px 0;font-size:11px;color:#94a3b8;text-decoration:line-through;'>{date_fmt}</p>"
                                )
                            else:
                                date_ctrl_html=(
                                    "<p style='margin:0 0 2px 0;font-size:10px;color:#94a3b8;'>Date planifiée</p>"
                                    f"<p style='margin:0 0 6px 0;font-size:11px;color:#334155;font-weight:500;'>{date_fmt}</p>"
                                )
                            label_html = ("<p style='margin:0 0 4px 0;font-size:11px;color:#64748B;'>⚙️ "+c_label+"</p>") if c_label else ""
                            carte_html=(
                                f"<div style='background:white;border-top:4px solid {c_col};padding:14px;border-radius:8px;box-shadow:0 2px 8px rgba(0,0,0,0.06);margin-bottom:8px;'>"
                                f"<p style='margin:0 0 8px 0;font-size:12px;font-weight:700;color:#1E293B;'>{c_ins}</p>"
                                f"<p style='margin:0 0 4px 0;font-size:11px;color:#475569;'>🏢 <b>{c_site}</b></p>"
                                f"{label_html}"
                                "<hr style='border:none;border-top:1px solid #F1F5F9;margin:8px 0;'>"
                                f"{date_ctrl_html}"
                                "<p style='margin:0 0 2px 0;font-size:10px;color:#94a3b8;'>Prochaine échéance</p>"
                                f"<p style='margin:0 0 6px 0;font-size:11px;color:#334155;font-weight:500;'>{next_fmt}</p>"
                                f"<span style='display:inline-block;padding:3px 8px;border-radius:4px;font-size:10px;font-weight:600;background:{c_col}22;color:{c_col};'>{c_stat} — {j_txt}</span>"
                                "</div>"
                            )
                            st.markdown(carte_html, unsafe_allow_html=True)
                elif evenements and jour_sel is None:
                    st.info("💡 Cliquez sur un jour coloré du calendrier pour voir les détails du contrôle.")


# ------------------------------------------------------------------------------
# BLOC 3 — UI Streamlit (à insérer dans `with tab_exigences:`)
# ------------------------------------------------------------------------------
        st.divider()
        st.markdown("### 🗓️ Calendrier de contrôle réglementaire")


        annee_ref_calendrier = datetime.date.today().year

        if st.button("🗓️ Générer le calendrier", use_container_width=True, key="btn_gen_calendrier"):
            with st.spinner("Construction du calendrier..."):
                df_calendrier = construire_calendrier_controle(df_rapports, annee_ref_calendrier)
                if df_calendrier.empty:
                    st.warning("Aucune donnée exploitable pour construire le calendrier.")
                    st.session_state["df_calendrier"] = None
                else:
                    st.session_state["df_calendrier"] = df_calendrier

        if st.session_state.get("df_calendrier") is not None:
            df_cal = st.session_state["df_calendrier"]
            st.dataframe(df_cal, hide_index=True, use_container_width=True)

            # ----- Taux de réalisation 2026 : barres horizontales par site + infographie globale -----
            taux = calculer_taux_realisation(df_cal)
            taux_meg = taux.get("MEG", 0.0)
            taux_sgb = taux.get("SGB", 0.0)
            taux_global = taux.get("Global", 0.0)

            st.markdown(f"#### 📊 Taux de réalisation des contrôles — {annee_ref_calendrier}")
            col_bars, col_gauge = st.columns([2, 1])

            with col_bars:
                df_bars = pd.DataFrame({"Site": ["MEG", "SGB"], "Taux (%)": [taux_meg, taux_sgb]})
                fig_bars = px.bar(
                    df_bars, x="Taux (%)", y="Site", text="Taux (%)", color="Site",
                    orientation="h",
                    color_discrete_map={"MEG": "#2563EB", "SGB": "#059669"},
                    range_x=[0, 100],
                )
                fig_bars.update_traces(texttemplate='%{text}%', textposition='outside', cliponaxis=False)
                fig_bars.update_layout(
                    title=f"Taux de réalisation par site — {annee_ref_calendrier}", title_x=0.5,
                    showlegend=False, xaxis_title="Taux (%)", yaxis_title="",
                    margin=dict(t=40, b=10, l=10, r=30), height=280,
                    paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)',
                )
                st.plotly_chart(fig_bars, use_container_width=True, config={'displayModeBar': False})

            with col_gauge:
                fig_gauge = go.Figure(go.Indicator(
                    mode="gauge+number",
                    value=taux_global,
                    number={'suffix': "%", 'font': {'size': 100, 'color': "#0F172A"}},
                    title={'text': f"Taux global {annee_ref_calendrier}", 'font': {'size': 14, 'color': "#334155"}},
                    gauge={
                        'axis': {'range': [0, 100], 'tickwidth': 1, 'tickcolor': "#94A3B8"},
                        'bar': {'color': "#1E3A8A", 'thickness': 0.8},
                        'bgcolor': "white",
                        'borderwidth': 0,
                        'steps': [{'range': [0, 100], 'color': "#E2E8F0"}],
                    },
                ))
                fig_gauge.update_layout(
                    margin=dict(t=40, b=10, l=20, r=20), height=280,
                    paper_bgcolor='rgba(0,0,0,0)',
                    transition_duration=600, transition_easing="cubic-in-out",
                )
                st.plotly_chart(fig_gauge, use_container_width=True, config={'displayModeBar': False})

            dl1, dl2 = st.columns(2)
            with dl1:
                pdf_cal = generer_calendrier_controle_pdf(df_cal, annee_ref_calendrier)
                st.download_button(
                    "📥 Télécharger en PDF", data=pdf_cal,
                    file_name=f"Calendrier_Controle_{annee_ref_calendrier}_{datetime.date.today().strftime('%d_%m_%Y')}.pdf",
                    mime="application/pdf", use_container_width=True, key="dl_calendrier_pdf",
                )
            with dl2:
                excel_cal = generer_calendrier_controle_excel(df_cal, annee_ref_calendrier)
                st.download_button(
                    "📥 Télécharger en Excel", data=excel_cal,
                    file_name=f"Calendrier_Controle_{annee_ref_calendrier}_{datetime.date.today().strftime('%d_%m_%Y')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True, key="dl_calendrier_excel",
                )
        # ---- Relance des échéances proches (notification manuelle par e-mail) ----
        # Réservé au profil Admin uniquement.
            if role == "Admin" and password_correct:
                st.markdown("<br><hr style='border-color:#E2E8F0;'>",unsafe_allow_html=True)
                st.markdown("<br>",unsafe_allow_html=True)
                st.markdown("### 📧 Relance des échéances")
                st.caption("Envoie un e-mail récapitulatif des contrôles en retard ou à venir sous 30 jours. ")
                df_calendrier_relance = construire_calendrier_controle(df_rapports)
                echeances_proches = extraire_echeances_proches(df_calendrier_relance, jours_horizon=30)
                if echeances_proches:
                    st.warning(f"⚠️ {len(echeances_proches)} échéance(s) en retard ou à venir sous 30 jours.")
                else:
                    st.success("✅ Aucune échéance en retard ou à venir sous 30 jours.")
                destinataires_defaut = ""
                try:
                    destinataires_defaut = ", ".join(st.secrets["notifications"]["destinataires"])
                except Exception:
                    pass
                destinataires_saisis = st.text_input("Destinataires (séparés par des virgules) :",
                                                       value=destinataires_defaut, placeholder="a@domaine.com, b@domaine.com")
                if st.button("📧 Envoyer la relance maintenant", use_container_width=True, disabled=not echeances_proches):
                    emails = [e.strip() for e in destinataires_saisis.split(",") if format_email_valide(e.strip())]
                    if not emails:
                        st.error("Aucune adresse e-mail valide renseignée.")
                    else:
                        corps = construire_message_relance(echeances_proches)
                        nb_ok, nb_ko = 0, 0
                        for dest in emails:
                            ok_env, err_env = envoyer_email(dest, "Relance — Echéances de contrôle réglementaire", corps)
                            nb_ok += 1 if ok_env else 0
                            nb_ko += 0 if ok_env else 1
                        if nb_ok:
                            st.success(f"✅ Relance envoyée à {nb_ok} destinataire(s).")
                            journaliser_action(utilisateur_courant(), "Relance échéances envoyée", f"{nb_ok} destinataire(s)")
                        if nb_ko:
                            st.error(f"❌ Échec d'envoi pour {nb_ko} destinataire(s) (vérifier la configuration SMTP).")


    # ---- ONGLET EXIGENCES ----
    with tab_exigences:
        st.markdown("<p style='font-size:1.2rem;font-weight:700;color:#0F172A;margin-bottom:15px;'>📌 Exigences réglementaires</p>", unsafe_allow_html=True)

        df_exig = lire_exigences()


    # ===== SECTION 1 : CONTRAT D'ABONNEMENT =====
        st.markdown("### 📄 Contrat d'abonnement 2026")

        def lien_telechargement_direct(lien: str) -> str:
            """Convertit un lien Google Drive (vue/partage) en lien de téléchargement direct.
            Si ce n'est pas un lien Google Drive reconnu, renvoie le lien tel quel."""
            if not lien:
                return lien
            m = re.search(r"/file/d/([a-zA-Z0-9_-]+)", lien)
            if not m:
                m = re.search(r"[?&]id=([a-zA-Z0-9_-]+)", lien)
            if m:
                file_id = m.group(1)
                return f"https://drive.google.com/uc?export=download&id={file_id}"
            return lien

        def lien_apercu_drive(lien: str) -> str:
            """Convertit un lien Google Drive en lien d'aperçu (preview) qui s'ouvre
            directement dans Drive sans télécharger le fichier."""
            if not lien:
                return lien
            m = re.search(r"/file/d/([a-zA-Z0-9_-]+)", lien)
            if not m:
                m = re.search(r"[?&]id=([a-zA-Z0-9_-]+)", lien)
            if m:
                file_id = m.group(1)
                return f"https://drive.google.com/file/d/{file_id}/preview"
            return lien

        lien_contrat = ""
        if not df_exig.empty and "Type" in df_exig.columns:
            ligne_c = df_exig[df_exig["Type"] == "Contrat"]
            if not ligne_c.empty:
                lien_contrat = str(ligne_c.iloc[0].get("Lien_PDF", "")).strip()

        col_contrat, col_action = st.columns([5, 1])
        with col_contrat:
            if lien_contrat and lien_contrat.lower() != "nan":
                lien_dl = lien_telechargement_direct(lien_contrat)
                lien_apercu = lien_apercu_drive(lien_contrat)

                st.markdown(
                    "<div style='background:white;padding:16px 20px;border-radius:10px;"
                    "box-shadow:0 2px 8px rgba(0,0,0,0.05);border-left:4px solid #1E3A8A;"
                    "display:flex;align-items:center;justify-content:space-between;'>"
                    "<span style='font-size:14px;font-weight:600;color:#1E293B;'>📑 Contrat d'abonnement 2026</span>"
                    "<span>"
                    f"<a href='{lien_apercu}' target='_blank' rel='noopener' style='text-decoration:none;background:#1E3A8A;"
                    "color:white;padding:8px 14px;border-radius:6px;font-size:13px;font-weight:600;margin-right:8px;'>"
                    "👁️ Consulter</a>"
                    f"<a href='{lien_dl}' download target='_blank' rel='noopener' style='text-decoration:none;background:#16A34A;"
                    "color:white;padding:8px 14px;border-radius:6px;font-size:13px;font-weight:600;'>"
                    "📥 Télécharger</a>"
                    "</span>"
                    "</div>", unsafe_allow_html=True)
            else:
                st.info("Aucun contrat n'a encore été ajouté.")

        if role == "Admin" and password_correct:
            with st.expander("✏️ Gérer le contrat"):
                nouveau_lien = st.text_input("Lien Google Drive du contrat PDF :",
                    value=lien_contrat if lien_contrat.lower() != "nan" else "",
                    placeholder="https://drive.google.com/file/d/...")
                bc1, bc2 = st.columns(2)
                with bc1:
                    if st.button("💾", use_container_width=True):
                        if nouveau_lien.strip():
                            ok, err = ecrire_contrat(nouveau_lien.strip())
                            if ok:
                                st.success("✅ Contrat mis à jour !")
                                st.rerun()
                            else:
                                st.error(f"Erreur : {err}")
                        else:
                            st.warning("Veuillez coller un lien.")
                with bc2:
                    if st.button("🗑️", use_container_width=True):
                        if supprimer_contrat():
                            st.success("✅ Contrat supprimé.")
                            st.rerun()
                        else:
                            st.error("Erreur lors de la suppression.")

        st.markdown("<br><hr style='border-color:#E2E8F0;'>", unsafe_allow_html=True)     
        

    # ===== SECTION 3 : LISTE DES ÉQUIPEMENTS (ARBORESCENCE) =====
        st.markdown("### 🏭 Liste des équipements soumis au contrôle")

        df_equip = pd.DataFrame()
        if not df_exig.empty and "Type" in df_exig.columns:
            df_equip = df_exig[df_exig["Type"] == "Equipement"].copy()
            if "Nombre" in df_equip.columns:
                df_equip["Nombre"] = pd.to_numeric(df_equip["Nombre"], errors="coerce").fillna(0).astype(int)

    # Initialisation propre du Session State
        if "site_exig_sel" not in st.session_state: 
            st.session_state.site_exig_sel = None
        if "ins_exig_sel" not in st.session_state: 
            st.session_state.ins_exig_sel = None

    # Niveau 1 : choix du site
        st.markdown("<p style='font-size:13px;color:#64748B;font-weight:600;margin-bottom:8px;'>Sélectionnez un site :</p>", unsafe_allow_html=True)
        s1, s2, s3 = st.columns([1, 1, 3])
    
        with s1:
            actif_sgb = (st.session_state.site_exig_sel == "SGB")
            if st.button("🏢 SGB", use_container_width=True, type="primary" if actif_sgb else "secondary"):
                st.session_state.site_exig_sel = "SGB"
                st.session_state.ins_exig_sel = None  # Reset l'installation si on change de site
                st.rerun()
            
        with s2:
            actif_meg = (st.session_state.site_exig_sel == "MEG")
            if st.button("🏢 MEG", use_container_width=True, type="primary" if actif_meg else "secondary"):
                st.session_state.site_exig_sel = "MEG"
                st.session_state.ins_exig_sel = None  # Reset l'installation si on change de site
                st.rerun()

    # --- CORRECTION DE LA LOGIQUE D'AFFICHAGE ---
    # On se base TOUJOURS sur le session_state actuel, pas sur le clic du bouton direct
        if st.session_state.site_exig_sel:
            site_sel = st.session_state.site_exig_sel
            st.markdown(f"<p style='font-size:13px;color:#64748B;font-weight:600;margin:16px 0 8px 0;'>Installations — Site {site_sel} :</p>", unsafe_allow_html=True)

            df_site = df_equip[df_equip["Site"] == site_sel] if not df_equip.empty else pd.DataFrame()

            NOMS_COURTS_INS = {
                "Installations électriques": "⚡ Électriques",
                "Equipements de levage":     "🏗️ Levage",
                "Sécurité incendie":         "🔥 Incendie",
                "Installations de gaz":      "🔵 Gaz",
                "Appareil pression de gaz":  "⚙️ Pression gaz",
            }

        # Création dynamique des boutons des installations
            ins_cols = st.columns(5)
            for i, (ins, couleur) in enumerate(COULEURS_INS.items()):
                with ins_cols[i % 5]:
                    nb_total_ins = int(df_site[df_site["Installation"] == ins]["Nombre"].sum()) if not df_site.empty else 0
                    actif_ins = (st.session_state.ins_exig_sel == ins)
                    label_court = NOMS_COURTS_INS.get(ins, ins)
                
                    if st.button(f"{label_court} ({nb_total_ins})", key=f"ins_btn_{ins}", use_container_width=True,
                                 type="primary" if actif_ins else "secondary",
                                 help=f"{nb_total_ins} équipement(s) au total"):
                        st.session_state.ins_exig_sel = ins
                        st.rerun()

        # Niveau 3 : sous-équipements de l'istallation choisie
            if st.session_state.ins_exig_sel:
                ins_sel = st.session_state.ins_exig_sel
                st.markdown(f"<p style='font-size:13px;color:#64748B;font-weight:600;margin:16px 0 8px 0;'>Sous-équipements — {ins_sel} ({site_sel}) :</p>", unsafe_allow_html=True)

                df_ins = df_site[df_site["Installation"] == ins_sel] if not df_site.empty else pd.DataFrame()
                couleur_ins = COULEURS_INS.get(ins_sel, "#94a3b8")

                if df_ins.empty:
                    st.info(f"Aucun sous-équipement enregistré pour {ins_sel} sur le site {site_sel}.")
                else:
                    eq_cols = st.columns(3)
                    for idx, (_, row_eq) in enumerate(df_ins.iterrows()):
                        with eq_cols[idx % 3]:
                            st.markdown(
                                f"<div style='background:white;border-left:4px solid {couleur_ins};"
                                "padding:14px;border-radius:8px;box-shadow:0 2px 6px rgba(0,0,0,0.05);margin-bottom:10px;'>"
                                f"<p style='margin:0;font-size:13px;font-weight:600;color:#1E293B;'>{row_eq.get('Sous_equipement','')}</p>"
                                f"<p style='margin:6px 0 0 0;font-size:24px;font-weight:800;color:{couleur_ins};'>{int(row_eq.get('Nombre',0))}</p>"
                                "</div>", unsafe_allow_html=True)

            # Gestion (ajout/suppression) — admin uniquement
                if role == "Admin" and password_correct:
                    with st.expander("✏️ Gérer les sous-équipements"):
                        st.markdown("**Ajouter un sous-équipement :**")
                        ac1, ac2, ac3 = st.columns([2, 1, 1])
                        with ac1:
                            nouv_seq = st.text_input("Nom du sous-équipement", key="nouv_seq_nom")
                        with ac2:
                            nouv_nb = st.number_input("Nombre", min_value=1, value=1, key="nouv_seq_nb")
                        with ac3:
                            st.write("")
                            st.write("")
                            if st.button("➕ Ajouter", use_container_width=True):
                                if nouv_seq.strip():
                                    ok, err = ajouter_equipement(site_sel, ins_sel, nouv_seq.strip(), nouv_nb)
                                    if ok:
                                        st.success("✅ Ajouté !")
                                        st.rerun()
                                    else:
                                        st.error(f"Erreur : {err}")
                                else:
                                    st.warning("Veuillez saisir un nom.")

                        if not df_ins.empty:
                            st.markdown("<br>**Supprimer un sous-équipement :**", unsafe_allow_html=True)
                            for orig_idx, row_eq in df_ins.iterrows():
                                dc1, dc2 = st.columns([5, 1])
                                with dc1:
                                    st.write(f"{row_eq.get('Sous_equipement','')} — {int(row_eq.get('Nombre',0))} unité(s)")
                                with dc2:
                                    if st.button("🗑️", key=f"del_eq_{orig_idx}"):
                                        num_ligne_sheet = orig_idx + 2
                                        if supprimer_equipement_ligne(num_ligne_sheet):
                                            st.success("Supprimé !")
                                            st.cache_data.clear()
                                            st.rerun()
                                        else:
                                            st.error("Erreur lors de la suppression.")
        else:
            st.info("👆 Sélectionnez un site (SGB ou MEG) pour voir les installations")

        st.divider()

        if not df_exig.empty:
            st.markdown("### 📄 Check-lists des équipements contractés")
            col_sgb, col_meg = st.columns(2)
            date_str = datetime.date.today().strftime('%d_%m_%Y')

            with col_sgb:
                st.markdown(
                    "<div style='background:white;padding:16px 20px;border-radius:10px;"
                    "box-shadow:0 2px 8px rgba(0,0,0,0.05);border-left:4px solid #1E3A8A;"
                    "margin-bottom:10px;'>"
                    "<span style='font-size:14px;font-weight:600;color:#1E293B;'>📑 Rapport d'inspection — SGB</span>"
                    "</div>", unsafe_allow_html=True)
                if st.button("👁️ Consulter le rapport", use_container_width=True, key="consult_sgb", type="primary"):
                    with st.spinner("Préparation du rapport SGB..."):
                        try:
                            st.session_state["pdf_sgb"] = generer_rapport_equipements_pdf(df_exig, "SGB")
                        except Exception as e:
                            st.session_state["pdf_sgb"] = None
                            st.error(f"Erreur PDF SGB : {e}")
                if st.session_state.get("pdf_sgb"):
                    afficher_apercu_pdf(st.session_state["pdf_sgb"])
                    st.download_button(
                        label="📥 Télécharger le rapport SGB",
                        data=st.session_state["pdf_sgb"],
                        file_name=f"Rapport_Inspection_SGB_{date_str}.pdf",
                        mime="application/pdf",
                        use_container_width=True,
                        key="dl_sgb"
                    )

            with col_meg:
                st.markdown(
                    "<div style='background:white;padding:16px 20px;border-radius:10px;"
                    "box-shadow:0 2px 8px rgba(0,0,0,0.05);border-left:4px solid #1E3A8A;"
                    "margin-bottom:10px;'>"
                    "<span style='font-size:14px;font-weight:600;color:#1E293B;'>📑 Rapport d'inspection — MEG</span>"
                    "</div>", unsafe_allow_html=True)
                if st.button("👁️ Consulter le rapport", use_container_width=True, key="consult_meg", type="primary"):
                    with st.spinner("Préparation du rapport MEG..."):
                        try:
                            st.session_state["pdf_meg"] = generer_rapport_equipements_pdf(df_exig, "MEG")
                        except Exception as e:
                            st.session_state["pdf_meg"] = None
                            st.error(f"Erreur PDF MEG : {e}")
                if st.session_state.get("pdf_meg"):
                    afficher_apercu_pdf(st.session_state["pdf_meg"])
                    st.download_button(
                        label="📥 Télécharger le rapport MEG",
                        data=st.session_state["pdf_meg"],
                        file_name=f"Rapport_Inspection_MEG_{date_str}.pdf",
                        mime="application/pdf",
                        use_container_width=True,
                        key="dl_meg"
                    )
                    


    

    
    # ---- ONGLET 3 : PRÉSENCE & VISITES ----
    if tab3 and role=="Admin" and password_correct:
        with tab3:
            st.markdown("<p style='font-size:1.2rem;font-weight:700;color:#1E3A8A;'>👥 Suivi des visiteurs</p>",unsafe_allow_html=True)
            col_r,_=st.columns([1,5])
            with col_r:
                if st.button("🔄"): st.rerun()
            st.markdown("### 🟢 Présence en temps réel")
            with st.spinner("Chargement..."):
                df_presence=lire_presence()
            if df_presence.empty:
                st.info("Aucun visiteur enregistré.")
            else:
                nb_en_ligne=len(df_presence[df_presence["Statut"].str.contains("🟢")])
                nb_recent  =len(df_presence[df_presence["Statut"].str.contains("🟡")])
                nb_offline =len(df_presence[df_presence["Statut"].str.contains("🔴")])
                p1,p2,p3=st.columns(3)
                with p1:
                    st.markdown(f"""<div style="background:#F0FDF4;padding:16px;border-radius:10px;border-left:4px solid #10B981;margin-bottom:16px;">
                        <p style="margin:0;font-size:11px;color:#064E3B;font-weight:700;text-transform:uppercase;">🟢 En ligne</p>
                        <p style="margin:4px 0 0 0;font-size:32px;color:#065F46;font-weight:800;">{nb_en_ligne}</p></div>""",unsafe_allow_html=True)
                with p2:
                    st.markdown(f"""<div style="background:#FFFBEB;padding:16px;border-radius:10px;border-left:4px solid #F59E0B;margin-bottom:16px;">
                        <p style="margin:0;font-size:11px;color:#78350F;font-weight:700;text-transform:uppercase;">🟡 Récemment actif</p>
                        <p style="margin:4px 0 0 0;font-size:32px;color:#92400E;font-weight:800;">{nb_recent}</p></div>""",unsafe_allow_html=True)
                with p3:
                    st.markdown(f"""<div style="background:#FEF2F2;padding:16px;border-radius:10px;border-left:4px solid #EF4444;margin-bottom:16px;">
                        <p style="margin:0;font-size:11px;color:#7F1D1D;font-weight:700;text-transform:uppercase;">🔴 Hors ligne</p>
                        <p style="margin:4px 0 0 0;font-size:32px;color:#991B1B;font-weight:800;">{nb_offline}</p></div>""",unsafe_allow_html=True)
                st.dataframe(df_presence,column_config={
                    "Email":st.column_config.TextColumn("📧 Visiteur"),
                    "Dernière activité":st.column_config.TextColumn("🕐 Dernière activité"),
                    "Statut":st.column_config.TextColumn("Statut"),
                    "Activité":st.column_config.TextColumn("⏱️ Détail")},
                    hide_index=True,use_container_width=True)

            st.markdown("<br>",unsafe_allow_html=True)
            st.markdown("### 📋 Historique complet des accès")
            with st.spinner("Chargement des logs..."):
                df_logs=lire_logs()
            if df_logs.empty:
                st.info("Aucun log enregistré.")
            else:
                nb_total=len(df_logs)
                col_em=[c for c in df_logs.columns if "email" in c.lower() or "mail" in c.lower()]
                nb_uniq=df_logs[col_em[0]].nunique() if col_em else 0
                l1,l2=st.columns(2)
                with l1:
                    st.markdown(f"""<div style="background:white;padding:16px;border-radius:10px;box-shadow:0 2px 6px rgba(0,0,0,0.05);border-left:4px solid #1E3A8A;margin-bottom:16px;">
                        <p style="margin:0;font-size:11px;color:#64748B;font-weight:600;text-transform:uppercase;">Total visites</p>
                        <p style="margin:4px 0 0 0;font-size:28px;color:#0F172A;font-weight:700;">{nb_total}</p></div>""",unsafe_allow_html=True)
                with l2:
                    st.markdown(f"""<div style="background:white;padding:16px;border-radius:10px;box-shadow:0 2px 6px rgba(0,0,0,0.05);border-left:4px solid #0EA5E9;margin-bottom:16px;">
                        <p style="margin:0;font-size:11px;color:#64748B;font-weight:600;text-transform:uppercase;">Visiteurs uniques</p>
                        <p style="margin:4px 0 0 0;font-size:28px;color:#0F172A;font-weight:700;">{nb_uniq}</p></div>""",unsafe_allow_html=True)
                st.dataframe(df_logs,column_config={
                    "Date":st.column_config.TextColumn("📅 Date & Heure"),
                    "Email":st.column_config.TextColumn("📧 E-mail")},
                    hide_index=True,use_container_width=True)

            # ---- Journal d'audit (traçabilité des actions de modification) ----
            st.markdown("<br><hr style='border-color:#E2E8F0;'>",unsafe_allow_html=True)
            st.markdown("<br>",unsafe_allow_html=True)
            st.markdown("### 🧾 Journal d'audit")
            st.caption("Historique des créations, modifications, suppressions effectuées dans l'application.")
            with st.spinner("Chargement du journal d'audit..."):
                df_audit = lire_audit_log()
            if df_audit.empty:
                st.info("Aucune action journalisée.")
            else:
                st.dataframe(df_audit.sort_index(ascending=False), hide_index=True, use_container_width=True)

            

            
    # ---- ONGLET 4 : KPI (Admin uniquement — vue complète) ----
    if tab_kpi and role=="Admin" and password_correct:
        with tab_kpi:
            st.markdown("<p style='font-size:1.2rem;font-weight:700;color:#1E3A8A;'>📊 Indicateurs de performance</p>",unsafe_allow_html=True)
            col_r_kpi,_=st.columns([1,5])
            with col_r_kpi:
                if st.button("🔄",key="refresh_kpi"): st.cache_data.clear(); st.rerun()

            # ---- Préparation des données de contrôle (même logique que l'onglet Planification) ----
            col_ins_k   = [c for c in df_rapports.columns if "ins" in c.lower()]
            col_date_k  = [c for c in df_rapports.columns if "date" in c.lower() and "reelle" not in c.lower() and "réelle" not in c.lower() and "prochaine" not in c.lower() and "planifi" not in c.lower()]
            col_site_k  = [c for c in df_rapports.columns if "site" in c.lower()]
            col_label_k = [c for c in df_rapports.columns if "equip" in c.lower() or "label" in c.lower() or "nom" in c.lower()]
            col_reelle_k= [c for c in df_rapports.columns if "reelle" in c.lower() or "réelle" in c.lower()]
            col_planifiee_k = [c for c in df_rapports.columns if "planifi" in c.lower()]

            if df_rapports.empty or not col_ins_k or not col_date_k:
                st.info("Données insuffisantes dans l'onglet « Rapports » pour calculer les KPI.")
                kpi_data = None
            else:
                df_k = df_rapports.copy()
                df_k["_date_brute"]  = pd.to_datetime(df_k[col_date_k[0]], dayfirst=True, errors='coerce')
                df_k["_date_reelle"] = pd.to_datetime(df_k[col_reelle_k[0]], dayfirst=True, errors='coerce') if col_reelle_k else pd.NaT
                # Date planifiée de référence pour le respect de délai : si la colonne dédiée existe
                # dans le Sheet on l'utilise, sinon (tant qu'elle n'a pas été ajoutée) on considère
                # provisoirement que la date planifiée = la date de dernière visite (donc toujours respecté).
                if col_planifiee_k:
                    df_k["_date_planifiee"] = pd.to_datetime(df_k[col_planifiee_k[0]], dayfirst=True, errors='coerce')
                else:
                    df_k["_date_planifiee"] = df_k["_date_reelle"]
                df_k = df_k.dropna(subset=["_date_brute"])

                cles_k=[]
                if col_site_k:  cles_k.append(col_site_k[0])
                cles_k.append(col_ins_k[0])
                if col_label_k: cles_k.append(col_label_k[0])
                df_k = df_k.sort_values("_date_brute", ascending=True)
                df_k = df_k.drop_duplicates(subset=cles_k, keep="last")

                # ---- KPI 1 : Taux de réalisation 2026, calculé PAR INSTALLATION (et non par équipement) ----
                # Contrôles attendus en 2026 selon la périodicité de chaque installation :
                #   - Installations électriques : périodicité 6 mois -> 2 contrôles/an
                #   - Les 4 autres types        : périodicité 12 mois -> 1 contrôle/an
                # => 2 sites (SGB, MEG) x (1 élec x2 + 4 autres x1) = 12 contrôles attendus au total sur l'année

                SITES_SUIVIS = ["SGB", "MEG"]
                INSTALLATIONS_SUIVIES = list(PERIODICITE.keys())

                def nb_campagnes_attendues(installation):
                    return round(12 / PERIODICITE.get(installation, 12))

                nb_total_2026 = sum(nb_campagnes_attendues(ins) for ins in INSTALLATIONS_SUIVIES) * len(SITES_SUIVIS)

                # Un contrôle est compté comme réalisé en 2026 dès lors que sa DATE RÉELLE de visite
                # tombe en 2026, quelle que soit l'année de l'échéance théorique associée.
                df_realises_2026 = df_k[df_k["_date_reelle"].notna() & (df_k["_date_reelle"].dt.year == 2026)].copy()
                if col_site_k:
                    df_realises_2026 = df_realises_2026[df_realises_2026[col_site_k[0]].astype(str).str.strip().isin(SITES_SUIVIS)]

                nb_realises_2026 = 0
                for site in SITES_SUIVIS:
                    for ins in INSTALLATIONS_SUIVIES:
                        attendu = nb_campagnes_attendues(ins)
                        df_grp = df_realises_2026[df_realises_2026[col_ins_k[0]].astype(str).str.strip() == ins]
                        if col_site_k:
                            df_grp = df_grp[df_grp[col_site_k[0]].astype(str).str.strip() == site]
                        # Une "campagne" réalisée = une date réelle distincte (mois/échéance) pour cette
                        # installation sur ce site, plafonnée au nombre de contrôles attendus par an.
                        nb_campagnes_realisees = df_grp["_date_brute"].nunique() if not df_grp.empty else 0
                        nb_realises_2026 += min(nb_campagnes_realisees, attendu)

                nb_restants_2026 = nb_total_2026 - nb_realises_2026
                taux1 = round(nb_realises_2026/nb_total_2026*100,1) if nb_total_2026>0 else 0
                


                # ---- KPI 2 : Taux de respect de délai de visite (écart ≤ 1 mois entre date planifiée et date de dernière visite) ----
                # On ne considère que les contrôles dont la visite a réellement eu lieu en 2026
                # (certains contrôles réalisés figurent avec une date réelle en 2025 et ne doivent pas être comptés ici).
                df_realises_k = df_k[df_k["_date_reelle"].notna() & (df_k["_date_reelle"].dt.year == 2026)].copy()
                nb_visites_realisees = len(df_realises_k)
                if nb_visites_realisees > 0:
                    df_realises_k["_ecart"] = (df_realises_k["_date_reelle"] - df_realises_k["_date_planifiee"]).dt.days.abs()
                    nb_respectes = int((df_realises_k["_ecart"] <= 31).sum())
                else:
                    nb_respectes = 0
                nb_non_respectes = nb_visites_realisees - nb_respectes
                taux2 = round(nb_respectes/nb_visites_realisees*100,1) if nb_visites_realisees>0 else 0

                # ---- Sous-totaux PAR SITE, pour le dashboard comparatif SGB vs MEG ----
                comparatif_sites = {}
                for site_c in SITES_SUIVIS:
                    total_site_c = sum(nb_campagnes_attendues(ins) for ins in INSTALLATIONS_SUIVIES)
                    realises_site_c = 0
                    for ins in INSTALLATIONS_SUIVIES:
                        attendu_c = nb_campagnes_attendues(ins)
                        df_grp_c = df_realises_2026[df_realises_2026[col_ins_k[0]].astype(str).str.strip() == ins]
                        if col_site_k:
                            df_grp_c = df_grp_c[df_grp_c[col_site_k[0]].astype(str).str.strip() == site_c]
                        realises_site_c += min(df_grp_c["_date_brute"].nunique() if not df_grp_c.empty else 0, attendu_c)
                    df_realises_site_c = df_realises_k[df_realises_k[col_site_k[0]].astype(str).str.strip() == site_c] if col_site_k else df_realises_k.iloc[0:0]
                    nb_visites_site_c = len(df_realises_site_c)
                    nb_respectes_site_c = int((df_realises_site_c["_ecart"] <= 31).sum()) if nb_visites_site_c > 0 and "_ecart" in df_realises_site_c.columns else 0
                    comparatif_sites[site_c] = {
                        "taux_realisation": round(realises_site_c/total_site_c*100,1) if total_site_c else 0,
                        "realises": realises_site_c, "total": total_site_c,
                        "taux_delai": round(nb_respectes_site_c/nb_visites_site_c*100,1) if nb_visites_site_c else 0,
                        "visites": nb_visites_site_c,
                    }



                kpi_data = {
                    "kpi1": {"taux":taux1, "realises":nb_realises_2026, "restants":nb_restants_2026, "total":nb_total_2026},
                    "kpi2": {"taux":taux2, "respectes":nb_respectes, "non_respectes":nb_non_respectes, "total":nb_visites_realisees}
                }




                k1c,k2c = st.columns(2)

                with k1c:
                    st.markdown("<p style='text-align:center;font-weight:600;color:#1E293B;font-size:14px;'>Taux de réalisation 2026</p></p>",unsafe_allow_html=True)
                    if nb_total_2026>0:
                        dfp1=pd.DataFrame({"Statut":["Réalisés","Restants"],"Nombre":[nb_realises_2026,nb_restants_2026]})
                        fig1=px.pie(dfp1,values="Nombre",names="Statut",hole=0.6,color="Statut",
                                    color_discrete_map={"Réalisés":"#10B981","Restants":"#EF4444"})
                        fig1.update_traces(textposition='inside',textinfo='percent')
                        fig1.update_layout(margin=dict(t=10,b=10,l=10,r=10),height=260,showlegend=False,
                                            paper_bgcolor='rgba(0,0,0,0)',plot_bgcolor='rgba(0,0,0,0)',
                                            transition_duration=500,transition_easing="cubic-in-out")
                        st.plotly_chart(fig1,use_container_width=True,config={'displayModeBar':False})
                        st.markdown(f"<p style='text-align:center;font-size:13px;color:#64748B;'>{taux1}% réalisés ({nb_realises_2026}/{nb_total_2026} contrôles d'installation)</p>",unsafe_allow_html=True)
                    else:
                        st.info("Aucune échéance d'installation en 2026.")

                with k2c:
                    st.markdown("<p style='text-align:center;font-weight:600;color:#1E293B;font-size:14px;'>Respect délai de visite (≤ 1 mois)</p>",unsafe_allow_html=True)
                    if nb_visites_realisees>0:
                        dfp2=pd.DataFrame({"Statut":["Respecté","Non respecté"],"Nombre":[nb_respectes,nb_non_respectes]})
                        fig2=px.pie(dfp2,values="Nombre",names="Statut",hole=0.6,color="Statut",
                                    color_discrete_map={"Respecté":"#0EA5E9","Non respecté":"#EF4444"})
                        fig2.update_traces(textposition='inside',textinfo='percent')
                        fig2.update_layout(margin=dict(t=10,b=10,l=10,r=10),height=260,showlegend=False,
                                            paper_bgcolor='rgba(0,0,0,0)',plot_bgcolor='rgba(0,0,0,0)',
                                            transition_duration=500,transition_easing="cubic-in-out")
                        st.plotly_chart(fig2,use_container_width=True,config={'displayModeBar':False})
                        st.markdown(f"<p style='text-align:center;font-size:13px;color:#64748B;'>{taux2}% respectés ({nb_respectes}/{nb_visites_realisees})</p>",unsafe_allow_html=True)
                    else:
                        st.info("Aucune visite réalisée à ce jour.")

            st.markdown("<br><hr style='border-color:#E2E8F0;'>",unsafe_allow_html=True)

            # ================= TAUX DE NON-CONFORMITÉ DE SITE (CARTOGRAPHIE) =================
            st.markdown("<p style='font-size:1.2rem;font-weight:700;color:#0F172A;'>🗺️ Taux de non-conformité de site</p>",unsafe_allow_html=True)

            carto_b64 = _charger_cartographie_b64()
            if carto_b64:
                vc1,vc2 = st.columns([5,1])
                with vc2:
                    st.markdown(f"<a href='{LUCID_CARTOGRAPHIE_URL}' target='_blank' style='display:inline-block;background:#1E3A8A;color:white;padding:8px 14px;border-radius:8px;text-decoration:none;font-weight:600;font-size:13px;text-align:center;'>🔗 Ouvrir dans Lucid</a>",unsafe_allow_html=True)
                components.html(f"""
                <div id="carto-viewer" style="position:relative;width:100%;height:620px;overflow:hidden;
                     background:#F8FAFC;border:1px solid #E2E8F0;border-radius:12px;">
                  <img id="carto-img" src="data:image/png;base64,{carto_b64}"
                       style="position:absolute;top:0;left:0;transform-origin:0 0;cursor:grab;user-select:none;max-width:none;visibility:hidden;"
                       draggable="false"/>
                  <div style="position:absolute;bottom:14px;right:14px;display:flex;gap:8px;z-index:10;">
                    <button id="carto-zoom-in" style="width:36px;height:36px;border-radius:8px;border:1px solid #CBD5E1;background:white;font-size:16px;font-weight:700;cursor:pointer;box-shadow:0 2px 6px rgba(0,0,0,0.08);">➕</button>
                    <button id="carto-zoom-out" style="width:36px;height:36px;border-radius:8px;border:1px solid #CBD5E1;background:white;font-size:16px;font-weight:700;cursor:pointer;box-shadow:0 2px 6px rgba(0,0,0,0.08);">➖</button>
                    <button id="carto-reset" style="width:36px;height:36px;border-radius:8px;border:1px solid #CBD5E1;background:white;font-size:15px;cursor:pointer;box-shadow:0 2px 6px rgba(0,0,0,0.08);">⟳</button>
                  </div>
                  <div style="position:absolute;top:10px;left:14px;font-size:11px;color:#64748B;background:rgba(255,255,255,0.85);
                       padding:4px 10px;border-radius:6px;">🖱️ Molette pour zoomer • Glisser pour déplacer</div>
                </div>
                <script>
                (function(){{
                    let baseScale=1, scale=1, posX=0, posY=0, isDragging=false, startX=0, startY=0;
                    const viewer=document.getElementById('carto-viewer');
                    const img=document.getElementById('carto-img');

                    function apply(){{ img.style.transform = 'translate('+posX+'px,'+posY+'px) scale('+scale+')'; }}

                    function fitToView(){{
                        const cw = viewer.clientWidth, ch = viewer.clientHeight;
                        const nw = img.naturalWidth, nh = img.naturalHeight;
                        if(!nw || !nh) return;
                        baseScale = Math.min(cw/nw, ch/nh);
                        scale = baseScale;
                        posX = (cw - nw*scale)/2;
                        posY = (ch - nh*scale)/2;
                        img.style.visibility = 'visible';
                        apply();
                    }}

                    function zoom(factor){{
                        scale*=factor;
                        scale=Math.max(baseScale*0.9, Math.min(scale, baseScale*8));
                        apply();
                    }}

                    if(img.complete && img.naturalWidth){{ fitToView(); }}
                    img.addEventListener('load', fitToView);

                    document.getElementById('carto-zoom-in').addEventListener('click', function(){{ zoom(1.25); }});
                    document.getElementById('carto-zoom-out').addEventListener('click', function(){{ zoom(0.8); }});
                    document.getElementById('carto-reset').addEventListener('click', fitToView);
                    img.addEventListener('wheel', function(e){{
                        e.preventDefault();
                        zoom(e.deltaY<0 ? 1.1 : 0.9);
                    }}, {{passive:false}});
                    img.addEventListener('mousedown', function(e){{ isDragging=true; startX=e.clientX-posX; startY=e.clientY-posY; img.style.cursor='grabbing'; }});
                    window.addEventListener('mouseup', function(){{ isDragging=false; img.style.cursor='grab'; }});
                    window.addEventListener('mousemove', function(e){{ if(!isDragging) return; posX=e.clientX-startX; posY=e.clientY-startY; apply(); }});
                }})();
                </script>
                """, height=630, scrolling=False)
            else:
                st.warning("⚠️ Fichier « Cartographie.png » introuvable. Placez-le dans le même dossier que l'application (à côté de app.py) pour l'afficher ici.")
                st.markdown(f"[🔗 Consulter la cartographie sur Lucid]({LUCID_CARTOGRAPHIE_URL})")

            st.markdown("<br><hr style='border-color:#E2E8F0;'>",unsafe_allow_html=True)

            # ================= Actions de contrôle =================
            st.markdown("<p style='font-size:1.2rem;font-weight:700;color:#0F172A;'>📌 Actions de contrôle</p>",unsafe_allow_html=True)

            with st.spinner("Chargement des actions..."):
                df_reserve = lire_points_reserve()

            with st.expander("➕ Ajouter une action"):
                r1,r2,r3,r4 = st.columns([1,1.5,1.5,1])
                with r1:
                    res_site = st.selectbox("Site",["SGB","MEG"],key="res_site_new")
                with r2:
                    res_ins = st.selectbox("Installation",list(PERIODICITE.keys()),key="res_ins_new")
                with r3:
                    res_seq = st.text_input("Sous-équipement",key="res_seq_new")
                with r4:
                    res_nb = st.number_input("Nb points",min_value=1,value=1,key="res_nb_new")
                if st.button("💾 Enregistrer",key="btn_add_reserve"):
                    if res_seq.strip():
                        ok,err = ajouter_point_reserve(res_site,res_ins,res_seq.strip(),res_nb)
                        if ok:
                            st.success("✅ Point de réserve ajouté !")
                            st.cache_data.clear()
                            st.rerun()
                        else:
                            st.error(f"Erreur : {err}")
                    else:
                        st.warning("Veuillez saisir un sous-équipement.")

            if df_reserve.empty:
                st.info("Aucun point de réserve enregistré. Utilisez le formulaire ci-dessus pour en ajouter.")
            else:
                if "Nombre" in df_reserve.columns:
                    df_reserve["Nombre"] = pd.to_numeric(df_reserve["Nombre"],errors="coerce").fillna(0).astype(int)

                with st.container(border=True):
                    st.markdown("<p style='font-weight:600;color:#1E293B;margin:0 0 10px 0;font-size:13px;'>🔍 Filtrer les Actions de contrôle</p>",unsafe_allow_html=True)
                    fr1,fr2,fr3 = st.columns(3)
                    sites_dispo = ["Tous"]+sorted(df_reserve["Site"].dropna().unique().tolist()) if "Site" in df_reserve.columns else ["Tous"]
                    inss_dispo  = ["Tous"]+sorted(df_reserve["Installation"].dropna().unique().tolist()) if "Installation" in df_reserve.columns else ["Tous"]
                    with fr1: f_res_site = st.selectbox("Site",sites_dispo,key="f_res_site")
                    with fr2: f_res_ins  = st.selectbox("Installation",inss_dispo,key="f_res_ins")
                    with fr3: f_res_seq  = st.text_input("Recherche sous-équipement",key="f_res_seq")

                df_reserve_f = df_reserve.copy()
                if f_res_site!="Tous" and "Site" in df_reserve_f.columns:
                    df_reserve_f = df_reserve_f[df_reserve_f["Site"]==f_res_site]
                if f_res_ins!="Tous" and "Installation" in df_reserve_f.columns:
                    df_reserve_f = df_reserve_f[df_reserve_f["Installation"]==f_res_ins]
                if f_res_seq.strip() and "Sous_equipement" in df_reserve_f.columns:
                    df_reserve_f = df_reserve_f[df_reserve_f["Sous_equipement"].astype(str).str.contains(f_res_seq.strip(),case=False,na=False)]

                st.dataframe(df_reserve_f.rename(columns={
                    "Site":"Site","Installation":"Installation","Sous_equipement":"Sous équipement","Nombre":"Nombre des actions de contrôle"
                }),hide_index=True,use_container_width=True)

                st.markdown("<br>",unsafe_allow_html=True)

                # --- Répartition par site : graphique centré ---
                csite1,csite2,csite3 = st.columns([1,2,1])
                with csite2:
                    if "Site" in df_reserve_f.columns and not df_reserve_f.empty:
                        df_by_site = df_reserve_f.groupby("Site")["Nombre"].sum().reset_index()
                        figS = px.pie(df_by_site,values="Nombre",names="Site",hole=0.6,
                                      color_discrete_sequence=['#1E3A8A','#0EA5E9','#94A3B8'])
                        figS.update_traces(textposition='inside',textinfo='percent+label')
                        figS.update_layout(title="Répartition par site",title_x=0.5,margin=dict(t=40,b=10,l=10,r=10),height=280,
                                            paper_bgcolor='rgba(0,0,0,0)',plot_bgcolor='rgba(0,0,0,0)')
                        st.plotly_chart(figS,use_container_width=True,config={'displayModeBar':False})
                    else:
                        st.info("Aucune donnée à afficher pour le graphe par site.")

                st.markdown("<br>",unsafe_allow_html=True)
                st.markdown("<p style='font-weight:700;font-size:14px;color:#0F172A;text-align:center;margin-bottom:10px;'>Répartition par installation</p>",unsafe_allow_html=True)

                # --- Répartition par installation : MEG (gauche) | légende (milieu) | SGB (droite) ---
                if "Installation" in df_reserve_f.columns and "Site" in df_reserve_f.columns and not df_reserve_f.empty:
                    all_inss = sorted(df_reserve_f["Installation"].dropna().unique().tolist())
                    palette = px.colors.qualitative.Set1
                    color_map = {ins: palette[i % len(palette)] for i,ins in enumerate(all_inss)}

                    gins1,gins2,gins3 = st.columns([2,1,2])

                    with gins1:
                        df_meg_ins = df_reserve_f[df_reserve_f["Site"]=="MEG"].groupby("Installation")["Nombre"].sum().reset_index()
                        if not df_meg_ins.empty:
                            figMEG = px.pie(df_meg_ins,values="Nombre",names="Installation",hole=0.6,
                                             color="Installation",color_discrete_map=color_map)
                            figMEG.update_traces(textposition='inside',textinfo='percent',showlegend=False)
                            figMEG.update_layout(title="MEG",title_x=0.5,showlegend=False,
                                                  margin=dict(t=40,b=10,l=10,r=10),height=260,
                                                  paper_bgcolor='rgba(0,0,0,0)',plot_bgcolor='rgba(0,0,0,0)')
                            st.plotly_chart(figMEG,use_container_width=True,config={'displayModeBar':False})
                        else:
                            st.info("Aucune donnée MEG.")

                    with gins2:
                        legende_items = "".join(
                            f"<div style='display:flex;align-items:center;gap:8px;margin-bottom:12px;'>"
                            f"<span style='width:12px;height:12px;min-width:12px;border-radius:3px;background:{color_map[ins]};display:inline-block;'></span>"
                            f"<span style='font-size:11.5px;color:#334155;'>{ins}</span>"
                            f"</div>"
                            for ins in all_inss
                        )
                        legende_html = f"<div style='padding-top:35px;'>{legende_items}</div>"
                        st.markdown(legende_html,unsafe_allow_html=True)

                    with gins3:
                        df_sgb_ins = df_reserve_f[df_reserve_f["Site"]=="SGB"].groupby("Installation")["Nombre"].sum().reset_index()
                        if not df_sgb_ins.empty:
                            figSGB = px.pie(df_sgb_ins,values="Nombre",names="Installation",hole=0.6,
                                             color="Installation",color_discrete_map=color_map)
                            figSGB.update_traces(textposition='inside',textinfo='percent',showlegend=False)
                            figSGB.update_layout(title="SGB",title_x=0.5,showlegend=False,
                                                  margin=dict(t=40,b=10,l=10,r=10),height=260,
                                                  paper_bgcolor='rgba(0,0,0,0)',plot_bgcolor='rgba(0,0,0,0)')
                            st.plotly_chart(figSGB,use_container_width=True,config={'displayModeBar':False})
                        else:
                            st.info("Aucune donnée SGB.")
                else:
                    st.info("Aucune donnée à afficher pour le graphe par installation.")

                with st.expander("🗑️ Supprimer une action"):
                    for orig_idx,row_r in df_reserve.iterrows():
                        dcx1,dcx2 = st.columns([5,1])
                        with dcx1:
                            st.write(f"{row_r.get('Site','')} — {row_r.get('Installation','')} — {row_r.get('Sous_equipement','')} — {row_r.get('Nombre',0)} pt(s)")
                        with dcx2:
                            if st.button("🗑️",key=f"del_res_{orig_idx}"):
                                num_ligne_sheet = orig_idx+2
                                if supprimer_ligne_generique("PointsReserve",num_ligne_sheet,4):
                                    st.success("Supprimé !")
                                    st.cache_data.clear()
                                    st.rerun()
                                else:
                                    st.error("Erreur lors de la suppression.")

            # ================= Actions de contrôle PAR NATURE =================
            st.markdown("<br><hr style='border-color:#E2E8F0;'>",unsafe_allow_html=True)
            st.markdown("<p style='font-size:1.2rem;font-weight:700;color:#0F172A;'>🧭 Actions de contrôle par nature</p>",unsafe_allow_html=True)

            with st.spinner("Chargement des actions par nature..."):
                df_nature = lire_points_reserve_nature()

            with st.expander("➕ Ajouter une ligne"):
                nt1,nt2,nt3,nt4 = st.columns([1,1.5,1,1.5])
                with nt1:
                    nat_site = st.selectbox("Site",["SGB","MEG"],key="nat_site_new")
                with nt2:
                    nat_ins = st.selectbox("Installation",list(PERIODICITE.keys()),key="nat_ins_new")
                with nt3:
                    nat_nb = st.number_input("Nb points",min_value=1,value=1,key="nat_nb_new")
                with nt4:
                    nat_code = st.selectbox("Nature",list(NATURE_PILOTE.keys()),
                                             format_func=lambda c: f"{c} — {NATURE_PILOTE[c][0]}",key="nat_code_new")
                nat_pilote_auto = NATURE_PILOTE[nat_code][1]
                st.markdown(
                    f"<p style='font-size:12.5px;color:#64748B;margin-top:-4px;'>Pilote assigné automatiquement : "
                    f"<b style='color:#1E3A8A;'>{nat_pilote_auto}</b></p>",unsafe_allow_html=True)
                if st.button("💾 Enregistrer",key="btn_add_nature"):
                    ok,err = ajouter_point_reserve_nature(nat_site,nat_ins,nat_nb,nat_code)
                    if ok:
                        st.success("✅ Ligne ajoutée !")
                        st.cache_data.clear()
                        st.rerun()
                    else:
                        st.error(f"Erreur : {err}")

            if df_nature.empty:
                st.info("Aucune donnée enregistrée. Utilisez le formulaire ci-dessus pour en ajouter.")
            else:
                if "Nombre" in df_nature.columns:
                    df_nature["Nombre"] = pd.to_numeric(df_nature["Nombre"],errors="coerce").fillna(0).astype(int)

                with st.container(border=True):
                    st.markdown("<p style='font-weight:600;color:#1E293B;margin:0 0 10px 0;font-size:13px;'>🔍 Filtrer</p>",unsafe_allow_html=True)
                    fn1,fn2 = st.columns(2)
                    sites_dispo_n = ["Tous"]+sorted(df_nature["Site"].dropna().unique().tolist()) if "Site" in df_nature.columns else ["Tous"]
                    inss_dispo_n  = ["Tous"]+sorted(df_nature["Installation"].dropna().unique().tolist()) if "Installation" in df_nature.columns else ["Tous"]
                    with fn1: f_nat_site = st.selectbox("Site",sites_dispo_n,key="f_nat_site")
                    with fn2: f_nat_ins  = st.selectbox("Installation",inss_dispo_n,key="f_nat_ins")

                df_nature_f = df_nature.copy()
                if f_nat_site!="Tous" and "Site" in df_nature_f.columns:
                    df_nature_f = df_nature_f[df_nature_f["Site"]==f_nat_site]
                if f_nat_ins!="Tous" and "Installation" in df_nature_f.columns:
                    df_nature_f = df_nature_f[df_nature_f["Installation"]==f_nat_ins]

                st.dataframe(df_nature_f.rename(columns={
                    "Site":"Site","Installation":"Installation","Nombre":"Nombre des actions","Nature":"Nature","Pilote":"Pilote"
                }),hide_index=True,use_container_width=True)

                st.markdown("<br>",unsafe_allow_html=True)
                st.markdown("<p style='font-weight:700;font-size:14px;color:#0F172A;text-align:center;margin-bottom:10px;'>Répartition par site : Nature et Pilote</p>",unsafe_allow_html=True)

                # --- Grille 2x2 : ligne 1 = SGB (Nature | Pilote), ligne 2 = MEG (Nature | Pilote) ---
                all_natures = [v[0] for v in NATURE_PILOTE.values()]
                palette_nat = px.colors.qualitative.Set2
                color_map_nat = {n: palette_nat[i % len(palette_nat)] for i,n in enumerate(all_natures)}

                # Les pilotes sont parfois combinés (ex: "BT + Maintenance", "BT + HSE + RH + DG").
                # On isole chaque entité (Maintenance, BT, HSE, RH, DG, DMTN, Chef service BT, ...)
                # pour calculer un % propre à chacune, même quand elle est partagée entre plusieurs natures.
                entites_atomiques = sorted(set(
                    e.strip() for v in NATURE_PILOTE.values() for e in v[1].split("+") if e.strip()
                ))
                palette_pil = px.colors.qualitative.Set1
                color_map_pil = {p: palette_pil[i % len(palette_pil)] for i,p in enumerate(entites_atomiques)}

                def _pie_nature_site(df_src, site, color_map, titre):
                    if "Nombre" not in df_src.columns or "Nature" not in df_src.columns or "Site" not in df_src.columns:
                        st.info(f"Aucune donnée {site}.")
                        return
                    d = df_src[df_src["Site"]==site].groupby("Nature")["Nombre"].sum().reset_index()
                    if d.empty:
                        st.info(f"Aucune donnée {site}.")
                        return
                    fig = px.pie(d,values="Nombre",names="Nature",hole=0.6,color="Nature",color_discrete_map=color_map)
                    fig.update_traces(textposition='inside',textinfo='percent')
                    fig.update_layout(title=titre,title_x=0.5,margin=dict(t=40,b=10,l=10,r=10),height=280,
                                       paper_bgcolor='rgba(0,0,0,0)',plot_bgcolor='rgba(0,0,0,0)',
                                       legend=dict(font=dict(size=9)))
                    st.plotly_chart(fig,use_container_width=True,config={'displayModeBar':False})

                def _bar_pilote_site(df_src, site, color_map, titre):
                    if "Nombre" not in df_src.columns or "Pilote" not in df_src.columns or "Site" not in df_src.columns:
                        st.info(f"Aucune donnée {site}.")
                        return
                    d = df_src[df_src["Site"]==site]
                    if d.empty:
                        st.info(f"Aucune donnée {site}.")
                        return
                    total = d["Nombre"].sum()
                    compte = {}
                    for _,row in d.iterrows():
                        for e in str(row["Pilote"]).split("+"):
                            e = e.strip()
                            if not e: continue
                            compte[e] = compte.get(e,0) + row["Nombre"]
                    if not compte or total==0:
                        st.info(f"Aucune donnée {site}.")
                        return
                    dd = pd.DataFrame({"Pilote":list(compte.keys()),"Nombre":list(compte.values())})
                    dd["Pourcentage"] = (dd["Nombre"]/total*100).round(1)
                    dd = dd.sort_values("Pourcentage",ascending=True)
                    fig = px.bar(dd,x="Pourcentage",y="Pilote",orientation="h",text="Pourcentage",
                                 color="Pilote",color_discrete_map=color_map)
                    fig.update_traces(texttemplate='%{text}%',textposition='outside',cliponaxis=False)
                    fig.update_layout(title=titre,title_x=0.5,showlegend=False,
                                       xaxis_title="% des actions",yaxis_title="",
                                       margin=dict(t=40,b=10,l=10,r=30),height=280,
                                       paper_bgcolor='rgba(0,0,0,0)',plot_bgcolor='rgba(0,0,0,0)')
                    st.plotly_chart(fig,use_container_width=True,config={'displayModeBar':False})

                if {"Nature","Pilote","Site","Nombre"}.issubset(df_nature_f.columns):
                    g1, g2 = st.columns(2)
                    with g1: _pie_nature_site(df_nature_f,"SGB",color_map_nat,"SGB — % par nature")
                    with g2: _bar_pilote_site(df_nature_f,"SGB",color_map_pil,"SGB — % par pilote")
                    g3,g4 = st.columns(2)
                    with g3: _pie_nature_site(df_nature_f,"MEG",color_map_nat,"MEG — % par nature")
                    with g4: _bar_pilote_site(df_nature_f,"MEG",color_map_pil,"MEG — % par pilote")
                else:
                    st.info("Aucune donnée à afficher pour les graphes.")

                with st.expander("🗑️ Supprimer une ligne"):
                    for orig_idx,row_n in df_nature.iterrows():
                        dnx1,dnx2 = st.columns([5,1])
                        with dnx1:
                            st.write(f"{row_n.get('Site','')} — {row_n.get('Installation','')} — {row_n.get('Nombre',0)} pt(s) — {row_n.get('Nature','')} — {row_n.get('Pilote','')}")
                        with dnx2:
                            if st.button("🗑️",key=f"del_nat_{orig_idx}"):
                                num_ligne_sheet = orig_idx+2
                                if supprimer_ligne_generique("PointsReserveNature",num_ligne_sheet,5):
                                    st.success("Supprimé !")
                                    st.cache_data.clear()
                                    st.rerun()
                                else:
                                    st.error("Erreur lors de la suppression.")

            # ================= RAPPORT PDF PAR PILOTE (CODIFICATION EXTERNE) =================
            st.markdown("<br><hr style='border-color:#E2E8F0;'>",unsafe_allow_html=True)
            st.markdown("<p style='font-size:1.2rem;font-weight:700;color:#0F172A;'>📄 Rapport des actions par Pilote</p>",unsafe_allow_html=True)

            entites_pilote_codif = sorted(set(
                e.strip() for v in NATURE_PILOTE.values() for e in v[1].split("+") if e.strip()
            ))
            cpil1,cpil2 = st.columns([3,1])
            with cpil1:
                pilote_codif_choisi = st.selectbox("Pilote",entites_pilote_codif,key="pilote_codif_select")
            with cpil2:
                st.write("")
                st.write("")
                charger_classeur_pilote = st.button("🔍 Charger les installations",use_container_width=True,key="btn_charger_classeur_pilote")

            if charger_classeur_pilote:
                with st.spinner("Lecture des classeurs de codification (MEG et SGB)..."):
                    df_codif_brut, err = codif_charger_toutes_actions()
                    if err and df_codif_brut.empty:
                        st.session_state["classeur_codif_pilote"] = None
                        st.error(err)
                    else:
                        if err:
                            st.warning(err)
                        st.session_state["classeur_codif_pilote"] = df_codif_brut
                        st.session_state["pdf_pilote"] = None

            classeur_pilote = st.session_state.get("classeur_codif_pilote")

            if classeur_pilote is None or (hasattr(classeur_pilote,"empty") and classeur_pilote.empty):
                st.info("👆 Cliquez sur « Charger les installations » pour sélectionner précisément "
                         "le site et l'installation à inclure dans le rapport.")
            else:
                df_codif = classeur_pilote.copy()
                # Exclut les actions déjà cochées comme réalisées par un responsable
                df_realisees_pilote = lire_actions_realisees()
                if not df_realisees_pilote.empty:
                    cles_faites = set(df_realisees_pilote.apply(_cle_action, axis=1))
                    df_codif = df_codif[~df_codif.apply(_cle_action, axis=1).isin(cles_faites)]

                df_codif["Nature"] = df_codif["Code"].map(lambda c: NATURE_PILOTE.get(c,("",""))[0])
                codes_ok = _codes_pour_pilote(pilote_codif_choisi)
                df_pilote_codif = df_codif[df_codif["Code"].isin(codes_ok)]

                if df_pilote_codif.empty:
                    st.info(f"Aucune action restante pour le pilote « {pilote_codif_choisi} » "
                            f"(codes recherchés : {', '.join(codes_ok) if codes_ok else '—'}).")
                else:
                        installations_dispo = sorted(df_pilote_codif["Installation"].unique().tolist())
                        sites_dispo = sorted(df_pilote_codif["Site"].dropna().unique().tolist())

                        cfil1,cfil2 = st.columns(2)
                        with cfil1:
                            if sites_dispo:
                                site_filtre_pilote = st.selectbox(
                                    "Site", ["Tous"]+sites_dispo, key="site_filtre_pilote_codif"
                                )
                            else:
                                site_filtre_pilote = "Tous"
                        installations_apres_site = [
                            i for i in installations_dispo
                            if site_filtre_pilote == "Tous" or i in
                               df_pilote_codif[df_pilote_codif["Site"]==site_filtre_pilote]["Installation"].unique()
                        ]
                        with cfil2:
                            installation_filtre_pilote = st.selectbox(
                                "Installation", ["Toutes"]+installations_apres_site,
                                key="installation_filtre_pilote_codif"
                            )

                        if installation_filtre_pilote == "Toutes":
                            df_filtre_codif = df_pilote_codif[df_pilote_codif["Installation"].isin(installations_apres_site)]
                        else:
                            df_filtre_codif = df_pilote_codif[df_pilote_codif["Installation"] == installation_filtre_pilote]


                        lancer_rapport_pilote = st.button(
                            "👁️ Générer",use_container_width=True,key="btn_gen_rapport_pilote",type="primary"
                        )

                        if lancer_rapport_pilote:
                            if df_filtre_codif.empty:
                                st.session_state["pdf_pilote"] = None
                                st.info(f"Aucune action trouvée pour le pilote « {pilote_codif_choisi} » "
                                        f"avec les filtres sélectionnés.")
                            else:
                                try:
                                    st.session_state["pdf_pilote"] = generer_rapport_pilote_pdf(
                                        pilote_codif_choisi, df_filtre_codif,
                                        "https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcR6q1BtDSDgVnJZFo0hOBfQJoDS6OYiub-qfQ&s"
                                    )
                                    st.session_state["nb_actions_pilote"] = len(df_filtre_codif)
                                    st.session_state["pilote_pdf_nom"] = pilote_codif_choisi
                                except Exception as e:
                                    st.session_state["pdf_pilote"] = None
                                    st.error(f"Erreur lors de la génération du PDF : {e}")

            if st.session_state.get("pdf_pilote"):
                st.success(f"{st.session_state.get('nb_actions_pilote',0)} action(s) trouvée(s) pour "
                           f"« {st.session_state.get('pilote_pdf_nom','')} ».")
                afficher_apercu_pdf_grille(st.session_state["pdf_pilote"], colonnes=2)
                st.download_button(
                    label="📥 Télécharger le rapport PDF",
                    data=st.session_state["pdf_pilote"],
                    file_name=f"Rapport_{st.session_state.get('pilote_pdf_nom','pilote')}_{datetime.date.today().strftime('%d_%m_%Y')}.pdf",
                    mime="application/pdf",
                    use_container_width=True,
                    key="dl_pilote"
                )

            # ================= RAPPORT PDF PREMIUM =================
            st.markdown("<br><hr style='border-color:#E2E8F0;'>",unsafe_allow_html=True)
            st.markdown("<p style='font-size:1.2rem;font-weight:700;color:#0F172A;'>📄 Rapport PDF </p>",unsafe_allow_html=True)
           
            if kpi_data is None:
                st.info("Le rapport PDF nécessite des données KPI disponibles (onglet « Rapports » non vide).")
            else:
                st.markdown(
                    "<div style='background:white;padding:16px 20px;border-radius:10px;"
                    "box-shadow:0 2px 8px rgba(0,0,0,0.05);border-left:4px solid #1E3A8A;"
                    "margin-bottom:10px;'>"
                    "<span style='font-size:14px;font-weight:600;color:#1E293B;'>📑 Rapport PDF — Synthèse KPI</span>"
                    "</div>", unsafe_allow_html=True)
                if st.button("👁️ Consulter le rapport", use_container_width=True, key="consult_kpi", type="primary"):
                    with st.spinner("Préparation du rapport PDF..."):
                        try:
                            st.session_state["pdf_kpi"] = generer_rapport_kpi_pdf(
                                kpi_data,
                                df_reserve,
                                df_nature,
                                carto_b64,
                                "https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcR6q1BtDSDgVnJZFo0hOBfQJoDS6OYiub-qfQ&s"
                            )
                        except Exception as e:
                            st.session_state["pdf_kpi"] = None
                            st.error(f"Erreur lors de la génération du PDF : {e}")
                if st.session_state.get("pdf_kpi"):
                    afficher_apercu_pdf_grille(st.session_state["pdf_kpi"], colonnes=2)
                    st.download_button(
                        label="📥 Télécharger le rapport PDF",
                        data=st.session_state["pdf_kpi"],
                        file_name=f"Rapport_KPI_{datetime.date.today().strftime('%d_%m_%Y')}.pdf",
                        mime="application/pdf",
                        use_container_width=True,
                        key="dl_kpi"
                    )
            # ---- Export global (toutes les données, tous sites/années, en un seul fichier) ----
            st.markdown("<br><hr style='border-color:#E2E8F0;'>",unsafe_allow_html=True)
            st.markdown("<br>",unsafe_allow_html=True)
            st.markdown("### 📦 Export global")
            st.caption("Génère un classeur Excel regroupant tous les rapports.")
            if st.button("📥 Générer le classeur", use_container_width=True):
                with st.spinner("Génération de l'export global..."):
                    excel_global = generer_export_global_excel()
                st.download_button("⬇️ Télécharger le classeur", data=excel_global,
                    file_name=f"Export_Global_{datetime.date.today().strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True)
                journaliser_action(utilisateur_courant(), "Export global généré")




    # ---- ONGLET 4 : KPI (Responsable — vue filtrée sur ses propres actions/rapport) ----
    if tab_kpi and role=="Responsable" and st.session_state.responsable_connecte:
        with tab_kpi:
            compte_resp = RESPONSABLES.get(st.session_state.responsable_actif, {})
            nom_resp = compte_resp.get("nom", st.session_state.responsable_actif)
            entites_resp = compte_resp.get("entites", [])
            site_resp = compte_resp.get("site")

            st.markdown(f"<p style='font-size:1.2rem;font-weight:700;color:#1E3A8A;'>📊 Indicateurs — {nom_resp}</p>",unsafe_allow_html=True)
            st.markdown(
                "<div style='background:#EFF6FF;border-left:4px solid #2a78d6;padding:10px 14px;border-radius:6px;margin-bottom:14px;'>"
                "<p style='margin:0;font-size:12px;color:#1e40af;font-weight:600;'>Consultation seule : Veuillez sélectionner votre site</p>"
                "</div>", unsafe_allow_html=True)

            with st.spinner("Chargement des actions par nature..."):
                df_nature_r = lire_points_reserve_nature()

            if df_nature_r.empty:
                st.info("Aucune donnée disponible pour le moment.")
            else:
                if "Nombre" in df_nature_r.columns:
                    df_nature_r["Nombre"] = pd.to_numeric(df_nature_r["Nombre"],errors="coerce").fillna(0).astype(int)

                # Ne garder que les lignes concernant les entités du responsable connecté
                def _concerne_responsable(pilote_str):
                    entites = [e.strip() for e in str(pilote_str).split("+") if e.strip()]
                    return any(e in entites_resp for e in entites)

                if "Pilote" in df_nature_r.columns:
                    df_nature_r = df_nature_r[df_nature_r["Pilote"].apply(_concerne_responsable)]

                if df_nature_r.empty:
                    st.info("Aucune action trouvée pour votre périmètre.")
                else:
                    if site_resp:
                        # Compte restreint à un seul site (ex : Saber -> MEG, Chafik -> SGB)
                        st.session_state.site_kpi_responsable = site_resp
                        st.markdown(f"<p style='font-size:12px;color:#64748B;margin-bottom:6px;'>🏭 Site : <strong>{site_resp}</strong></p>",unsafe_allow_html=True)
                    else:
                        if "site_kpi_responsable" not in st.session_state:
                            st.session_state.site_kpi_responsable = "SGB"

                        rcol1,rcol2,_ = st.columns([1,1,4])
                        with rcol1:
                            if st.button("🏭 SGB", key="btn_kpi_sgb_resp", type=("primary" if st.session_state.site_kpi_responsable=="SGB" else "secondary"), use_container_width=True):
                                st.session_state.site_kpi_responsable = "SGB"
                                st.rerun()
                        with rcol2:
                            if st.button("🏭 MEG", key="btn_kpi_meg_resp", type=("primary" if st.session_state.site_kpi_responsable=="MEG" else "secondary"), use_container_width=True):
                                st.session_state.site_kpi_responsable = "MEG"
                                st.rerun()

                    site_choisi_r = st.session_state.site_kpi_responsable
                    st.markdown(f"<p style='font-weight:700;font-size:14px;color:#0F172A;text-align:center;margin:14px 0 10px 0;'>Répartition des actions — Site {site_choisi_r}</p>",unsafe_allow_html=True)

                    all_natures_r = [v[0] for v in NATURE_PILOTE.values()]
                    palette_nat_r = px.colors.qualitative.Set2
                    color_map_nat_r = {n: palette_nat_r[i % len(palette_nat_r)] for i,n in enumerate(all_natures_r)}

                    if {"Nature","Site","Nombre"}.issubset(df_nature_r.columns):
                        d_site_r = df_nature_r[df_nature_r["Site"]==site_choisi_r]
                        if d_site_r.empty:
                            st.info(f"Aucune donnée {site_choisi_r} pour votre périmètre.")
                        else:
                            total_r = int(d_site_r["Nombre"].sum())
                            st.markdown(f"""<div style="background:white;padding:14px;border-radius:10px;box-shadow:0 2px 6px rgba(0,0,0,0.05);border-left:4px solid #1E3A8A;margin-bottom:14px;">
                                <p style="margin:0;font-size:11px;color:#64748B;font-weight:600;text-transform:uppercase;">Total actions — {site_choisi_r}</p>
                                <p style="margin:4px 0 0 0;font-size:28px;color:#0F172A;font-weight:700;">{total_r}</p></div>""",unsafe_allow_html=True)

                            dv_r = d_site_r.groupby("Nature")["Nombre"].sum().reset_index()

                            gcol1,gcol2 = st.columns(2)
                            with gcol1:
                                fig_r = px.pie(dv_r,values="Nombre",names="Nature",hole=0.6,color="Nature",color_discrete_map=color_map_nat_r)
                                fig_r.update_traces(textposition='inside',textinfo='percent+value')
                                fig_r.update_layout(title=f"{site_choisi_r} — répartition par nature",title_x=0.5,margin=dict(t=40,b=10,l=10,r=10),height=320,
                                                     paper_bgcolor='rgba(0,0,0,0)',plot_bgcolor='rgba(0,0,0,0)',legend=dict(font=dict(size=9)))
                                st.plotly_chart(fig_r,use_container_width=True,config={'displayModeBar':False})

                            with gcol2:
                                if "Installation" in d_site_r.columns:
                                    di_r = d_site_r.groupby("Installation")["Nombre"].sum().reset_index().sort_values("Nombre",ascending=True)
                                    palette_ins_r = px.colors.qualitative.Set3
                                    color_map_ins_r = {i: palette_ins_r[k % len(palette_ins_r)] for k,i in enumerate(di_r["Installation"])}
                                    fig_ins_r = px.bar(di_r,x="Nombre",y="Installation",orientation="h",text="Nombre",
                                                        color="Installation",color_discrete_map=color_map_ins_r)
                                    fig_ins_r.update_traces(textposition='outside',cliponaxis=False)
                                    fig_ins_r.update_layout(title=f"{site_choisi_r} — répartition par installation",title_x=0.5,showlegend=False,
                                                             xaxis_title="Nombre d'actions",yaxis_title="",
                                                             margin=dict(t=40,b=10,l=10,r=30),height=320,
                                                             paper_bgcolor='rgba(0,0,0,0)',plot_bgcolor='rgba(0,0,0,0)')
                                    st.plotly_chart(fig_ins_r,use_container_width=True,config={'displayModeBar':False})
                                else:
                                    st.info("Aucune donnée d'installation à afficher.")

                            st.dataframe(d_site_r.rename(columns={
                                "Site":"Site","Installation":"Installation","Nombre":"Nombre des actions","Nature":"Nature","Pilote":"Pilote"
                            }),hide_index=True,use_container_width=True)
                    else:
                        st.info("Aucune donnée à afficher pour les graphes.")

            # ================= RAPPORT PDF ASSOCIÉ AU RESPONSABLE =================
            st.markdown("<br><hr style='border-color:#E2E8F0;'>",unsafe_allow_html=True)
            st.markdown("<p style='font-size:1.2rem;font-weight:700;color:#0F172A;'>📄 Mon rapport des actions</p>",unsafe_allow_html=True)

            if len(entites_resp) > 1:
                entite_pdf_choisie = st.selectbox("Périmètre", entites_resp, key="entite_pdf_responsable")
            else:
                entite_pdf_choisie = entites_resp[0] if entites_resp else None
                

            if st.button("Générer mon rapport", use_container_width=True, key="btn_gen_rapport_responsable", type="primary") and entite_pdf_choisie:
                with st.spinner("Lecture des classeurs de codification (MEG et SGB)..."):
                    df_codif_r, err = codif_charger_toutes_actions()
                    if err and df_codif_r.empty:
                        st.session_state["pdf_responsable"] = None
                        st.error(err)
                    else:
                        if err:
                            st.warning(err)
                        # Exclut les actions déjà cochées comme réalisées (onglet « Suivi des actions »)
                        df_realisees_r = lire_actions_realisees()
                        if not df_realisees_r.empty:
                            cles_faites_r = set(df_realisees_r.apply(_cle_action, axis=1))
                            df_codif_r = df_codif_r[~df_codif_r.apply(_cle_action, axis=1).isin(cles_faites_r)]

                        df_codif_r["Nature"] = df_codif_r["Code"].map(lambda c: NATURE_PILOTE.get(c,("",""))[0])
                        codes_ok_r = _codes_pour_pilote(entite_pdf_choisie)
                        df_filtre_codif_r = df_codif_r[df_codif_r["Code"].isin(codes_ok_r)]
                        if site_resp and "Site" in df_filtre_codif_r.columns:
                            df_filtre_codif_r = df_filtre_codif_r[df_filtre_codif_r["Site"].astype(str).str.strip().str.upper() == site_resp.upper()]
                        if df_filtre_codif_r.empty:
                            st.session_state["pdf_responsable"] = None
                            st.info(f"Aucune action restante pour « {entite_pdf_choisie} »" + (f" — site {site_resp}." if site_resp else "."))
                        else:
                            try:
                                st.session_state["pdf_responsable"] = generer_rapport_pilote_pdf(
                                    entite_pdf_choisie, df_filtre_codif_r,
                                    "https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcR6q1BtDSDgVnJZFo0hOBfQJoDS6OYiub-qfQ&s"
                                )
                                st.session_state["nb_actions_responsable"] = len(df_filtre_codif_r)
                            except Exception as e:
                                st.session_state["pdf_responsable"] = None
                                st.error(f"Erreur lors de la génération du PDF : {e}")

            if st.session_state.get("pdf_responsable"):
                st.success(f"{st.session_state.get('nb_actions_responsable',0)} action(s) trouvée(s) pour « {entite_pdf_choisie} ».")
                afficher_apercu_pdf_grille(st.session_state["pdf_responsable"], colonnes=2)
                st.download_button(
                    label="📥 Télécharger mon rapport PDF",
                    data=st.session_state["pdf_responsable"],
                    file_name=f"Rapport_{entite_pdf_choisie}_{datetime.date.today().strftime('%d_%m_%Y')}.pdf",
                    mime="application/pdf",
                    use_container_width=True,
                    key="dl_responsable"
                )

    # ---- ONGLET 4 : KPI (Visiteur — vue simplifiée en lecture seule) ----
    if tab_kpi and role=="Visiteur":
        with tab_kpi:
            st.markdown("<p style='font-size:1.2rem;font-weight:700;color:#1E3A8A;'>📊 Actions de contrôle par nature</p>",unsafe_allow_html=True)

            with st.spinner("Chargement des actions par nature..."):
                df_nature_v = lire_points_reserve_nature()

            if df_nature_v.empty:
                st.info("Aucune donnée disponible pour le moment.")
            else:
                if "Nombre" in df_nature_v.columns:
                    df_nature_v["Nombre"] = pd.to_numeric(df_nature_v["Nombre"],errors="coerce").fillna(0).astype(int)

                if "site_kpi_visiteur" not in st.session_state:
                    st.session_state.site_kpi_visiteur = "SGB"

                bcol1,bcol2,_ = st.columns([1,1,4])
                with bcol1:
                    if st.button("🏭 SGB", key="btn_kpi_sgb", type=("primary" if st.session_state.site_kpi_visiteur=="SGB" else "secondary"), use_container_width=True):
                        st.session_state.site_kpi_visiteur = "SGB"
                        st.rerun()
                with bcol2:
                    if st.button("🏭 MEG", key="btn_kpi_meg", type=("primary" if st.session_state.site_kpi_visiteur=="MEG" else "secondary"), use_container_width=True):
                        st.session_state.site_kpi_visiteur = "MEG"
                        st.rerun()

                site_choisi = st.session_state.site_kpi_visiteur
                st.markdown(f"<p style='font-weight:700;font-size:14px;color:#0F172A;text-align:center;margin:14px 0 10px 0;'>Répartition — Site {site_choisi}</p>",unsafe_allow_html=True)

                all_natures_v = [v[0] for v in NATURE_PILOTE.values()]
                palette_nat_v = px.colors.qualitative.Set2
                color_map_nat_v = {n: palette_nat_v[i % len(palette_nat_v)] for i,n in enumerate(all_natures_v)}

                entites_atomiques_v = sorted(set(
                    e.strip() for v in NATURE_PILOTE.values() for e in v[1].split("+") if e.strip()
                ))
                palette_pil_v = px.colors.qualitative.Set1
                color_map_pil_v = {p: palette_pil_v[i % len(palette_pil_v)] for i,p in enumerate(entites_atomiques_v)}

                if {"Nature","Pilote","Site"}.issubset(df_nature_v.columns):
                    vg1,vg2 = st.columns(2)
                    with vg1:
                        dv = df_nature_v[df_nature_v["Site"]==site_choisi].groupby("Nature")["Nombre"].sum().reset_index()
                        if dv.empty:
                            st.info(f"Aucune donnée {site_choisi}.")
                        else:
                            figv1 = px.pie(dv,values="Nombre",names="Nature",hole=0.6,color="Nature",color_discrete_map=color_map_nat_v)
                            figv1.update_traces(textposition='inside',textinfo='percent')
                            figv1.update_layout(title=f"{site_choisi} — % par nature",title_x=0.5,margin=dict(t=40,b=10,l=10,r=10),height=300,
                                                 paper_bgcolor='rgba(0,0,0,0)',plot_bgcolor='rgba(0,0,0,0)',legend=dict(font=dict(size=9)))
                            st.plotly_chart(figv1,use_container_width=True,config={'displayModeBar':False})
                    with vg2:
                        dsite = df_nature_v[df_nature_v["Site"]==site_choisi]
                        total_v = dsite["Nombre"].sum()
                        compte_v = {}
                        for _,row in dsite.iterrows():
                            for e in str(row["Pilote"]).split("+"):
                                e = e.strip()
                                if not e: continue
                                compte_v[e] = compte_v.get(e,0) + row["Nombre"]
                        if not compte_v or total_v==0:
                            st.info(f"Aucune donnée {site_choisi}.")
                        else:
                            ddv = pd.DataFrame({"Pilote":list(compte_v.keys()),"Nombre":list(compte_v.values())})
                            ddv["Pourcentage"] = (ddv["Nombre"]/total_v*100).round(1)
                            ddv = ddv.sort_values("Pourcentage",ascending=True)
                            figv2 = px.bar(ddv,x="Pourcentage",y="Pilote",orientation="h",text="Pourcentage",
                                           color="Pilote",color_discrete_map=color_map_pil_v)
                            figv2.update_traces(texttemplate='%{text}%',textposition='outside',cliponaxis=False)
                            figv2.update_layout(title=f"{site_choisi} — % par pilote",title_x=0.5,showlegend=False,
                                                 xaxis_title="% des actions",yaxis_title="",
                                                 margin=dict(t=40,b=10,l=10,r=30),height=300,
                                                 paper_bgcolor='rgba(0,0,0,0)',plot_bgcolor='rgba(0,0,0,0)')
                            st.plotly_chart(figv2,use_container_width=True,config={'displayModeBar':False})
                else:
                    st.info("Aucune donnée à afficher pour les graphes.")

    # ---- ONGLET 5 : SUIVI DES ACTIONS (cocher les actions terminées, historique persistant) ----
    if tab_suivi:
        with tab_suivi:
            st.markdown("<p style='font-size:1.2rem;font-weight:700;color:#1E3A8A;'>✅ Suivi des actions</p>",unsafe_allow_html=True)

            est_admin_suivi = (role == "Admin" and password_correct)

            if not est_admin_suivi:
                st.markdown(
                    "<div style='background:#EFF6FF;border-left:4px solid #2a78d6;padding:10px 14px;border-radius:6px;margin-bottom:14px;'>"
                    "<p style='margin:0;font-size:12px;color:#1e40af;'>Cochez les actions terminées.</p>"
                    "</div>", unsafe_allow_html=True)

            if est_admin_suivi:
                entites_disponibles_suivi = sorted(set(
                    e.strip() for v in NATURE_PILOTE.values() for e in v[1].split("+") if e.strip()
                ))
                pilote_suivi_choisi = st.selectbox("Responsable à suivre", entites_disponibles_suivi, key="pilote_suivi_admin")
                nom_responsable_suivi = nom_pour_pilote_site(pilote_suivi_choisi, None)
                site_resp_suivi = None
            else:
                compte_resp_suivi = RESPONSABLES.get(st.session_state.responsable_actif, {})
                entites_resp_suivi = compte_resp_suivi.get("entites", [])
                nom_responsable_suivi = compte_resp_suivi.get("nom", st.session_state.responsable_actif)
                site_resp_suivi = compte_resp_suivi.get("site")
                if len(entites_resp_suivi) > 1:
                    pilote_suivi_choisi = st.selectbox("Périmètre", entites_resp_suivi, key="pilote_suivi_resp")
                else:
                    pilote_suivi_choisi = entites_resp_suivi[0] if entites_resp_suivi else None

            if not pilote_suivi_choisi:
                st.info("Aucun périmètre associé à ce profil.")
            else:
                with st.spinner("Chargement des actions (MEG et SGB)..."):
                    df_codif_suivi, err_suivi = codif_charger_toutes_actions()
                    df_realisees_suivi = lire_actions_realisees()
                    df_encours_suivi = lire_suivi_encours()

                if err_suivi and df_codif_suivi.empty:
                    st.error(err_suivi)
                else:
                    if err_suivi:
                        st.warning(err_suivi)

                    codes_ok_suivi = _codes_pour_pilote(pilote_suivi_choisi)
                    df_pilote_suivi = df_codif_suivi[df_codif_suivi["Code"].isin(codes_ok_suivi)].copy()
                    if site_resp_suivi and "Site" in df_pilote_suivi.columns:
                        df_pilote_suivi = df_pilote_suivi[df_pilote_suivi["Site"].astype(str).str.strip().str.upper() == site_resp_suivi.upper()]
                    df_pilote_suivi["Cle"] = df_pilote_suivi.apply(_cle_action, axis=1)

                    cles_faites_suivi = set()
                    df_hist_pilote = pd.DataFrame()
                    if not df_realisees_suivi.empty and "Pilote" in df_realisees_suivi.columns:
                        df_hist_pilote = df_realisees_suivi[df_realisees_suivi["Pilote"] == pilote_suivi_choisi]
                        cles_faites_suivi = set(df_hist_pilote.apply(_cle_action, axis=1))

                    df_restantes = df_pilote_suivi[~df_pilote_suivi["Cle"].isin(cles_faites_suivi)]

                    total_pilote = len(df_pilote_suivi)
                    nb_realisees = len(cles_faites_suivi & set(df_pilote_suivi["Cle"]))
                    taux = round((nb_realisees/total_pilote*100), 1) if total_pilote else 0.0

                    # Dernier statut « En cours » connu (Type + commentaire), par clé d'action
                    info_encours_par_cle = {}
                    if not df_encours_suivi.empty:
                        for _, r_enc in df_encours_suivi.iterrows():
                            info_encours_par_cle[r_enc["Cle"]] = {
                                "Type": r_enc.get("Type", "") or "",
                                "Commentaire": r_enc.get("Commentaire", "") or "",
                            }

                    col_liste, col_graphe = st.columns([3, 1])

                    with col_liste:
                        if df_restantes.empty:
                            st.markdown(f"<p style='font-weight:700;font-size:14px;color:#0F172A;'>Actions restantes — {nom_responsable_suivi}</p>",unsafe_allow_html=True)
                            st.success("🎉 Toutes les actions de ce périmètre sont réalisées !")
                        else:
                            site_f_suivi = "Tous"
                            if site_resp_suivi:
                                # Compte restreint à un seul site : pas de sélecteur de site
                                st.markdown(f"<p style='font-weight:700;font-size:14px;color:#0F172A;'>Actions restantes — {nom_responsable_suivi}</p>",unsafe_allow_html=True)
                                df_apres_site_suivi = df_restantes
                                installations_dispo_suivi = sorted(df_apres_site_suivi["Installation"].dropna().unique().tolist())
                                install_f_suivi = st.selectbox("Installation", ["Toutes"]+installations_dispo_suivi, key="installation_filtre_suivi")
                            else:
                                sites_dispo_suivi = sorted(df_restantes["Site"].dropna().unique().tolist())
                                fcol1, fcol2 = st.columns(2)
                                with fcol1:
                                    site_f_suivi = st.selectbox("Site", ["Tous"]+sites_dispo_suivi, key="site_filtre_suivi")
                                if est_admin_suivi:
                                    nom_responsable_suivi = nom_pour_pilote_site(pilote_suivi_choisi, None if site_f_suivi == "Tous" else site_f_suivi)
                                st.markdown(f"<p style='font-weight:700;font-size:14px;color:#0F172A;'>Actions restantes — {nom_responsable_suivi}</p>",unsafe_allow_html=True)
                                df_apres_site_suivi = df_restantes if site_f_suivi == "Tous" else df_restantes[df_restantes["Site"] == site_f_suivi]
                                with fcol2:
                                    installations_dispo_suivi = sorted(df_apres_site_suivi["Installation"].dropna().unique().tolist())
                                    install_f_suivi = st.selectbox("Installation", ["Toutes"]+installations_dispo_suivi, key="installation_filtre_suivi")
                            df_affiche_suivi = df_apres_site_suivi if install_f_suivi == "Toutes" else df_apres_site_suivi[df_apres_site_suivi["Installation"] == install_f_suivi]

                            # ---- Recalcul des statistiques (jauge + graphes) selon le site sélectionné ----
                            site_effectif_stats = site_resp_suivi if site_resp_suivi else (None if site_f_suivi == "Tous" else site_f_suivi)
                            if site_effectif_stats and "Site" in df_pilote_suivi.columns:
                                df_pilote_suivi_stats = df_pilote_suivi[df_pilote_suivi["Site"].astype(str).str.strip().str.upper() == site_effectif_stats.upper()]
                                df_restantes = df_restantes[df_restantes["Site"].astype(str).str.strip().str.upper() == site_effectif_stats.upper()]
                            else:
                                df_pilote_suivi_stats = df_pilote_suivi
                            total_pilote = len(df_pilote_suivi_stats)
                            nb_realisees = len(cles_faites_suivi & set(df_pilote_suivi_stats["Cle"]))
                            taux = round((nb_realisees/total_pilote*100), 1) if total_pilote else 0.0

                            if not est_admin_suivi:
                                # ---- Vue responsable : un tableau, avec colonnes Statut / Type de suivi / Commentaire ----
                                TYPES_SUIVI = ["Immédiat", "Sous-traitance", "Planifié"]

                                df_edit = df_affiche_suivi[["Site", "Installation", "Désignation", "Observation", "Code"]].copy()
                                df_edit.insert(0, "Statut", "En cours")
                                df_edit["Type de suivi"] = df_affiche_suivi["Cle"].map(
                                    lambda c: info_encours_par_cle.get(c, {}).get("Type") or None
                                )
                                df_edit["Commentaire"] = df_affiche_suivi["Cle"].map(
                                    lambda c: info_encours_par_cle.get(c, {}).get("Commentaire", "")
                                )

                                df_edit_out = st.data_editor(
                                    df_edit,
                                    hide_index=True,
                                    use_container_width=True,
                                    disabled=["Site", "Installation", "Désignation", "Observation", "Code"],
                                    column_config={
                                        "Statut": st.column_config.SelectboxColumn("Statut", options=["En cours", "Terminé"], required=True),
                                        "Type de suivi": st.column_config.SelectboxColumn("Type de suivi (si en cours)", options=TYPES_SUIVI, required=False),
                                        "Commentaire": st.column_config.TextColumn("Commentaire (optionnel)"),
                                    },
                                    key="editeur_suivi_actions"
                                )

                                nb_termine_saisi = int((df_edit_out["Statut"] == "Terminé").sum()) if not df_edit_out.empty else 0
                                if st.button(f"💾 Enregistrer les modifications ({nb_termine_saisi} terminée(s))", type="primary",
                                             use_container_width=True, key="btn_valider_suivi"):
                                    erreur_maj = False
                                    nb_termine_ok, nb_encours_ok = 0, 0
                                    for idx, row_edit in df_edit_out.iterrows():
                                        row_orig = df_affiche_suivi.loc[idx].copy()
                                        row_orig["Pilote"] = pilote_suivi_choisi
                                        statut_c = row_edit["Statut"]
                                        type_c = row_edit.get("Type de suivi")
                                        commentaire_c = row_edit.get("Commentaire", "")
                                        if statut_c == "Terminé":
                                            ok = marquer_actions_realisees(pd.DataFrame([row_orig]), nom_responsable_suivi)
                                            erreur_maj = erreur_maj or not ok
                                            nb_termine_ok += 1 if ok else 0
                                        else:
                                            info_prealable = info_encours_par_cle.get(row_orig["Cle"], {})
                                            type_defaut = info_prealable.get("Type", "")
                                            commentaire_defaut = info_prealable.get("Commentaire", "")
                                            if (type_c or "") == (type_defaut or "") and (commentaire_c or "") == (commentaire_defaut or ""):
                                                continue  # rien n'a changé pour cette action, inutile de ré-écrire
                                            ok = enregistrer_statut_en_cours(row_orig, type_c or TYPES_SUIVI[0], commentaire_c or "", nom_responsable_suivi)
                                            erreur_maj = erreur_maj or not ok
                                            nb_encours_ok += 1 if ok else 0

                                    if erreur_maj:
                                        st.error("Une erreur est survenue lors de l'enregistrement (vérifiez les onglets « ActionsRealisees » / « SuiviActions »).")
                                    else:
                                        st.success(f"{nb_termine_ok} action(s) marquée(s) terminée(s), {nb_encours_ok} mise(s) à jour « en cours ».")
                                        st.cache_data.clear()
                                        st.rerun()
                            else:
                                # ---- Vue admin : tableau récapitulatif en lecture seule ----
                                df_recap = df_affiche_suivi[["Site", "Installation", "Désignation", "Observation", "Code"]].copy()
                                df_recap["Type de suivi"] = df_affiche_suivi["Cle"].map(
                                    lambda c: info_encours_par_cle.get(c, {}).get("Type") or "Non défini"
                                )
                                df_recap["Commentaire"] = df_affiche_suivi["Cle"].map(
                                    lambda c: info_encours_par_cle.get(c, {}).get("Commentaire", "")
                                )
                                st.dataframe(df_recap, hide_index=True, use_container_width=True)
                                st.caption("ℹ️ Vue administrateur en lecture seule.")

                        with st.expander(f"🗂️ Historique des actions réalisées ({len(df_hist_pilote)})"):
                            if df_hist_pilote.empty:
                                st.info("Aucune action réalisée pour le moment.")
                            else:
                                cols_hist = [c for c in ["DateRealisation", "Site", "Installation", "Désignation", "Observation", "Code", "Responsable"] if c in df_hist_pilote.columns]
                                df_aff_hist = df_hist_pilote[cols_hist]
                                if "DateRealisation" in df_aff_hist.columns:
                                    df_aff_hist = df_aff_hist.sort_values("DateRealisation", ascending=False)
                                st.dataframe(df_aff_hist, hide_index=True, use_container_width=True)

                    with col_graphe:
                        st.markdown("<p style='font-weight:700;font-size:13px;color:#0F172A;text-align:center;'>Taux de réalisation</p>",unsafe_allow_html=True)
                        fig_gauge = go.Figure(go.Indicator(
                            mode="gauge+number",
                            value=taux,
                            number={"suffix": "%"},
                            gauge={
                                "axis": {"range": [0, 100]},
                                "bar": {"color": "#1baf7a"},
                                "steps": [
                                    {"range": [0, 50], "color": "#FEE2E2"},
                                    {"range": [50, 80], "color": "#FEF3C7"},
                                    {"range": [80, 100], "color": "#DCFCE7"},
                                ],
                            }
                        ))
                        fig_gauge.update_layout(height=220, margin=dict(t=20, b=10, l=20, r=20), paper_bgcolor='rgba(0,0,0,0)',
                                                 transition_duration=600, transition_easing="cubic-in-out")
                        st.plotly_chart(fig_gauge, use_container_width=True, config={'displayModeBar': False})
                        st.markdown(f"<p style='text-align:center;font-size:12px;color:#64748B;'>{nb_realisees} / {total_pilote} action(s) réalisée(s)</p>",unsafe_allow_html=True)

                    # ---- Graphes complémentaires (Admin uniquement), affichés sous le tableau ----
                    if est_admin_suivi:
                        st.markdown("<br><hr style='border-color:#E2E8F0;'>",unsafe_allow_html=True)
                        st.markdown(f"<p style='font-size:1.05rem;font-weight:700;color:#1E3A8A;'>📈 Analyse du suivi — {nom_responsable_suivi}</p>",unsafe_allow_html=True)

                        gA, gB = st.columns(2)

                        # --- Graphe 1 : % Terminé vs En cours (pour le pilote sélectionné) ---
                        with gA:
                            en_cours_total = max(total_pilote - nb_realisees, 0)
                            if total_pilote:
                                df_g1 = pd.DataFrame({"Statut": ["Terminé", "En cours"], "Nombre": [nb_realisees, en_cours_total]})
                                fig1 = px.pie(df_g1, values="Nombre", names="Statut", hole=0.6, color="Statut",
                                              color_discrete_map={"Terminé": "#16A34A", "En cours": "#F97316"})
                                fig1.update_traces(textposition='inside', textinfo='percent+value')
                                fig1.update_layout(title="Terminé vs En cours", title_x=0.5,
                                                    margin=dict(t=40,b=10,l=10,r=10), height=300,
                                                    paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)',
                                                    legend=dict(font=dict(size=9)),
                                                    transition_duration=500, transition_easing="cubic-in-out")
                                st.plotly_chart(fig1, use_container_width=True, config={'displayModeBar': False})
                            else:
                                st.info("Aucune donnée.")

                        # --- Graphe 2 : % Immédiat / Sous-traitance / Planifié parmi les actions en cours ---
                        with gB:
                            cles_en_cours_g = set(df_restantes["Cle"]) if "Cle" in df_restantes.columns else set()
                            compte_type_g = {"Immédiat": 0, "Sous-traitance": 0, "Planifié": 0, "Non défini": 0}
                            for cle_g in cles_en_cours_g:
                                t_g = info_encours_par_cle.get(cle_g, {}).get("Type") or "Non défini"
                                if t_g not in compte_type_g:
                                    t_g = "Non défini"
                                compte_type_g[t_g] += 1
                            df_g2 = pd.DataFrame({"Type": list(compte_type_g.keys()), "Nombre": list(compte_type_g.values())})
                            df_g2 = df_g2[df_g2["Nombre"] > 0]
                            if not df_g2.empty:
                                fig2 = px.pie(df_g2, values="Nombre", names="Type", hole=0.6, color="Type",
                                              color_discrete_map={"Immédiat": "#2563EB", "Sous-traitance": "#F59E0B",
                                                                   "Planifié": "#8B5CF6", "Non défini": "#94A3B8"})
                                fig2.update_traces(textposition='inside', textinfo='percent+value')
                                fig2.update_layout(title="Répartition des actions en cours", title_x=0.5,
                                                    margin=dict(t=40,b=10,l=10,r=10), height=300,
                                                    paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)',
                                                    legend=dict(font=dict(size=9)),
                                                    transition_duration=500, transition_easing="cubic-in-out")
                                st.plotly_chart(fig2, use_container_width=True, config={'displayModeBar': False})
                            else:
                                st.info("Aucune action en cours.")
